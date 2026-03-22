"""
compare_boq.py — Compare system-generated BOQ Excel against approved BOQ.
Usage: python compare_boq.py [output_boq.xlsx]
"""
import sys
import os
sys.path.insert(0, os.path.dirname(__file__))

import openpyxl
from pathlib import Path
from src.config import DATA_DIR, OUTPUT_BOQ
from src.utils import safe_float

# Find the most recent output BOQ
if len(sys.argv) > 1:
    sys_path = Path(sys.argv[1])
else:
    boq_files = sorted(OUTPUT_BOQ.glob("G303_BOQ_*.xlsx"), key=lambda p: p.stat().st_mtime)
    if not boq_files:
        print("No G303_BOQ_*.xlsx found in output/boq/")
        sys.exit(1)
    sys_path = boq_files[-1]

ap_path = DATA_DIR / "approved_boq.xlsx"

print(f"System BOQ : {sys_path.name}")
print(f"Approved   : {ap_path.name}")
print()

# Load approved BOQ
ap_wb = openpyxl.load_workbook(str(ap_path), data_only=True, read_only=True)
ap_ws = ap_wb["G303-BOQ"]

# Load system BOQ (the BOQ sheet should be "BOQ" or "G303-BOQ")
sys_wb = openpyxl.load_workbook(str(sys_path), data_only=True, read_only=True)
sys_sheet = "BOQ" if "BOQ" in sys_wb.sheetnames else sys_wb.sheetnames[0]
sys_ws = sys_wb[sys_sheet]

DATA_START = 9
TOLERANCE = 0.05   # 5% tolerance

ok = 0
diff = 0
skip = 0
diff_items = []

ap_rows = list(ap_ws.iter_rows(min_row=DATA_START, values_only=True))
sys_rows = list(sys_ws.iter_rows(min_row=DATA_START, values_only=True))

results = []
for row_i, (ap_row, sys_row) in enumerate(zip(ap_rows, sys_rows), DATA_START):
    if len(ap_row) < 3 or len(sys_row) < 3:
        continue

    ap_sc   = str(ap_row[0]).strip() if ap_row[0] is not None else ""
    ap_desc = str(ap_row[1]).strip() if ap_row[1] is not None else ""
    ap_qty  = ap_row[2]
    sys_qty = sys_row[2]   # col C = calculated qty in system output

    # Col E in approved = notes (if any); col E in system = RATE
    # Skip notes/headers
    if ap_qty is None and sys_qty is None:
        # Header row
        conf = ""
        note = ""
        results.append((row_i, "HEAD", ap_sc, ap_desc, ap_qty, sys_qty, conf, note))
        continue

    ap_v  = safe_float(ap_qty)
    sys_v = safe_float(sys_qty)

    # Get confidence and notes from system BOQ (cols G, I)
    conf = str(sys_row[6]).strip() if len(sys_row) > 6 and sys_row[6] is not None else ""
    note = str(sys_row[8]).strip() if len(sys_row) > 8 and sys_row[8] is not None else ""

    if ap_v is None or sys_v is None:
        skip += 1
        status = "SKIP"
    elif ap_v == 0 and sys_v == 0:
        ok += 1
        status = "OK  "
    elif ap_v == 0:
        skip += 1
        status = "SKIP"
    else:
        pct = abs(sys_v - ap_v) / max(abs(ap_v), 0.01)
        if pct <= TOLERANCE:
            ok += 1
            status = "OK  "
        else:
            diff += 1
            status = "DIFF"

    results.append((row_i, status, ap_sc, ap_desc, ap_qty, sys_qty, conf, note))

# Print output
out_lines = []
for (row_i, status, sc, desc, ap_qty, sys_qty, conf, note) in results:
    if status == "HEAD":
        line = f"[{status}] SYS={sys_qty!s:<10} AP={ap_qty!s:<10} {conf:<6} {sc}  {desc}"
    else:
        line = f"[{status}] SYS={sys_qty!s:<10} AP={ap_qty!s:<10} {conf:<6} {desc}"
    out_lines.append(line)
    if note and status in ("DIFF", "HEAD"):
        out_lines.append(f"         NOTE: {note[:100]}")

print("\n" + "="*80)
print("\n".join(out_lines))
print("="*80)
print(f"Compared: {ok+diff}  OK: {ok}  DIFF: {diff}")

# Write to file
out_file = Path(__file__).parent / "output" / "comparison_latest.txt"
out_file.parent.mkdir(parents=True, exist_ok=True)
with open(str(out_file), "w", encoding="utf-8") as f:
    for line in out_lines:
        f.write(line + "\n")
    f.write("="*80 + "\n")
    f.write(f"Compared: {ok+diff}  OK: {ok}  DIFF: {diff}\n")

print(f"\nSaved to: {out_file.name}")

ap_wb.close()
sys_wb.close()
