"""
loader.py — Loads standard_model_G303.xlsx into memory.

Reads three sheets:
  1. Standard BOQ — all items with stock codes, descriptions, units, standard qty, rate
  2. Standard Geometry — dimensions of standard G303
  3. Rules Library — scaling rules for dependent items

Returns structured dicts/lists for downstream pipeline steps.
"""

import os
from typing import Any
import openpyxl


def load_standard_model(filepath: str) -> dict[str, Any]:
    """Load the standard model Excel workbook and return parsed data.

    Args:
        filepath: Path to standard_model_G303.xlsx

    Returns:
        dict with keys: 'standard_boq', 'standard_geometry', 'rules'

    Raises:
        FileNotFoundError: If the file does not exist.
        ValueError: If required sheets are missing.
    """
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"Standard model file not found: {filepath}")

    wb = openpyxl.load_workbook(filepath, data_only=True)

    required_sheets = ["Standard BOQ", "Standard Geometry", "Rules Library"]
    missing = [s for s in required_sheets if s not in wb.sheetnames]
    if missing:
        raise ValueError(f"Missing required sheets in standard model: {missing}")

    standard_boq = _load_standard_boq(wb["Standard BOQ"])
    standard_geometry = _load_standard_geometry(wb["Standard Geometry"])
    rules = _load_rules_library(wb["Rules Library"])

    wb.close()

    return {
        "standard_boq": standard_boq,
        "standard_geometry": standard_geometry,
        "rules": rules,
    }


def _load_standard_boq(ws) -> list[dict]:
    """Parse the Standard BOQ sheet into a list of item dicts."""
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not rows:
        return []

    headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    items = []

    for row in rows[1:]:
        if not row or all(cell is None for cell in row):
            continue
        item = {}
        for i, header in enumerate(headers):
            if i < len(row):
                item[header] = row[i]
        # Normalise key names for consistent downstream use
        item = _normalise_boq_keys(item)
        if item.get("description"):
            items.append(item)

    return items


def _normalise_boq_keys(item: dict) -> dict:
    """Map various possible column names to canonical keys."""
    key_map = {
        "item no": "item_no",
        "item_no": "item_no",
        "item number": "item_no",
        "no": "item_no",
        "stock code": "stock_code",
        "stock_code": "stock_code",
        "code": "stock_code",
        "description": "description",
        "desc": "description",
        "unit": "unit",
        "uom": "unit",
        "qty": "qty",
        "quantity": "qty",
        "standard qty": "qty",
        "standard_qty": "qty",
        "rate": "rate",
        "rate (pgk)": "rate",
        "rate_pgk": "rate",
        "amount": "amount",
        "amount (pgk)": "amount",
        "category": "category",
        "section": "category",
        "notes": "notes",
    }
    normalised = {}
    for k, v in item.items():
        canonical = key_map.get(str(k).strip().lower(), str(k).strip().lower())
        normalised[canonical] = v
    return normalised


def _load_standard_geometry(ws) -> dict:
    """Parse the Standard Geometry sheet into a dict of element → value.

    Expected format: two columns — Element, Value
    e.g. 'total_floor_area', 120.5
    """
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not rows:
        return {}

    headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    geometry = {}

    # Detect format: key-value pair (2 cols) or tabular
    if len(headers) >= 2 and "element" in headers[0]:
        for row in rows[1:]:
            if row and row[0] is not None:
                key = str(row[0]).strip().lower().replace(" ", "_")
                value = row[1] if len(row) > 1 else None
                try:
                    value = float(value) if value is not None else 0.0
                except (ValueError, TypeError):
                    pass
                geometry[key] = value
    else:
        # Tabular format: headers are element names, single data row
        for i, header in enumerate(headers):
            if header and i < len(rows[1]) if len(rows) > 1 else False:
                key = header.replace(" ", "_")
                value = rows[1][i]
                try:
                    value = float(value) if value is not None else 0.0
                except (ValueError, TypeError):
                    pass
                geometry[key] = value

    return geometry


def _load_rules_library(ws) -> list[dict]:
    """Parse the Rules Library sheet into a list of rule dicts.

    Expected columns: Rule ID, Target Item, Depends On, Formula, Description
    """
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not rows:
        return []

    headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    rules = []

    for row in rows[1:]:
        if not row or all(cell is None for cell in row):
            continue
        rule = {}
        for i, header in enumerate(headers):
            if i < len(row):
                canonical = header.replace(" ", "_")
                rule[canonical] = row[i]
        if rule.get("rule_id") or rule.get("target_item"):
            rules.append(rule)

    return rules


def load_rate_library(filepath: str) -> list[dict]:
    """Load the rate library Excel file.

    Args:
        filepath: Path to rate_library.xlsx

    Returns:
        List of rate dicts with keys: stock_code, description, unit, rate, category, notes
    """
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"Rate library file not found: {filepath}")

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not rows:
        wb.close()
        return []

    headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    rates = []

    key_map = {
        "stock code": "stock_code",
        "stock_code": "stock_code",
        "code": "stock_code",
        "description": "description",
        "desc": "description",
        "unit": "unit",
        "uom": "unit",
        "rate": "rate",
        "rate (pgk)": "rate",
        "rate_pgk": "rate",
        "category": "category",
        "notes": "notes",
    }

    for row in rows[1:]:
        if not row or all(cell is None for cell in row):
            continue
        item = {}
        for i, header in enumerate(headers):
            if i < len(row):
                canonical = key_map.get(header, header.replace(" ", "_"))
                item[canonical] = row[i]
        if item.get("description"):
            rates.append(item)

    wb.close()
    return rates
