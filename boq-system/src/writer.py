"""
writer.py — Writes final BOQ Excel with all sheets.

Sheet 1: Main BOQ — full priced BOQ with confidence and source
Sheet 2: Change Log — element-by-element delta comparison
Sheet 3: Variance Report — bulk items comparison
Sheet 4: Assumptions — every rule, AI estimate, fallback documented
Sheet 5: Items for Review — LOW confidence + QA flags + AI-estimated rates
Sheet 6: QA Summary — pass/fail for each QA check

RULE: Excel formulas not hardcoded values.
  Amount column = =Qty*Rate formula
  Totals = =SUM() formulas
"""

from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter


# Color coding per confidence level
CONFIDENCE_FILLS = {
    "HIGH": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),    # Green
    "MEDIUM": PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),   # Blue
    "LOW": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),      # Orange
    "REVIEW": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),   # Red
}

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
NORMAL_FONT = Font(name="Calibri", size=10)
BOLD_FONT = Font(name="Calibri", bold=True, size=10)
TITLE_FONT = Font(name="Calibri", bold=True, size=14)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

QA_STATUS_FILLS = {
    "PASS": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "WARN": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    "FAIL": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
}


def write_output(
    project_name: str,
    project_boq: list[dict],
    change_log: list[dict],
    standard_boq: list[dict],
    qa_flags: list[dict],
    output_dir: str = "output",
) -> str:
    """Write the complete BOQ Excel output.

    Args:
        project_name: Name of the project (used in filename).
        project_boq: Final priced project BOQ items.
        change_log: Change log from change_detector.
        standard_boq: Original standard BOQ for variance report.
        qa_flags: QA check results.
        output_dir: Directory to save the output file.

    Returns:
        Path to the generated Excel file.
    """
    import os
    os.makedirs(output_dir, exist_ok=True)

    wb = Workbook()
    date_str = datetime.now().strftime("%Y%m%d")
    filename = f"{project_name}_BOQ_{date_str}.xlsx"
    filepath = os.path.join(output_dir, filename)

    # Sheet 1: Main BOQ
    _write_main_boq(wb.active, project_boq, project_name)

    # Sheet 2: Change Log
    ws2 = wb.create_sheet("Change Log")
    _write_change_log(ws2, change_log)

    # Sheet 3: Variance Report
    ws3 = wb.create_sheet("Variance Report")
    _write_variance_report(ws3, project_boq, standard_boq)

    # Sheet 4: Assumptions
    ws4 = wb.create_sheet("Assumptions")
    _write_assumptions(ws4, project_boq, change_log)

    # Sheet 5: Items for Review
    ws5 = wb.create_sheet("Items for Review")
    _write_review_items(ws5, project_boq, qa_flags)

    # Sheet 6: QA Summary
    ws6 = wb.create_sheet("QA Summary")
    _write_qa_summary(ws6, qa_flags)

    wb.save(filepath)
    return filepath


# --- Sheet 1: Main BOQ ---

def _write_main_boq(ws, boq: list[dict], project_name: str) -> None:
    ws.title = "Main BOQ"

    # Title row
    ws.merge_cells("A1:K1")
    ws["A1"] = f"Bill of Quantities — {project_name}"
    ws["A1"].font = TITLE_FONT

    ws.merge_cells("A2:K2")
    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = NORMAL_FONT

    # Headers
    headers = [
        "Item No", "Stock Code", "Description", "Unit", "Qty",
        "Rate (PGK)", "Amount (PGK)", "Confidence", "Source",
        "Rate Source", "Notes",
    ]
    col_widths = [10, 14, 45, 8, 12, 14, 16, 12, 14, 14, 40]

    header_row = 4
    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col)].width = width

    # Data rows
    data_start = header_row + 1
    for i, item in enumerate(boq):
        row = data_start + i
        ws.cell(row=row, column=1, value=item.get("item_no", i + 1)).border = THIN_BORDER
        ws.cell(row=row, column=2, value=item.get("stock_code", "")).border = THIN_BORDER
        ws.cell(row=row, column=3, value=item.get("description", "")).border = THIN_BORDER
        ws.cell(row=row, column=4, value=item.get("unit", "")).border = THIN_BORDER

        qty_cell = ws.cell(row=row, column=5, value=item.get("qty", 0))
        qty_cell.number_format = '#,##0.00'
        qty_cell.border = THIN_BORDER

        rate_cell = ws.cell(row=row, column=6, value=item.get("rate", 0) or 0)
        rate_cell.number_format = '#,##0.00'
        rate_cell.border = THIN_BORDER

        # Amount = Qty * Rate (FORMULA, not hardcoded)
        amount_cell = ws.cell(
            row=row, column=7,
            value=f"=E{row}*F{row}",
        )
        amount_cell.number_format = '#,##0.00'
        amount_cell.border = THIN_BORDER

        ws.cell(row=row, column=8, value=item.get("confidence", "")).border = THIN_BORDER
        ws.cell(row=row, column=9, value=item.get("source", "")).border = THIN_BORDER
        ws.cell(row=row, column=10, value=item.get("rate_source", "")).border = THIN_BORDER
        ws.cell(row=row, column=11, value=item.get("notes", "")).border = THIN_BORDER

        # Apply confidence color to entire row
        confidence = str(item.get("confidence", "")).upper()
        fill = CONFIDENCE_FILLS.get(confidence)
        if fill:
            for col in range(1, 12):
                ws.cell(row=row, column=col).fill = fill

        # Font
        for col in range(1, 12):
            ws.cell(row=row, column=col).font = NORMAL_FONT

    # Total row
    total_row = data_start + len(boq)
    ws.cell(row=total_row, column=6, value="TOTAL:").font = BOLD_FONT
    ws.cell(row=total_row, column=6).alignment = Alignment(horizontal="right")
    total_cell = ws.cell(
        row=total_row, column=7,
        value=f"=SUM(G{data_start}:G{total_row - 1})",
    )
    total_cell.font = BOLD_FONT
    total_cell.number_format = '#,##0.00'
    total_cell.border = THIN_BORDER

    # Legend
    legend_row = total_row + 2
    ws.cell(row=legend_row, column=1, value="Legend:").font = BOLD_FONT
    for i, (level, fill) in enumerate(CONFIDENCE_FILLS.items()):
        r = legend_row + 1 + i
        ws.cell(row=r, column=1, value=level).fill = fill
        ws.cell(row=r, column=1).font = NORMAL_FONT
        labels = {
            "HIGH": "Unchanged from standard or BOM-verified",
            "MEDIUM": "Recalculated from geometry or rules",
            "LOW": "AI estimate or assumed",
            "REVIEW": "QA flag — must check before use",
        }
        ws.cell(row=r, column=2, value=labels.get(level, "")).font = NORMAL_FONT

    # Freeze panes
    ws.freeze_panes = f"A{data_start}"

    # Auto-filter
    ws.auto_filter.ref = f"A{header_row}:K{total_row - 1}"


# --- Sheet 2: Change Log ---

def _write_change_log(ws, change_log: list[dict]) -> None:
    headers = ["Element", "Standard Value", "Project Value", "Unit", "Delta", "Delta %", "Changed", "Impact"]
    col_widths = [25, 16, 16, 8, 14, 12, 10, 50]

    ws.merge_cells("A1:H1")
    ws["A1"] = "Change Log — Standard vs Project Geometry"
    ws["A1"].font = TITLE_FONT

    header_row = 3
    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col)].width = width

    for i, entry in enumerate(change_log):
        row = header_row + 1 + i
        ws.cell(row=row, column=1, value=entry.get("element", "")).border = THIN_BORDER
        ws.cell(row=row, column=2, value=entry.get("standard_value", 0)).border = THIN_BORDER
        ws.cell(row=row, column=3, value=entry.get("project_value", 0)).border = THIN_BORDER
        ws.cell(row=row, column=4, value=entry.get("unit", "")).border = THIN_BORDER

        delta_cell = ws.cell(row=row, column=5, value=entry.get("delta", 0))
        delta_cell.number_format = '#,##0.00'
        delta_cell.border = THIN_BORDER

        pct_cell = ws.cell(row=row, column=6, value=entry.get("delta_pct", 0))
        pct_cell.number_format = '0.0"%"'
        pct_cell.border = THIN_BORDER

        changed = "YES" if entry.get("changed") else "NO"
        changed_cell = ws.cell(row=row, column=7, value=changed)
        changed_cell.border = THIN_BORDER
        if entry.get("changed"):
            changed_cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

        ws.cell(row=row, column=8, value=entry.get("impact", "")).border = THIN_BORDER

    ws.freeze_panes = f"A{header_row + 1}"


# --- Sheet 3: Variance Report ---

def _write_variance_report(ws, project_boq: list[dict], standard_boq: list[dict]) -> None:
    ws.merge_cells("A1:G1")
    ws["A1"] = "Variance Report — Standard vs Project Quantities"
    ws["A1"].font = TITLE_FONT

    headers = ["Item No", "Description", "Unit", "Standard Qty", "Project Qty", "Variance", "Variance %"]
    col_widths = [10, 45, 8, 14, 14, 14, 14]

    header_row = 3
    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col)].width = width

    # Build standard qty lookup
    std_map = {}
    for item in standard_boq:
        key = item.get("stock_code") or item.get("item_no") or item.get("description", "")
        std_map[key] = item

    row_num = header_row + 1
    for item in project_boq:
        key = item.get("stock_code") or item.get("item_no") or item.get("description", "")
        std_item = std_map.get(key, {})

        std_qty = float(std_item.get("qty", 0) or 0)
        proj_qty = float(item.get("qty", 0) or 0)
        variance = proj_qty - std_qty
        variance_pct = (variance / std_qty * 100) if std_qty != 0 else (100 if variance != 0 else 0)

        if abs(variance) < 0.01:
            continue  # Only show items with variance

        ws.cell(row=row_num, column=1, value=item.get("item_no", "")).border = THIN_BORDER
        ws.cell(row=row_num, column=2, value=item.get("description", "")).border = THIN_BORDER
        ws.cell(row=row_num, column=3, value=item.get("unit", "")).border = THIN_BORDER

        ws.cell(row=row_num, column=4, value=std_qty).border = THIN_BORDER
        ws.cell(row=row_num, column=4).number_format = '#,##0.00'

        ws.cell(row=row_num, column=5, value=proj_qty).border = THIN_BORDER
        ws.cell(row=row_num, column=5).number_format = '#,##0.00'

        # Variance as formula
        ws.cell(row=row_num, column=6, value=f"=E{row_num}-D{row_num}").border = THIN_BORDER
        ws.cell(row=row_num, column=6).number_format = '#,##0.00'

        # Variance % as formula
        ws.cell(
            row=row_num, column=7,
            value=f'=IF(D{row_num}=0,"NEW",F{row_num}/D{row_num}*100)',
        ).border = THIN_BORDER
        ws.cell(row=row_num, column=7).number_format = '0.0'

        # Color code large variances
        if abs(variance_pct) > 50:
            for c in range(1, 8):
                ws.cell(row=row_num, column=c).fill = PatternFill(
                    start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"
                )

        row_num += 1

    ws.freeze_panes = f"A{header_row + 1}"


# --- Sheet 4: Assumptions ---

def _write_assumptions(ws, project_boq: list[dict], change_log: list[dict]) -> None:
    ws.merge_cells("A1:D1")
    ws["A1"] = "Assumptions & Calculation Basis"
    ws["A1"].font = TITLE_FONT

    headers = ["Category", "Item", "Assumption / Rule Applied", "Source"]
    col_widths = [18, 35, 55, 18]

    header_row = 3
    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col)].width = width

    row_num = header_row + 1

    # Document each recalculated item
    for item in project_boq:
        source = item.get("source", "")
        if source in ("recalculated", "rule-scaled", "new", "removed"):
            ws.cell(row=row_num, column=1, value=source.upper()).border = THIN_BORDER
            ws.cell(row=row_num, column=2, value=item.get("description", "")).border = THIN_BORDER
            ws.cell(row=row_num, column=3, value=item.get("notes", "")).border = THIN_BORDER
            ws.cell(row=row_num, column=4, value=item.get("confidence", "")).border = THIN_BORDER
            row_num += 1

    # Document AI-estimated rates
    for item in project_boq:
        if item.get("rate_source") == "AI-estimate":
            ws.cell(row=row_num, column=1, value="AI RATE").border = THIN_BORDER
            ws.cell(row=row_num, column=2, value=item.get("description", "")).border = THIN_BORDER
            ws.cell(
                row=row_num, column=3,
                value=f"Rate {item.get('rate', 0)} PGK estimated by GPT-4o. Verify before use.",
            ).border = THIN_BORDER
            ws.cell(row=row_num, column=4, value="AI-estimate").border = THIN_BORDER
            row_num += 1

    # Document geometry changes
    for entry in change_log:
        if entry.get("changed"):
            ws.cell(row=row_num, column=1, value="GEOMETRY").border = THIN_BORDER
            ws.cell(row=row_num, column=2, value=entry.get("element", "")).border = THIN_BORDER
            ws.cell(
                row=row_num, column=3,
                value=f"Changed from {entry['standard_value']} to {entry['project_value']} "
                      f"({entry['delta_pct']:+.1f}%). {entry.get('impact', '')}",
            ).border = THIN_BORDER
            ws.cell(row=row_num, column=4, value="DXF comparison").border = THIN_BORDER
            row_num += 1


# --- Sheet 5: Items for Review ---

def _write_review_items(ws, project_boq: list[dict], qa_flags: list[dict]) -> None:
    ws.merge_cells("A1:F1")
    ws["A1"] = "Items Requiring Manual Review"
    ws["A1"].font = TITLE_FONT

    headers = ["Item No", "Description", "Reason for Review", "Current Qty", "Current Rate", "Confidence"]
    col_widths = [10, 40, 45, 14, 14, 12]

    header_row = 3
    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col)].width = width

    row_num = header_row + 1

    # LOW confidence items
    for item in project_boq:
        confidence = str(item.get("confidence", "")).upper()
        if confidence in ("LOW", "REVIEW"):
            ws.cell(row=row_num, column=1, value=item.get("item_no", "")).border = THIN_BORDER
            ws.cell(row=row_num, column=2, value=item.get("description", "")).border = THIN_BORDER
            ws.cell(row=row_num, column=3, value=f"Low confidence: {item.get('notes', '')}").border = THIN_BORDER
            ws.cell(row=row_num, column=4, value=item.get("qty", 0)).border = THIN_BORDER
            ws.cell(row=row_num, column=5, value=item.get("rate", 0) or "MISSING").border = THIN_BORDER
            ws.cell(row=row_num, column=6, value=confidence).border = THIN_BORDER

            fill = CONFIDENCE_FILLS.get(confidence)
            if fill:
                for c in range(1, 7):
                    ws.cell(row=row_num, column=c).fill = fill
            row_num += 1

    # AI-estimated rates
    for item in project_boq:
        if item.get("rate_source") == "AI-estimate":
            ws.cell(row=row_num, column=1, value=item.get("item_no", "")).border = THIN_BORDER
            ws.cell(row=row_num, column=2, value=item.get("description", "")).border = THIN_BORDER
            ws.cell(row=row_num, column=3, value="Rate is AI-estimated — verify against supplier quotes").border = THIN_BORDER
            ws.cell(row=row_num, column=4, value=item.get("qty", 0)).border = THIN_BORDER
            ws.cell(row=row_num, column=5, value=item.get("rate", 0)).border = THIN_BORDER
            ws.cell(row=row_num, column=6, value=item.get("confidence", "")).border = THIN_BORDER
            row_num += 1

    # Manual-required rates
    for item in project_boq:
        if item.get("rate_source") == "manual-required":
            ws.cell(row=row_num, column=1, value=item.get("item_no", "")).border = THIN_BORDER
            ws.cell(row=row_num, column=2, value=item.get("description", "")).border = THIN_BORDER
            ws.cell(row=row_num, column=3, value="No rate found — manual entry required").border = THIN_BORDER
            ws.cell(row=row_num, column=4, value=item.get("qty", 0)).border = THIN_BORDER
            ws.cell(row=row_num, column=5, value="MISSING").border = THIN_BORDER
            ws.cell(row=row_num, column=6, value=item.get("confidence", "")).border = THIN_BORDER

            for c in range(1, 7):
                ws.cell(row=row_num, column=c).fill = CONFIDENCE_FILLS["REVIEW"]
            row_num += 1

    # QA flagged items
    for flag in qa_flags:
        if flag.get("status") in ("FAIL", "WARN") and flag.get("items"):
            for flagged_item in flag["items"]:
                ws.cell(row=row_num, column=1, value="QA").border = THIN_BORDER
                ws.cell(row=row_num, column=2, value=str(flagged_item)).border = THIN_BORDER
                ws.cell(row=row_num, column=3, value=f"QA {flag['check']}: {flag['details']}").border = THIN_BORDER
                ws.cell(row=row_num, column=4, value="").border = THIN_BORDER
                ws.cell(row=row_num, column=5, value="").border = THIN_BORDER
                ws.cell(row=row_num, column=6, value="REVIEW").border = THIN_BORDER

                for c in range(1, 7):
                    ws.cell(row=row_num, column=c).fill = QA_STATUS_FILLS.get(flag["status"], QA_STATUS_FILLS["WARN"])
                row_num += 1

    ws.freeze_panes = f"A{header_row + 1}"


# --- Sheet 6: QA Summary ---

def _write_qa_summary(ws, qa_flags: list[dict]) -> None:
    ws.merge_cells("A1:D1")
    ws["A1"] = "Quality Assurance Summary"
    ws["A1"].font = TITLE_FONT

    headers = ["Check", "Status", "Details", "Flagged Items"]
    col_widths = [30, 10, 60, 50]

    header_row = 3
    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col)].width = width

    for i, flag in enumerate(qa_flags):
        row = header_row + 1 + i
        ws.cell(row=row, column=1, value=flag.get("check", "")).border = THIN_BORDER
        status_cell = ws.cell(row=row, column=2, value=flag.get("status", ""))
        status_cell.border = THIN_BORDER

        fill = QA_STATUS_FILLS.get(flag.get("status", ""))
        if fill:
            status_cell.fill = fill

        ws.cell(row=row, column=3, value=flag.get("details", "")).border = THIN_BORDER

        items_str = "; ".join(str(item) for item in flag.get("items", []))
        ws.cell(row=row, column=4, value=items_str).border = THIN_BORDER

    # Summary counts
    summary_row = header_row + 1 + len(qa_flags) + 2
    pass_count = sum(1 for f in qa_flags if f.get("status") == "PASS")
    warn_count = sum(1 for f in qa_flags if f.get("status") == "WARN")
    fail_count = sum(1 for f in qa_flags if f.get("status") == "FAIL")

    ws.cell(row=summary_row, column=1, value="Summary:").font = BOLD_FONT
    ws.cell(row=summary_row + 1, column=1, value=f"PASS: {pass_count}").fill = QA_STATUS_FILLS["PASS"]
    ws.cell(row=summary_row + 2, column=1, value=f"WARN: {warn_count}").fill = QA_STATUS_FILLS["WARN"]
    ws.cell(row=summary_row + 3, column=1, value=f"FAIL: {fail_count}").fill = QA_STATUS_FILLS["FAIL"]

    ws.freeze_panes = f"A{header_row + 1}"
