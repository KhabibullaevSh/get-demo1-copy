"""Create sample standard_model_G303.xlsx and rate_library.xlsx for the BOQ system."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def create_standard_model():
    wb = openpyxl.Workbook()

    # ── Sheet 1: Standard BOQ ──
    ws = wb.active
    ws.title = "Standard BOQ"

    headers = ["Item No", "Stock Code", "Description", "Unit", "Qty", "Rate", "Amount", "Category", "Notes"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = THIN_BORDER

    boq_items = [
        # Structure
        [1, "STR-001", "Steel post 100x100x3.0mm x 3000mm", "no", 12, 285.00, None, "Structure", ""],
        [2, "STR-002", "Steel bearer 200x75x3.0mm", "lm", 45, 125.00, None, "Structure", ""],
        [3, "STR-003", "Steel joist 150x50x1.6mm", "lm", 120, 78.00, None, "Structure", ""],
        [4, "STR-004", "Concrete footing pad 600x600x300", "no", 12, 165.00, None, "Structure", ""],
        [5, "STR-005", "Post stirrup bracket galvanised", "no", 12, 45.00, None, "Structure", ""],
        [6, "STR-006", "Bearer bolt M16x200mm galvanised", "no", 48, 12.50, None, "Structure", ""],
        # Roof
        [7, "ROF-001", "Roof sheet Colorbond 0.42mm custom length", "lm", 85, 42.00, None, "Roof", ""],
        [8, "ROF-002", "Roof batten C-section 35x45x0.55mm", "lm", 180, 18.50, None, "Roof", ""],
        [9, "ROF-003", "Roof underlay sarking 1500mm wide", "m2", 145, 8.50, None, "Roof", ""],
        [10, "ROF-004", "Ridge capping Colorbond", "lm", 12, 35.00, None, "Roof", ""],
        [11, "ROF-005", "Barge board Colorbond", "lm", 24, 28.00, None, "Roof", ""],
        [12, "ROF-006", "Fascia board Colorbond", "lm", 36, 25.00, None, "Roof", ""],
        [13, "ROF-007", "Gutter quad 115mm Colorbond", "lm", 36, 22.00, None, "Roof", ""],
        [14, "ROF-008", "Downpipe 90mm Colorbond", "lm", 18, 18.00, None, "Roof", ""],
        [15, "ROF-009", "Roof screw 12-14x50mm Type 17", "no", 850, 0.45, None, "Roof", ""],
        [16, "ROF-010", "Roof insulation R2.5 batts", "m2", 130, 14.00, None, "Roof", ""],
        # Walls - External
        [17, "WAL-001", "Wall frame C-section 75x32x0.75mm", "lm", 210, 22.00, None, "Walls", ""],
        [18, "WAL-002", "External FC sheet 9mm 2400x1200", "no", 55, 68.00, None, "Walls", ""],
        [19, "WAL-003", "External wall wrap breathable membrane", "m2", 165, 6.50, None, "Walls", ""],
        [20, "WAL-004", "External corner flashing Colorbond", "lm", 16, 15.00, None, "Walls", ""],
        # Walls - Internal
        [21, "WAL-005", "Internal plasterboard 10mm 2400x1200", "no", 65, 32.00, None, "Walls", ""],
        [22, "WAL-006", "Plasterboard jointing compound 15kg", "no", 8, 28.00, None, "Walls", ""],
        [23, "WAL-007", "Plasterboard paper tape 75m roll", "no", 6, 8.50, None, "Walls", ""],
        [24, "WAL-008", "Wall screw 6g x 25mm fine thread", "box", 12, 18.00, None, "Walls", ""],
        # Ceiling
        [25, "CEL-001", "Ceiling batten steel 28x20x0.42mm", "lm", 160, 12.00, None, "Ceiling", ""],
        [26, "CEL-002", "Ceiling plasterboard 10mm 2400x1200", "no", 45, 32.00, None, "Ceiling", ""],
        [27, "CEL-003", "Ceiling insulation R2.0 batts", "m2", 110, 12.00, None, "Ceiling", ""],
        # Doors
        [28, "DOR-001", "Entrance door solid core 2040x820mm", "no", 1, 450.00, None, "Doors", ""],
        [29, "DOR-002", "Internal door hollow core 2040x720mm", "no", 5, 185.00, None, "Doors", ""],
        [30, "DOR-003", "Bathroom door 2040x620mm", "no", 1, 195.00, None, "Doors", ""],
        [31, "DOR-004", "Door frame pine 90x30mm set", "set", 7, 85.00, None, "Doors", ""],
        [32, "DOR-005", "Door hinge 100mm stainless steel", "pair", 21, 12.00, None, "Doors", "3 per door"],
        [33, "DOR-006", "Door handle lever set", "set", 7, 35.00, None, "Doors", ""],
        [34, "DOR-007", "Entrance door deadlock", "no", 1, 65.00, None, "Doors", ""],
        # Windows
        [35, "WIN-001", "Aluminium sliding window 1800x1200mm", "no", 2, 380.00, None, "Windows", ""],
        [36, "WIN-002", "Aluminium sliding window 1200x1200mm", "no", 2, 320.00, None, "Windows", ""],
        [37, "WIN-003", "Aluminium louvre window 600x600mm", "no", 2, 185.00, None, "Windows", ""],
        [38, "WIN-004", "Window flashing kit per window", "set", 6, 25.00, None, "Windows", ""],
        # Floor Finishes
        [39, "FLR-001", "Floor tile ceramic 300x300mm", "m2", 35, 42.00, None, "Floor", "Wet areas"],
        [40, "FLR-002", "Floor tile adhesive 20kg bag", "bag", 12, 28.00, None, "Floor", ""],
        [41, "FLR-003", "Floor vinyl sheet 2m wide", "m2", 75, 38.00, None, "Floor", "Living/bedrooms"],
        [42, "FLR-004", "Floor underlay 2mm foam", "m2", 75, 5.50, None, "Floor", ""],
        [43, "FLR-005", "Skirting board pine 60x12mm", "lm", 85, 8.50, None, "Floor", ""],
        # Paint
        [44, "PNT-001", "Interior wall paint acrylic 15L", "pail", 6, 145.00, None, "Paint", ""],
        [45, "PNT-002", "Exterior wall paint weather shield 15L", "pail", 4, 185.00, None, "Paint", ""],
        [46, "PNT-003", "Ceiling paint flat white 15L", "pail", 3, 125.00, None, "Paint", ""],
        [47, "PNT-004", "Primer/sealer 15L", "pail", 3, 110.00, None, "Paint", ""],
        # Verandah
        [48, "VER-001", "Verandah decking hardwood 90x19mm", "m2", 24, 85.00, None, "Verandah", ""],
        [49, "VER-002", "Verandah bearer hardwood 100x75mm", "lm", 12, 45.00, None, "Verandah", ""],
        [50, "VER-003", "Verandah joist hardwood 75x50mm", "lm", 28, 32.00, None, "Verandah", ""],
        [51, "VER-004", "Verandah post 90x90mm hardwood", "no", 4, 125.00, None, "Verandah", ""],
        [52, "VER-005", "Verandah handrail complete", "lm", 8, 65.00, None, "Verandah", ""],
        # Plumbing
        [53, "PLB-001", "PVC pipe 100mm DWV", "lm", 15, 28.00, None, "Plumbing", ""],
        [54, "PLB-002", "PVC pipe 50mm DWV", "lm", 12, 15.00, None, "Plumbing", ""],
        [55, "PLB-003", "Water pipe HDPE 20mm", "lm", 25, 8.50, None, "Plumbing", ""],
        [56, "PLB-004", "Toilet suite complete", "set", 1, 450.00, None, "Plumbing", ""],
        [57, "PLB-005", "Shower base and screen", "set", 1, 380.00, None, "Plumbing", ""],
        [58, "PLB-006", "Bathroom basin and pedestal", "set", 1, 220.00, None, "Plumbing", ""],
        [59, "PLB-007", "Kitchen sink stainless double bowl", "set", 1, 350.00, None, "Plumbing", ""],
        [60, "PLB-008", "Laundry tub stainless", "set", 1, 280.00, None, "Plumbing", ""],
        # Electrical
        [61, "ELC-001", "Twin & earth cable 2.5mm 100m", "roll", 3, 165.00, None, "Electrical", ""],
        [62, "ELC-002", "Power outlet double GPO", "no", 12, 18.00, None, "Electrical", ""],
        [63, "ELC-003", "Light switch single", "no", 8, 12.00, None, "Electrical", ""],
        [64, "ELC-004", "LED downlight 10W", "no", 14, 22.00, None, "Electrical", ""],
        [65, "ELC-005", "Switchboard 12-way", "no", 1, 285.00, None, "Electrical", ""],
        [66, "ELC-006", "Smoke detector hardwired", "no", 3, 45.00, None, "Electrical", ""],
        # Waterproofing
        [67, "WPF-001", "Waterproofing membrane liquid", "m2", 12, 35.00, None, "Waterproofing", "Wet areas"],
        [68, "WPF-002", "Waterproofing tape joints", "lm", 15, 8.00, None, "Waterproofing", ""],
        # Stair (if applicable)
        [69, "STR-010", "Stair stringer steel 250x3mm", "no", 2, 185.00, None, "Stairs", ""],
        [70, "STR-011", "Stair tread hardwood 250x32mm", "no", 8, 45.00, None, "Stairs", ""],
    ]

    for item in boq_items:
        row = item[0] + 1
        for col, val in enumerate(item, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.border = THIN_BORDER
        # Amount formula
        ws.cell(row=row, column=7, value=f"=E{row}*F{row}").border = THIN_BORDER

    # ── Sheet 2: Standard Geometry ──
    ws2 = wb.create_sheet("Standard Geometry")
    ws2.cell(row=1, column=1, value="Element").font = HEADER_FONT
    ws2.cell(row=1, column=1).fill = HEADER_FILL
    ws2.cell(row=1, column=2, value="Value").font = HEADER_FONT
    ws2.cell(row=1, column=2).fill = HEADER_FILL
    ws2.cell(row=1, column=3, value="Unit").font = HEADER_FONT
    ws2.cell(row=1, column=3).fill = HEADER_FILL

    geometry = [
        ["total_floor_area", 110.0, "m²"],
        ["total_wall_length", 82.0, "lm"],
        ["external_wall_length", 48.0, "lm"],
        ["internal_wall_length", 34.0, "lm"],
        ["roof_area", 135.0, "m²"],
        ["roof_perimeter", 52.0, "lm"],
        ["verandah_area", 24.0, "m²"],
        ["ceiling_area", 110.0, "m²"],
        ["door_count", 7, "no"],
        ["window_count", 6, "no"],
        ["post_count", 12, "no"],
        ["stair_count", 1, "no"],
        ["room_count", 7, "no"],
        ["floor_height", 600, "mm"],
        ["wall_height", 2700, "mm"],
        ["roof_pitch", 22.5, "degrees"],
    ]

    for i, (elem, val, unit) in enumerate(geometry):
        row = i + 2
        ws2.cell(row=row, column=1, value=elem).border = THIN_BORDER
        ws2.cell(row=row, column=2, value=val).border = THIN_BORDER
        ws2.cell(row=row, column=3, value=unit).border = THIN_BORDER

    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 10

    # ── Sheet 3: Rules Library ──
    ws3 = wb.create_sheet("Rules Library")
    rule_headers = ["Rule ID", "Target Item", "Depends On", "Formula", "Multiplier", "Description"]
    for col, h in enumerate(rule_headers, 1):
        c = ws3.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = THIN_BORDER

    rules = [
        ["R-001", "ROF-002", "roof_area", "roof_area / 0.9", None, "Roof battens at 900mm spacing"],
        ["R-002", "ROF-003", "roof_area", "roof_area * 1.08", None, "Roof underlay with 8% overlap"],
        ["R-003", "ROF-009", "roof_area", "proportional", 1.0, "Roof screws scale with roof area"],
        ["R-004", "ROF-006", "roof_perimeter", "roof_perimeter * 0.5", None, "Fascia ~50% of roof perimeter"],
        ["R-005", "ROF-007", "roof_perimeter", "roof_perimeter * 0.7", None, "Gutters ~70% of roof perimeter"],
        ["R-006", "ROF-008", "roof_perimeter", "roof_perimeter * 0.35", None, "Downpipes ~35% of perimeter"],
        ["R-007", "WAL-001", "total_wall_length", "total_wall_length * 2.7 / 0.6", None, "Wall studs at 600mm spacing x wall height"],
        ["R-008", "WAL-002", "external_wall_length", "external_wall_length * 2.7 / 2.88", None, "FC sheets: wall area / sheet area (2.88m²)"],
        ["R-009", "WAL-003", "external_wall_length", "external_wall_length * 2.7 * 1.05", None, "Wall wrap: ext wall area + 5% overlap"],
        ["R-010", "WAL-005", "internal_wall_length", "internal_wall_length * 2.7 * 2 / 2.88", None, "Plasterboard: both sides of internal walls"],
        ["R-011", "CEL-001", "ceiling_area", "ceiling_area / 0.45 * 1.05", None, "Ceiling battens at 450mm spacing + 5%"],
        ["R-012", "CEL-002", "ceiling_area", "ceiling_area / 2.88", None, "Ceiling boards: area / sheet size"],
        ["R-013", "CEL-003", "ceiling_area", "ceiling_area * 1.05", None, "Ceiling insulation + 5% overlap"],
        ["R-014", "STR-004", "post_count", "post_count", None, "One footing per post"],
        ["R-015", "STR-005", "post_count", "post_count", None, "One stirrup bracket per post"],
        ["R-016", "STR-006", "post_count", "post_count * 4", None, "4 bearer bolts per post"],
        ["R-017", "DOR-004", "door_count", "door_count", None, "One frame set per door"],
        ["R-018", "DOR-005", "door_count", "door_count * 3", None, "3 hinge pairs per door (1.5 pair per door x 2 = 3)"],
        ["R-019", "DOR-006", "door_count", "door_count", None, "One handle set per door"],
        ["R-020", "WIN-004", "window_count", "window_count", None, "One flashing kit per window"],
        ["R-021", "WAL-008", "total_wall_length", "proportional", 1.0, "Wall screws scale with wall length"],
        ["R-022", "WAL-006", "internal_wall_length", "proportional", 1.0, "Jointing compound scales with internal walls"],
        ["R-023", "ROF-010", "roof_area", "roof_area * 0.95", None, "Roof insulation ~95% of roof area"],
        ["R-024", "VER-001", "verandah_area", "verandah_area * 1.1", None, "Verandah decking + 10% waste"],
        ["R-025", "FLR-005", "total_wall_length", "total_wall_length * 0.9", None, "Skirting ~90% of wall length (less doors)"],
        ["R-026", "WPF-001", "total_floor_area", "proportional", 0.11, "Waterproofing ~11% of floor area (wet areas)"],
    ]

    for i, rule in enumerate(rules):
        row = i + 2
        for col, val in enumerate(rule, 1):
            c = ws3.cell(row=row, column=col, value=val)
            c.border = THIN_BORDER

    for i, w in enumerate([10, 14, 22, 40, 12, 45], 1):
        ws3.column_dimensions[chr(64 + i)].width = w

    wb.save("data/standard_model_G303.xlsx")
    print("Created data/standard_model_G303.xlsx")


def create_rate_library():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rate Library"

    headers = ["Stock Code", "Description", "Unit", "Rate", "Category", "Notes"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = THIN_BORDER

    rates = [
        ["STR-001", "Steel post 100x100x3.0mm x 3000mm", "no", 285.00, "Structure", "Supply only"],
        ["STR-002", "Steel bearer 200x75x3.0mm", "lm", 125.00, "Structure", "Supply only"],
        ["STR-003", "Steel joist 150x50x1.6mm", "lm", 78.00, "Structure", "Supply only"],
        ["STR-004", "Concrete footing pad 600x600x300", "no", 165.00, "Structure", "Including concrete"],
        ["STR-005", "Post stirrup bracket galvanised", "no", 45.00, "Structure", ""],
        ["STR-006", "Bearer bolt M16x200mm galvanised", "no", 12.50, "Structure", ""],
        ["ROF-001", "Roof sheet Colorbond 0.42mm", "lm", 42.00, "Roof", "Custom length"],
        ["ROF-002", "Roof batten C-section 35x45x0.55mm", "lm", 18.50, "Roof", ""],
        ["ROF-003", "Roof underlay sarking 1500mm", "m2", 8.50, "Roof", ""],
        ["ROF-004", "Ridge capping Colorbond", "lm", 35.00, "Roof", ""],
        ["ROF-005", "Barge board Colorbond", "lm", 28.00, "Roof", ""],
        ["ROF-006", "Fascia board Colorbond", "lm", 25.00, "Roof", ""],
        ["ROF-007", "Gutter quad 115mm Colorbond", "lm", 22.00, "Roof", ""],
        ["ROF-008", "Downpipe 90mm Colorbond", "lm", 18.00, "Roof", ""],
        ["ROF-009", "Roof screw 12-14x50mm Type 17", "no", 0.45, "Roof", "Per screw"],
        ["ROF-010", "Roof insulation R2.5 batts", "m2", 14.00, "Roof", ""],
        ["WAL-001", "Wall frame C-section 75x32x0.75mm", "lm", 22.00, "Walls", ""],
        ["WAL-002", "External FC sheet 9mm 2400x1200", "no", 68.00, "Walls", "Per sheet"],
        ["WAL-003", "External wall wrap breathable", "m2", 6.50, "Walls", ""],
        ["WAL-004", "External corner flashing Colorbond", "lm", 15.00, "Walls", ""],
        ["WAL-005", "Internal plasterboard 10mm 2400x1200", "no", 32.00, "Walls", "Per sheet"],
        ["WAL-006", "Plasterboard jointing compound 15kg", "no", 28.00, "Walls", ""],
        ["WAL-007", "Plasterboard paper tape 75m", "no", 8.50, "Walls", "Per roll"],
        ["WAL-008", "Wall screw 6g x 25mm fine thread", "box", 18.00, "Walls", "Per box of 500"],
        ["CEL-001", "Ceiling batten steel 28x20x0.42mm", "lm", 12.00, "Ceiling", ""],
        ["CEL-002", "Ceiling plasterboard 10mm 2400x1200", "no", 32.00, "Ceiling", "Per sheet"],
        ["CEL-003", "Ceiling insulation R2.0 batts", "m2", 12.00, "Ceiling", ""],
        ["DOR-001", "Entrance door solid core 2040x820mm", "no", 450.00, "Doors", ""],
        ["DOR-002", "Internal door hollow core 2040x720mm", "no", 185.00, "Doors", ""],
        ["DOR-003", "Bathroom door 2040x620mm", "no", 195.00, "Doors", ""],
        ["DOR-004", "Door frame pine 90x30mm set", "set", 85.00, "Doors", ""],
        ["DOR-005", "Door hinge 100mm stainless", "pair", 12.00, "Doors", ""],
        ["DOR-006", "Door handle lever set", "set", 35.00, "Doors", ""],
        ["DOR-007", "Entrance door deadlock", "no", 65.00, "Doors", ""],
        ["WIN-001", "Aluminium sliding window 1800x1200mm", "no", 380.00, "Windows", ""],
        ["WIN-002", "Aluminium sliding window 1200x1200mm", "no", 320.00, "Windows", ""],
        ["WIN-003", "Aluminium louvre window 600x600mm", "no", 185.00, "Windows", ""],
        ["WIN-004", "Window flashing kit", "set", 25.00, "Windows", "Per window"],
        ["FLR-001", "Floor tile ceramic 300x300mm", "m2", 42.00, "Floor", ""],
        ["FLR-002", "Floor tile adhesive 20kg bag", "bag", 28.00, "Floor", ""],
        ["FLR-003", "Floor vinyl sheet 2m wide", "m2", 38.00, "Floor", ""],
        ["FLR-004", "Floor underlay 2mm foam", "m2", 5.50, "Floor", ""],
        ["FLR-005", "Skirting board pine 60x12mm", "lm", 8.50, "Floor", ""],
        ["PNT-001", "Interior wall paint acrylic 15L", "pail", 145.00, "Paint", ""],
        ["PNT-002", "Exterior wall paint weather shield 15L", "pail", 185.00, "Paint", ""],
        ["PNT-003", "Ceiling paint flat white 15L", "pail", 125.00, "Paint", ""],
        ["PNT-004", "Primer/sealer 15L", "pail", 110.00, "Paint", ""],
        ["VER-001", "Verandah decking hardwood 90x19mm", "m2", 85.00, "Verandah", ""],
        ["VER-002", "Verandah bearer hardwood 100x75mm", "lm", 45.00, "Verandah", ""],
        ["VER-003", "Verandah joist hardwood 75x50mm", "lm", 32.00, "Verandah", ""],
        ["VER-004", "Verandah post 90x90mm hardwood", "no", 125.00, "Verandah", ""],
        ["VER-005", "Verandah handrail complete", "lm", 65.00, "Verandah", ""],
        ["PLB-001", "PVC pipe 100mm DWV", "lm", 28.00, "Plumbing", ""],
        ["PLB-002", "PVC pipe 50mm DWV", "lm", 15.00, "Plumbing", ""],
        ["PLB-003", "Water pipe HDPE 20mm", "lm", 8.50, "Plumbing", ""],
        ["PLB-004", "Toilet suite complete", "set", 450.00, "Plumbing", ""],
        ["PLB-005", "Shower base and screen", "set", 380.00, "Plumbing", ""],
        ["PLB-006", "Bathroom basin and pedestal", "set", 220.00, "Plumbing", ""],
        ["PLB-007", "Kitchen sink stainless double bowl", "set", 350.00, "Plumbing", ""],
        ["PLB-008", "Laundry tub stainless", "set", 280.00, "Plumbing", ""],
        ["ELC-001", "Twin & earth cable 2.5mm 100m", "roll", 165.00, "Electrical", ""],
        ["ELC-002", "Power outlet double GPO", "no", 18.00, "Electrical", ""],
        ["ELC-003", "Light switch single", "no", 12.00, "Electrical", ""],
        ["ELC-004", "LED downlight 10W", "no", 22.00, "Electrical", ""],
        ["ELC-005", "Switchboard 12-way", "no", 285.00, "Electrical", ""],
        ["ELC-006", "Smoke detector hardwired", "no", 45.00, "Electrical", ""],
        ["WPF-001", "Waterproofing membrane liquid", "m2", 35.00, "Waterproofing", ""],
        ["WPF-002", "Waterproofing tape joints", "lm", 8.00, "Waterproofing", ""],
        ["STR-010", "Stair stringer steel 250x3mm", "no", 185.00, "Stairs", ""],
        ["STR-011", "Stair tread hardwood 250x32mm", "no", 45.00, "Stairs", ""],
    ]

    for i, rate in enumerate(rates):
        row = i + 2
        for col, val in enumerate(rate, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.border = THIN_BORDER

    for i, w in enumerate([14, 45, 8, 12, 15, 20], 1):
        ws.column_dimensions[chr(64 + i)].width = w

    wb.save("data/rate_library.xlsx")
    print("Created data/rate_library.xlsx")


if __name__ == "__main__":
    create_standard_model()
    create_rate_library()
    print("Done!")
