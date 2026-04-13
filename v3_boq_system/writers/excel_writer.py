"""
excel_writer.py — Write BOQ to Excel with companion sheets.

Sheets produced:
  1. BOQ (Commercial) — baseline-aligned commercial presentation
       - Items grouped by commercial_package_code (remapped for commercial logic)
       - Within each section: main family rows first, accessories after, MR/PH last
       - Columns: STOCK CODE | DESCRIPTION | QTY | UNIT | RATE | AMOUNT | CONF | NOTES
       - Confidence shown as row background colour (no verbose column)
  2. Engine Truth   — full source-driven rows with confidence/evidence (all items)
  3. Traceability   — evidence + derivation rule for every item
  4. Manual Review  — items flagged for review
  5. QA Summary     — package completeness + provenance stats
  6. Source Summary — which files contributed which data
"""
from __future__ import annotations

import logging
import os
from collections import defaultdict
from typing import Any

log = logging.getLogger("boq.v3.excel_writer")

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False


# ── Colour palette — matched to final BOQ package code sections ───────────────
_SECTION_COLOURS_FINAL: dict[str, str] = {
    "50106": "FFD6E4F0",   # WPC Timber — light blue
    "50107": "FFD5E8D4",   # Structural — light green
    "50111": "FFDAE8FC",   # Fixings — sky
    "50112": "FFD5F5E3",   # Roof — mint
    "50113": "FFFFE6CC",   # Ext Cladding — peach
    "50114": "FFFCF3CF",   # Openings — yellow
    "50115": "FFFDEBD0",   # Int Linings — apricot
    "50116": "FFF5CBA7",   # Painting — warm orange
    "50117": "FFE8DAEF",   # Services — lavender
    "50118": "FFEAF4FB",   # Insulation — pale sky
    "50119": "FFFDEDEC",   # Electrical — pale pink
    "50124": "FFFEF9E7",   # Stairs — cream
    "50129": "FFF2F3F4",   # FFE — light grey
    "50199": "FFEEEEEE",   # Unclassified
}

# Backward-compat colours for legacy A–K section names
_SECTION_COLOURS_LEGACY: dict[str, str] = {
    "A - Structural Frame":   "FFD6E4F0",
    "B - Roof":               "FFD5F5E3",
    "C - Insulation":         "FFE8DAEF",
    "D - Openings":           "FFFCF3CF",
    "E - Linings & Ceilings": "FFFDEBD0",
    "F - Finishes":           "FFFDEDEC",
    "G - Floor System":       "FFEBF5FB",
    "H - Substructure":       "FFF2F3F4",
    "I - Services":           "FFEBF5FB",
    "J - Stairs":             "FFFEF9E7",
    "K - External Works":     "FFEAFAF1",
}

_HEADER_FILL               = "FF2C3E50"   # dark slate
_SECTION_HDR_FILL          = "FF3D5A80"   # medium blue  — section headers
_COMMERCIAL_BLOCK_HDR_FILL = "FF5B8DB8"   # steel blue   — commercial block headers
_TRADE_GROUP_HDR_FILL      = "FF5B8DB8"   # steel blue   — trade group headers (Phase 5 legacy)
_SUBGROUP_HDR_FILL         = "FFB8D0E8"   # light blue   — optional subgroup headers
_MR_FILL              = "FFFFF3CD"   # light amber for manual review rows
_PLACEHOLDER_FILL     = "FFFFE0E0"   # light red for placeholder rows

_THIN   = Side(style="thin",   color="FFCCCCCC")
_MEDIUM = Side(style="medium", color="FF999999")
_BORDER = Border(left=_THIN, right=_THIN, bottom=_THIN)
_SECTION_BORDER = Border(
    left=Side(style="medium", color="FF3D5A80"),
    right=Side(style="medium", color="FF3D5A80"),
    bottom=Side(style="medium", color="FF3D5A80"),
)

# Commercial BOQ confidence colours (row background)
_CONF_FILL = {
    "HIGH":   "FFD5F5E3",   # mint green
    "MEDIUM": "FFFEF9E7",   # cream
    "LOW":    "FFFDEBD0",   # apricot
}
_ACCESSORY_THRESHOLD = 400   # family_sort_key >= this → accessory row
_MR_THRESHOLD       = 1000   # family_sort_key >= this → MR row in commercial view

# Item names that should be consolidated when the same name appears multiple times
# within the same commercial section (e.g. per-door-type hardware rows).
# Consolidation sums quantities and shows a single aggregate row in the commercial view.
# Engine Truth sheet is NEVER affected — all per-type rows are preserved there.
# Display names (item_display_name) to consolidate in the commercial view.
# Items sharing the same display name, commercial section, and unit are summed
# into a single aggregate row.  Engine Truth is unaffected.
_HARDWARE_CONSOLIDATE_NAMES: frozenset = frozenset([
    # Door hardware (display names — aggregated across door types)
    "Door | Leaf", "Door | Frame Set", "Door | Hinge (pair)",
    "Door | Lockset", "Door | Stop", "Door | Closer, Hydraulic",
    # Window hardware (display names — aggregated across window types)
    "Window | Louvre Frame", "Window | Louvre Blade", "Window | Fly Screen",
    # Window flashings — one row per window type in engine; single total in commercial
    "Flashing | Window Head, Galvanised", "Flashing | Window Sill, Galvanised",
    # Timber architrave — door + window architrave shown as single lm total
    "Timber Architrave | 12 x 75 x 2400mm Timber",
])


def _apply_fill(cell, hex_argb: str) -> None:
    cell.fill = PatternFill(fill_type="solid", fgColor=hex_argb)


def _header_row(ws, headers: list[str], row: int = 1) -> None:
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor=_HEADER_FILL)
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def _set_col_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_boq_excel(
    boq_items:    list[dict],
    qa_report:    dict,
    output_path:  str,
    project_name: str = "",
    source_files: list[str] | None = None,
) -> str:
    """
    Write all output sheets to an Excel workbook.

    Returns the output file path.
    """
    if not _HAS_OPENPYXL:
        log.error("openpyxl not installed — Excel output skipped")
        return ""

    wb = openpyxl.Workbook()

    # ── Sheet 1: BOQ (Commercial — baseline-aligned presentation) ─────────────
    ws_boq = wb.active
    ws_boq.title = "BOQ"
    _write_commercial_boq_sheet(ws_boq, boq_items, project_name)

    # ── Sheet 2: Engine Truth (full source-driven rows) ───────────────────────
    ws_truth = wb.create_sheet("Engine Truth")
    _write_boq_sheet(ws_truth, boq_items, project_name)

    # ── Sheet 3: Traceability ─────────────────────────────────────────────────
    ws_trace = wb.create_sheet("Traceability")
    _write_traceability_sheet(ws_trace, boq_items)

    # ── Sheet 4: Manual Review ────────────────────────────────────────────────
    ws_mr = wb.create_sheet("Manual Review")
    _write_manual_review_sheet(ws_mr, boq_items)

    # ── Sheet 5: QA Summary ───────────────────────────────────────────────────
    ws_qa = wb.create_sheet("QA Summary")
    _write_qa_sheet(ws_qa, qa_report)

    # ── Sheet 6: Source Summary ───────────────────────────────────────────────
    ws_src = wb.create_sheet("Source Summary")
    _write_source_sheet(ws_src, boq_items, source_files or [])

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    log.info("Excel written: %s", output_path)
    return output_path


def _build_commercial_items(boq_items: list[dict]) -> list[dict]:
    """
    Prepare items for the commercial BOQ sheet.

    Consolidates duplicate hardware rows that have the same item_name within the
    same commercial_package_code section (e.g. Door Leaf appears once per door type
    in the engine but should be shown as a single aggregate row in the commercial view).

    Rules:
    - Only consolidates names listed in _HARDWARE_CONSOLIDATE_NAMES
    - All duplicates must share the same commercial_package_code and unit
    - None of the duplicates may be manual_review (MR items are never merged)
    - Quantity = sum of all individual quantities
    - Confidence = most conservative (lowest) of the group
    - Notes = notes from first item + consolidation annotation
    - family_sort_key = minimum across the group (earliest position in section)

    Engine Truth sheet is NOT affected — it calls _write_boq_sheet with full boq_items.
    """
    import copy

    _CONF_ORDER = {"HIGH": 2, "MEDIUM": 1, "LOW": 0}

    # First pass: collect items into groups keyed by (comm_pkg, display_name).
    # Using display_name (normalised) allows window flashings and architrave to
    # consolidate even though their item_names are type-specific.
    from collections import defaultdict
    groups: dict[tuple, list[dict]] = defaultdict(list)
    for item in boq_items:
        name = item.get("item_display_name") or item.get("item_name", "")
        pkg  = item.get("commercial_package_code", "50199")
        groups[(pkg, name)].append(item)

    # Build the consolidated list
    result: list[dict] = []
    seen_keys: set = set()

    for item in boq_items:
        name = item.get("item_display_name") or item.get("item_name", "")
        pkg  = item.get("commercial_package_code", "50199")
        key  = (pkg, name)

        if key in seen_keys:
            continue  # already handled as part of a merged group

        group = groups[key]

        # Only consolidate names in the allowlist with 2+ non-MR items sharing same unit
        units = {it.get("unit", "") for it in group}
        all_non_mr = not any(it.get("manual_review", False) for it in group)
        if (name in _HARDWARE_CONSOLIDATE_NAMES
                and len(group) > 1
                and len(units) == 1
                and all_non_mr):
            # Merge into a single aggregate row
            merged = copy.deepcopy(group[0])
            total_qty = sum(
                it["quantity"] for it in group
                if it.get("quantity") is not None
            )
            min_conf = min(
                group,
                key=lambda x: _CONF_ORDER.get(x.get("confidence", "LOW"), 0),
            )
            merged["quantity"] = round(total_qty, 3)
            merged["confidence"] = min_conf.get("confidence", "LOW")
            merged["family_sort_key"] = min(
                it.get("family_sort_key", 500) for it in group
            )
            base_note = group[0].get("notes", "") or ""
            merged["notes"] = (
                f"[Commercial total — {len(group)} types combined: "
                + ", ".join(
                    f"{it.get('notes','').split('swing=')[-1].split(')')[0] or str(it.get('quantity','?'))}"
                    for it in group
                )
                + f"] {base_note}"
            ).strip()
            result.append(merged)
            seen_keys.add(key)
        else:
            result.append(item)
            seen_keys.add(key) if len(group) == 1 else None

    return result


def _write_commercial_boq_sheet(ws, boq_items: list[dict], project_name: str) -> None:
    """
    Write the commercial BOQ sheet — baseline-aligned, clean presentation.

    Columns:
      A: STOCK CODE
      B: MATERIALS DESCRIPTION
      C: QTY
      D: UNIT
      E: RATE (PGK)
      F: AMOUNT (PGK)
      G: CONF
      H: NOTES

    Section structure:
      - Groups by commercial_package_code (remapped vs engine package_code)
      - Within each section: sorted by family_sort_key
        0-399: main material / structural rows
        400-999: accessory / fixing / adhesive rows (lighter visual treatment)
        1000+: manual review / placeholder rows (amber, at bottom of each section)
      - Accessory sub-group separated by a thin divider row
      - MR/placeholder sub-group separated by a thicker divider row
    """
    title = f"BILL OF QUANTITIES — {project_name}" if project_name else "BILL OF QUANTITIES"
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14, color="FF2C3E50")
    ws.merge_cells("A1:H1")
    ws.row_dimensions[1].height = 28

    headers = ["STOCK CODE", "MATERIALS DESCRIPTION", "QTY", "UNIT",
               "RATE (PGK)", "AMOUNT (PGK)", "CONF", "NOTES"]
    _header_row(ws, headers, row=2)
    ws.row_dimensions[2].height = 28

    # Consolidate repeated hardware rows (per-door-type → single aggregate per type)
    # before sorting. Engine Truth sheet receives unmodified boq_items.
    commercial_items = _build_commercial_items(boq_items)

    # Sort by commercial package → trade_group_sort_key (estimator mode) or
    # family_sort_key (commercial / engine mode).
    # trade_group_sort_key encodes both trade-group order and within-group
    # family order in a single integer, so the sort is always correct.
    _PKG_ORDER = ["50106","50107","50111","50112","50113",
                  "50114","50115","50116","50117","50118","50119","50124","50129","50199"]
    def _ckey(item: dict):
        code = item.get("commercial_package_code", "50199")
        try:
            idx = _PKG_ORDER.index(code)
        except ValueError:
            idx = len(_PKG_ORDER)
        # Prefer commercial_block_sort_key (Phase 6 — commercial blocks)
        # Fallback: trade_group_sort_key (Phase 5), then family_sort_key
        sk = item.get("commercial_block_sort_key",
              item.get("trade_group_sort_key",
              item.get("family_sort_key", 500)))
        return (idx, sk, item.get("item_no", ""))

    sorted_items = sorted(commercial_items, key=_ckey)

    current_comm_pkg = ""
    prev_fam_key = -1
    row_idx = 3

    for item in sorted_items:
        comm_pkg   = item.get("commercial_package_code", "50199")
        fam_key    = item.get("family_sort_key", 500)  # used for divider logic only
        is_mr      = item.get("manual_review", False)
        is_ph      = item.get("quantity_status") == "placeholder"
        drule      = item.get("derivation_rule", "")
        is_cb_hdr  = drule == "insert_commercial_block_headers"
        is_tg_hdr  = drule == "insert_trade_group_headers"    # Phase 5 legacy
        is_sg_hdr  = (item.get("export_class") == "export_only_grouping"
                      and not is_cb_hdr and not is_tg_hdr)
        conf       = item.get("confidence", "LOW")
        sec_label  = item.get("commercial_section_label", comm_pkg)

        # ── Section header (when commercial package changes) ──────────────────
        if comm_pkg != current_comm_pkg:
            if current_comm_pkg:
                ws.row_dimensions[row_idx].height = 6
                row_idx += 1

            hdr_cell = ws.cell(row=row_idx, column=1, value=sec_label)
            hdr_cell.font  = Font(bold=True, color="FFFFFFFF", size=11)
            hdr_cell.fill  = PatternFill(fill_type="solid", fgColor=_SECTION_HDR_FILL[2:])
            hdr_cell.alignment = Alignment(vertical="center")
            for col in range(1, len(headers) + 1):
                c = ws.cell(row=row_idx, column=col)
                c.fill   = PatternFill(fill_type="solid", fgColor=_SECTION_HDR_FILL[2:])
                c.border = _SECTION_BORDER
            ws.merge_cells(f"A{row_idx}:H{row_idx}")
            ws.row_dimensions[row_idx].height = 20
            row_idx += 1
            current_comm_pkg = comm_pkg
            prev_fam_key = -1

        # ── Commercial block header (estimator mode — Phase 6) ───────────────
        if is_cb_hdr:
            cb_name  = item.get("item_display_name", "")
            cb_cell  = ws.cell(row=row_idx, column=1, value=f"  {cb_name}")
            cb_cell.font      = Font(bold=True, color="FFFFFFFF", size=10)
            cb_cell.alignment = Alignment(vertical="center")
            for col in range(1, len(headers) + 1):
                c = ws.cell(row=row_idx, column=col)
                c.fill   = PatternFill(fill_type="solid",
                                       fgColor=_COMMERCIAL_BLOCK_HDR_FILL[2:])
                c.border = Border(
                    bottom=Side(style="medium", color="FF2A5F8E"),
                    left =Side(style="thin",   color="FF2A5F8E"),
                    right=Side(style="thin",   color="FF2A5F8E"),
                )
            ws.merge_cells(f"A{row_idx}:H{row_idx}")
            ws.row_dimensions[row_idx].height = 16
            row_idx += 1
            prev_fam_key = -1   # reset divider logic for each new commercial block
            continue

        # ── Trade group header (Phase 5 legacy — still supported) ─────────────
        if is_tg_hdr:
            tg_name  = item.get("item_display_name", "")
            tg_cell  = ws.cell(row=row_idx, column=1, value=f"  {tg_name}")
            tg_cell.font      = Font(bold=True, color="FFFFFFFF", size=10)
            tg_cell.alignment = Alignment(vertical="center")
            for col in range(1, len(headers) + 1):
                c = ws.cell(row=row_idx, column=col)
                c.fill   = PatternFill(fill_type="solid",
                                       fgColor=_TRADE_GROUP_HDR_FILL[2:])
                c.border = Border(
                    bottom=Side(style="medium", color="FF2A5F8E"),
                    left =Side(style="thin",   color="FF2A5F8E"),
                    right=Side(style="thin",   color="FF2A5F8E"),
                )
            ws.merge_cells(f"A{row_idx}:H{row_idx}")
            ws.row_dimensions[row_idx].height = 16
            row_idx += 1
            prev_fam_key = -1   # reset divider logic for each new trade group
            continue

        # ── Subgroup header (optional, legacy estimator mode) ─────────────────
        if is_sg_hdr:
            sg_name = item.get("item_display_name", "")
            sg_cell = ws.cell(row=row_idx, column=1, value=f"    {sg_name}")
            sg_cell.font      = Font(bold=True, italic=True,
                                     color="FF1A3A5C", size=9)
            sg_cell.alignment = Alignment(vertical="center")
            for col in range(1, len(headers) + 1):
                c = ws.cell(row=row_idx, column=col)
                c.fill   = PatternFill(fill_type="solid",
                                       fgColor=_SUBGROUP_HDR_FILL[2:])
                c.border = Border(bottom=Side(style="thin", color="FF8DB0CC"))
            ws.merge_cells(f"A{row_idx}:H{row_idx}")
            ws.row_dimensions[row_idx].height = 14
            row_idx += 1
            prev_fam_key = fam_key
            continue

        # ── Thin accessory divider (first time fam_key enters accessory range)
        elif (prev_fam_key < _ACCESSORY_THRESHOLD <= fam_key < _MR_THRESHOLD
              and not is_mr and not is_ph):
            for col in range(1, len(headers) + 1):
                c = ws.cell(row=row_idx, column=col)
                c.fill = PatternFill(fill_type="solid", fgColor="FFE8E8E8")
                c.border = Border(bottom=Side(style="thin", color="FFBBBBBB"))
            ws.row_dimensions[row_idx].height = 4
            row_idx += 1

        # ── MR / placeholder sub-group divider
        elif prev_fam_key < _MR_THRESHOLD <= fam_key:
            for col in range(1, len(headers) + 1):
                c = ws.cell(row=row_idx, column=col)
                c.fill = PatternFill(fill_type="solid", fgColor="FFFCE8C3")
                c.border = Border(bottom=Side(style="medium", color="FFCC9900"))
            label_cell = ws.cell(
                row=row_idx, column=1,
                value="▸ Items below require manual review / site confirmation",
            )
            label_cell.font = Font(italic=True, color="FF996600", size=9)
            ws.merge_cells(f"A{row_idx}:H{row_idx}")
            ws.row_dimensions[row_idx].height = 14
            row_idx += 1

        prev_fam_key = fam_key

        # ── Row fill based on confidence + MR/PH status ───────────────────────
        if is_ph:
            row_fill = _PLACEHOLDER_FILL[2:]
        elif is_mr:
            row_fill = _MR_FILL[2:]
        elif fam_key >= _ACCESSORY_THRESHOLD:
            row_fill = "FFF5F5F5"  # very light grey for accessories
        else:
            fill_hex = _CONF_FILL.get(conf, "FFFFFFFF")
            row_fill = fill_hex[2:] if fill_hex.startswith("FF") else fill_hex

        display_name = item.get("item_display_name") or item.get("item_name", "")
        stock_code   = item.get("item_code", "")
        qty          = item.get("quantity")
        unit         = item.get("unit", "")
        # Short notes for commercial view (strip long derivation text)
        notes_raw = item.get("notes", "")
        short_notes = notes_raw[:120] if notes_raw else ""

        values = [
            stock_code,
            display_name,
            qty if qty is not None else "",
            unit,
            None,   # RATE — not populated
            None,   # AMOUNT — not populated
            conf[:1],  # H/M/L single char
            short_notes,
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill   = PatternFill(fill_type="solid", fgColor=row_fill)
            cell.border = _BORDER
            if col_idx == 2:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col_idx == 3:
                cell.alignment = Alignment(horizontal="right", vertical="center")
            if col_idx == 8:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        row_idx += 1

    _set_col_widths(ws, [16, 50, 8, 6, 10, 12, 5, 45])
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:H{row_idx - 1}"


def _write_boq_sheet(ws, boq_items: list[dict], project_name: str) -> None:
    """
    Write Engine Truth BOQ sheet — full source-driven rows with traceability metadata.

    Column layout:
      A: STOCK CODE
      B: MATERIALS DESCRIPTION
      C: QTY
      D: (quantity basis note)
      E: RATE (PGK)
      F: AMOUNT (PGK)
      G: CONFIDENCE
      H: SOURCE
      I: NOTES / DRAWING REF

    Section header rows:
      Use engine BOQ numeric package codes (50106–50129).
    """
    # ── Title row ─────────────────────────────────────────────────────────────
    title = f"BILL OF QUANTITIES — {project_name}" if project_name else "BILL OF QUANTITIES"
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14, color="FF2C3E50")
    ws.merge_cells("A1:I1")
    ws.row_dimensions[1].height = 28

    # ── Column headers (row 2, matching final BOQ) ────────────────────────────
    headers = [
        "STOCK CODE",
        "MATERIALS DESCRIPTION",
        "QTY",
        "",                   # quantity basis / blank
        "RATE (PGK)",
        "AMOUNT (PGK)",
        "CONFIDENCE",
        "SOURCE",
        "NOTES / DRAWING REF",
    ]
    _header_row(ws, headers, row=2)
    ws.row_dimensions[2].height = 30

    current_pkg_code = ""
    row_idx = 3

    for item in boq_items:
        pkg_code     = item.get("package_code", "50199")
        section_lbl  = item.get("boq_section_final", item.get("boq_section", ""))
        fill_hex     = _SECTION_COLOURS_FINAL.get(pkg_code, "FFFFFFFF")
        # strip leading "FF" for openpyxl PatternFill
        fill_rgb     = fill_hex[2:] if fill_hex.startswith("FF") else fill_hex

        # ── Section header row (when package changes) ─────────────────────────
        if pkg_code != current_pkg_code:
            hdr_cell = ws.cell(row=row_idx, column=1, value=section_lbl)
            hdr_cell.font  = Font(bold=True, color="FFFFFFFF", size=10)
            hdr_cell.fill  = PatternFill(fill_type="solid", fgColor=_SECTION_HDR_FILL[2:])
            hdr_cell.alignment = Alignment(vertical="center")
            for col in range(1, len(headers) + 1):
                c = ws.cell(row=row_idx, column=col)
                c.fill   = PatternFill(fill_type="solid", fgColor=_SECTION_HDR_FILL[2:])
                c.border = _SECTION_BORDER
            ws.row_dimensions[row_idx].height = 18
            row_idx += 1
            current_pkg_code = pkg_code

        # ── Data row ──────────────────────────────────────────────────────────
        mr  = item.get("manual_review", False)
        ph  = item.get("quantity_status") == "placeholder"
        qty = item.get("quantity")

        if mr:
            row_fill = _MR_FILL[2:]
        elif ph:
            row_fill = _PLACEHOLDER_FILL[2:]
        else:
            row_fill = fill_rgb

        # Use normalized display name for column B; stock code for column A
        display_name = item.get("item_display_name") or item.get("item_name", "")
        stock_code   = item.get("item_code", "")
        qty_basis    = item.get("quantity_status", "")   # brief provenance in blank col
        source_tag   = item.get("source_evidence", "").split(":")[0].strip()[:30] if item.get("source_evidence") else ""

        values = [
            stock_code,       # A: STOCK CODE
            display_name,     # B: MATERIALS DESCRIPTION (normalized)
            qty if qty is not None else "",   # C: QTY
            qty_basis,        # D: blank / qty_basis note
            None,             # E: RATE (PGK) — not populated (requires rate library)
            None,             # F: AMOUNT (PGK) — not populated
            item.get("confidence", ""),       # G: CONFIDENCE
            source_tag,       # H: SOURCE
            item.get("notes", ""),            # I: NOTES / DRAWING REF
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill   = PatternFill(fill_type="solid", fgColor=row_fill)
            cell.border = _BORDER
            if col_idx == 2:   # description — wrap text
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col_idx == 9:   # notes
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col_idx == 3:   # qty — right-align
                cell.alignment = Alignment(horizontal="right", vertical="center")

        row_idx += 1

    # ── Column widths matching final BOQ proportions ──────────────────────────
    _set_col_widths(ws, [16, 48, 8, 14, 10, 12, 12, 20, 45])
    ws.freeze_panes = "A3"

    # ── Auto-filter on header row ─────────────────────────────────────────────
    ws.auto_filter.ref = f"A2:I{row_idx - 1}"


def _write_traceability_sheet(ws, boq_items: list[dict]) -> None:
    headers = [
        "Item No", "Package Code", "Section (Final BOQ)", "Item Name (Original)",
        "Item Display Name", "Unit", "Qty",
        "Status", "Quantity Basis", "Source Evidence", "Derivation Rule", "Confidence"
    ]
    _header_row(ws, headers)
    for row_idx, item in enumerate(boq_items, 2):
        vals = [
            item.get("item_no", ""),
            item.get("package_code", ""),
            item.get("boq_section_final", item.get("boq_section", "")),
            item.get("item_name", ""),
            item.get("item_display_name", ""),
            item.get("unit", ""),
            item.get("quantity"),
            item.get("quantity_status", ""),
            item.get("quantity_basis", ""),
            item.get("source_evidence", ""),
            item.get("derivation_rule", ""),
            item.get("confidence", ""),
        ]
        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = _BORDER
    _set_col_widths(ws, [8, 8, 38, 40, 40, 6, 10, 12, 30, 50, 50, 10])
    ws.freeze_panes = "A2"


def _write_manual_review_sheet(ws, boq_items: list[dict]) -> None:
    headers = [
        "Item No", "Package Code", "Item Display Name", "Unit", "Qty",
        "Confidence", "Status", "Notes / Action Required"
    ]
    _header_row(ws, headers)
    mr_items = [i for i in boq_items if i.get("manual_review")]
    for row_idx, item in enumerate(mr_items, 2):
        vals = [
            item.get("item_no", ""),
            item.get("package_code", ""),
            item.get("item_display_name") or item.get("item_name", ""),
            item.get("unit", ""),
            item.get("quantity"),
            item.get("confidence", ""),
            item.get("quantity_status", ""),
            item.get("notes", ""),
        ]
        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = PatternFill(fill_type="solid", fgColor="FFFFF3CD")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = _BORDER
    _set_col_widths(ws, [8, 10, 48, 6, 10, 10, 12, 60])
    ws.freeze_panes = "A2"


def _write_qa_sheet(ws, qa_report: dict) -> None:
    ws.cell(1, 1, "QA SUMMARY REPORT").font = Font(bold=True, size=14)

    row = 3
    prov = qa_report.get("provenance_summary", {})
    ws.cell(row, 1, "QUANTITY PROVENANCE").font = Font(bold=True)
    row += 1
    for label, key in [
        ("Total Items", "total"), ("Measured", "measured"),
        ("Calculated", "calculated"), ("Inferred", "inferred"),
        ("Placeholder", "placeholder"), ("Manual Review", "manual_review_count"),
    ]:
        ws.cell(row, 1, label)
        ws.cell(row, 2, prov.get(key, 0))
        row += 1

    row += 1
    conf = qa_report.get("confidence_summary", {})
    ws.cell(row, 1, "CONFIDENCE BREAKDOWN").font = Font(bold=True)
    row += 1
    for label in ["HIGH", "MEDIUM", "LOW"]:
        ws.cell(row, 1, label)
        ws.cell(row, 2, conf.get(label, 0))
        ws.cell(row, 3, f"{conf.get(f'pct_{label.lower()}', 0):.1f}%")
        row += 1

    row += 1
    ws.cell(row, 1, "PACKAGE COMPLETENESS (V3 sections)").font = Font(bold=True)
    row += 1
    ws.cell(row, 1, "Package").font = Font(bold=True)
    ws.cell(row, 2, "Status").font = Font(bold=True)
    ws.cell(row, 3, "Items").font = Font(bold=True)
    ws.cell(row, 4, "Measured").font = Font(bold=True)
    ws.cell(row, 5, "Calculated").font = Font(bold=True)
    ws.cell(row, 6, "Inferred").font = Font(bold=True)
    ws.cell(row, 7, "Placeholder").font = Font(bold=True)
    row += 1
    for pkg, data in qa_report.get("package_completeness", {}).items():
        ws.cell(row, 1, pkg)
        ws.cell(row, 2, data.get("status", ""))
        ws.cell(row, 3, data.get("item_count", 0))
        ws.cell(row, 4, data.get("measured", 0))
        ws.cell(row, 5, data.get("calculated", 0))
        ws.cell(row, 6, data.get("inferred", 0))
        ws.cell(row, 7, data.get("placeholder", 0))
        if data.get("status") == "MISSING":
            for col in range(1, 8):
                ws.cell(row, col).fill = PatternFill(fill_type="solid", fgColor="FFFFE0E0")
        row += 1

    row += 1
    ws.cell(row, 1, "WARNINGS / ISSUES").font = Font(bold=True)
    row += 1
    for w in qa_report.get("warnings", []):
        ws.cell(row, 1, w)
        row += 1

    _set_col_widths(ws, [40, 12, 8, 10, 10, 10, 12])


def _write_source_sheet(ws, boq_items: list[dict], source_files: list[str]) -> None:
    ws.cell(1, 1, "SOURCE SUMMARY").font = Font(bold=True, size=14)
    row = 3

    ws.cell(row, 1, "Input Files").font = Font(bold=True)
    row += 1
    for f in source_files:
        ws.cell(row, 1, os.path.basename(f))
        row += 1

    row += 1
    ws.cell(row, 1, "Quantity Status Breakdown by Source").font = Font(bold=True)
    row += 1
    src_counts: dict[str, dict] = {}
    for item in boq_items:
        ev = item.get("source_evidence", "unknown")
        src_label = ev.split(":")[0].strip() if ":" in ev else ev.split(" ")[0]
        src_counts.setdefault(src_label, {"count": 0})
        src_counts[src_label]["count"] += 1

    for src, data in sorted(src_counts.items()):
        ws.cell(row, 1, src)
        ws.cell(row, 2, data["count"])
        row += 1

    _set_col_widths(ws, [50, 12])
