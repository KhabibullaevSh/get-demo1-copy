"""
baseline_profiler.py — Reads a reference/estimator BOQ Excel workbook and
extracts a structured profile of its section layout, family distribution,
unit conventions, and stylistic flags.

The profiler is fully generic: it does not assume any specific project type,
building programme, or package code scheme.  It detects these from the file
itself using pattern rules.

Output schema (returned as a plain dict)
-----------------------------------------
{
  "sections": {
    "<code>": {
      "label": str,
      "row_count": int,
      "non_empty_rows": int,
      "families": list[str],          # unique family names present
      "family_counts": {str: int},    # family → row count
      "units_seen": {str: int},       # unit → frequency
      "dominant_unit": str,
      "style_flags": {
        "uses_stock_length_unit": bool,   # "len" appears as a unit
        "uses_set_unit": bool,            # "set" or "sets" present
        "uses_each": bool,                # "each" present
        "fixings_embedded": bool,         # fixings families inside this section
        "has_placeholder_rows": bool,     # qty=0 / DUP / placeholder rows
        "has_duplicate_rows": bool,
        "uses_lm_unit": bool,
        "uses_m2_unit": bool,
        "uses_nr_unit": bool,
        "uses_m3_unit": bool,
      }
    }
  },
  "global_flags": {
    "fixings_standalone_section": bool,
    "services_placeholder": bool,
    "ffe_section_present": bool,
    "stairs_section_present": bool,
    "insulation_section_present": bool,
    "electrical_section_present": bool,
    "hydraulics_section_present": bool,
  },
  "column_map": {str: int},   # detected header → 0-based column index
  "source_file": str,
  "total_rows_parsed": int,
}
"""

from __future__ import annotations
import re
from collections import Counter
from pathlib import Path

import openpyxl

from .family_classifier import classify


# ---------------------------------------------------------------------------
# Column-header patterns
# ---------------------------------------------------------------------------

_COL_PATTERNS: dict[str, list[str]] = {
    "stock_code":   ["stock code", "code", "item code"],
    "description":  ["description", "material", "item", "materials"],
    "quantity":     ["qty", "quantity", "count"],
    "unit":         ["unit"],
    "rate":         ["rate"],
    "amount":       ["amount", "total"],
    "confidence":   ["confidence", "conf"],
    "source":       ["source"],
    "notes":        ["note", "drawing ref", "ref"],
}

# Package/section header patterns (match cell values that look like section
# headings, e.g. "50106 - WPC" or "50107 - STRUCTURAL …")
# Must have whitespace around the separator to distinguish from stock codes
# like "50110-ADH-002" (no spaces, code immediately after dash).
_SECTION_HEADER_RE = re.compile(
    r"(?P<code>\d{5})\s+[-–—]\s+(?P<label>.+)", re.IGNORECASE
)
# Stock code pattern — if cell matches this it is NOT a section header
_STOCK_CODE_RE = re.compile(r"^\d{5}-[A-Z]{2,5}-\d", re.IGNORECASE)

# Units to normalise to canonical form
_UNIT_NORMALISE: dict[str, str] = {
    "each": "each", "ea": "each", "no": "nr", "nr": "nr",
    "lm": "lm", "m": "lm", "meter": "lm", "metre": "lm",
    "m2": "m2", "m²": "m2", "sqm": "m2",
    "m3": "m3", "m³": "m3",
    "len": "len", "length": "len", "lengths": "len",
    "set": "set", "sets": "set",
    "pair": "pair", "pairs": "pair",
    "roll": "roll", "rolls": "roll",
    "bag": "bag", "bags": "bag",
    "pcs": "pcs", "pc": "pcs",
    "kg": "kg", "t": "t",
    "l": "l", "litre": "l",
    "lm": "lm",
}

# Rows with these descriptions are considered duplicate/placeholder markers
_DUP_MARKERS = {"dup", "duplicate", "duplicate row", "n/a", "none"}


def _normalise_unit(raw: str | None) -> str:
    if not raw:
        return ""
    key = str(raw).strip().lower()
    return _UNIT_NORMALISE.get(key, key)


def _is_section_header(cell_value: str | None) -> tuple[str, str] | None:
    """Return (code, label) if the cell looks like a section header, else None.

    Rejects stock codes of the form ``NNNNN-XXX-NNN`` which also start with
    five digits but have no spaces around the separator.
    """
    if not cell_value:
        return None
    raw = str(cell_value).strip()
    # Reject stock codes immediately
    if _STOCK_CODE_RE.match(raw):
        return None
    m = _SECTION_HEADER_RE.match(raw)
    if m:
        label = m.group("label").strip()
        # Reject if label looks like the rest of a stock code (e.g. "ADH-002")
        if re.match(r"^[A-Z]{2,5}-\d", label):
            return None
        return m.group("code"), label
    return None


def _detect_columns(ws: openpyxl.worksheet.worksheet.Worksheet,
                    max_header_row: int = 12) -> dict[str, int]:
    """Scan the first *max_header_row* rows to find the header row.

    Returns a mapping of canonical column name → 0-based column index.
    """
    best: dict[str, int] = {}
    best_score = 0

    for r in range(1, max_header_row + 1):
        candidate: dict[str, int] = {}
        score = 0
        for c in range(1, ws.max_column + 1):
            raw = ws.cell(r, c).value
            if raw is None:
                continue
            text = str(raw).strip().lower()
            for col_name, patterns in _COL_PATTERNS.items():
                if col_name in candidate:
                    continue
                if any(p in text for p in patterns):
                    candidate[col_name] = c - 1  # 0-based
                    score += 1
        if score > best_score:
            best_score = score
            best = candidate

    return best


def _cell(ws, row: int, col_map: dict[str, int], name: str):
    idx = col_map.get(name)
    if idx is None:
        return None
    return ws.cell(row, idx + 1).value  # openpyxl is 1-based


def profile_baseline_boq(excel_path: str | Path) -> dict:
    """Read the estimator BOQ at *excel_path* and return a profile dict."""
    path = Path(excel_path)
    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)

    # Find the BOQ sheet — prefer a sheet named "BOQ", otherwise first sheet
    sheet_name = "BOQ" if "BOQ" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    col_map = _detect_columns(ws)

    sections: dict[str, dict] = {}
    current_code: str | None = None
    total_rows = 0

    desc_col = col_map.get("description", 1)   # 0-based fallback
    qty_col  = col_map.get("quantity", 2)
    unit_col = col_map.get("unit", 3)

    for r in range(1, ws.max_row + 1):
        # Read description cell — try the detected column first, then scan
        raw_desc = ws.cell(r, desc_col + 1).value if desc_col is not None else None

        # Detect section header anywhere in the row
        found_header = None
        for c in range(1, min(ws.max_column + 1, 6)):
            found_header = _is_section_header(ws.cell(r, c).value)
            if found_header:
                break

        if found_header:
            code, label = found_header
            current_code = code
            if code not in sections:
                sections[code] = {
                    "label": label,
                    "row_count": 0,
                    "non_empty_rows": 0,
                    "family_counts": Counter(),
                    "units_seen": Counter(),
                    "has_placeholder_rows": False,
                    "has_duplicate_rows": False,
                }
            continue

        if current_code is None:
            continue

        sec = sections[current_code]
        desc = str(raw_desc).strip() if raw_desc else ""
        if not desc or desc.lower() in _DUP_MARKERS:
            if desc.lower() in ("dup", "duplicate", "duplicate row"):
                sec["has_duplicate_rows"] = True
            continue

        raw_qty  = ws.cell(r, qty_col  + 1).value if qty_col  is not None else None
        raw_unit = ws.cell(r, unit_col + 1).value if unit_col is not None else None

        qty  = raw_qty  if raw_qty  is not None else None
        unit = _normalise_unit(raw_unit)

        # Detect placeholder (qty = 0 or None with a description)
        if qty == 0 or qty is None:
            sec["has_placeholder_rows"] = True

        family = classify(desc)
        sec["family_counts"][family] += 1
        if unit:
            sec["units_seen"][unit] += 1
        sec["row_count"] += 1
        if desc:
            sec["non_empty_rows"] += 1
        total_rows += 1

    wb.close()

    # Post-process sections
    result_sections: dict[str, dict] = {}
    for code, sec in sections.items():
        units = dict(sec["units_seen"])
        dominant_unit = max(units, key=units.get) if units else ""
        families_list = sorted(set(sec["family_counts"].keys()))
        family_counts = dict(sec["family_counts"])

        fixing_families = {"screw_fixing", "bolt_fixing", "grommet"}
        fixings_embedded = bool(fixing_families & set(families_list))

        result_sections[code] = {
            "label": sec["label"],
            "row_count": sec["row_count"],
            "non_empty_rows": sec["non_empty_rows"],
            "families": families_list,
            "family_counts": family_counts,
            "units_seen": units,
            "dominant_unit": dominant_unit,
            "style_flags": {
                "uses_stock_length_unit": "len" in units,
                "uses_set_unit": "set" in units or "pair" in units,
                "uses_each": "each" in units,
                "fixings_embedded": fixings_embedded,
                "has_placeholder_rows": sec["has_placeholder_rows"],
                "has_duplicate_rows": sec["has_duplicate_rows"],
                "uses_lm_unit": "lm" in units,
                "uses_m2_unit": "m2" in units,
                "uses_nr_unit": "nr" in units,
                "uses_m3_unit": "m3" in units,
            },
        }

    # Global flags
    all_families = {f for s in result_sections.values() for f in s["families"]}
    codes = set(result_sections.keys())

    global_flags = {
        "fixings_standalone_section": any(
            "fixing" in s["label"].lower() or "fix" in s["label"].lower()
            for s in result_sections.values()
        ),
        "services_placeholder": any(
            "hydraulic" in s["label"].lower() and s["row_count"] == 0
            for s in result_sections.values()
        ),
        "ffe_section_present": any("ffe" in s["label"].lower() or "furniture" in s["label"].lower()
                                   for s in result_sections.values()),
        "stairs_section_present": any("stair" in s["label"].lower()
                                      for s in result_sections.values()),
        "insulation_section_present": any("insulation" in s["label"].lower()
                                          for s in result_sections.values()),
        "electrical_section_present": any("electrical" in s["label"].lower()
                                          for s in result_sections.values()),
        "hydraulics_section_present": any("hydraulic" in s["label"].lower()
                                          for s in result_sections.values()),
    }

    return {
        "sections": result_sections,
        "global_flags": global_flags,
        "column_map": col_map,
        "source_file": str(path),
        "total_rows_parsed": total_rows,
    }
