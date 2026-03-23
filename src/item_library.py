"""
item_library.py — Load the approved BOQ / standard model as a reference library.

This module is a READ-ONLY reference — it does NOT impose the G303 structure
on all projects. It is used by boq_mapper.py for stock-code / description lookup.

The library is keyed by normalised description for fuzzy matching.
Keyword-based fallback matching is provided for common item categories so that
custom project descriptions (e.g. "Wall Frame LGS") can match library entries
(e.g. "Wall Frame 89×38 C-Stud") even when word overlap is low.
"""

from __future__ import annotations
import logging
import re
from pathlib import Path

log = logging.getLogger("boq.item_library")


def load_item_library(
    approved_boq_path: str | None,
    standard_model_path: str | None = None,
) -> dict:
    """
    Load item library from approved BOQ and/or standard model file.

    Priority:
      1. approved_boq_path  (G303-BOQ sheet)
      2. standard_model_path  (Standard BOQ sheet)

    Returns a dict keyed by normalised description:
    {
      "norm_desc": {
        "stock_code": str,
        "description": str,
        "unit": str,
        "section": str,
        "row_idx": int
      }
    }
    Returns empty dict if no files found or on error.
    """
    library: dict = {}

    # Try approved BOQ first
    if approved_boq_path and Path(approved_boq_path).exists():
        _load_from_approved_boq(approved_boq_path, library)
        log.info(
            "Item library loaded from approved BOQ: %d items  (%s)",
            len(library), Path(approved_boq_path).name,
        )
        if library:
            return library

    # Fallback: standard model
    if standard_model_path and Path(standard_model_path).exists():
        _load_from_standard_model(standard_model_path, library)
        log.info(
            "Item library loaded from standard model: %d items  (%s)",
            len(library), Path(standard_model_path).name,
        )
        return library

    log.warning(
        "No item library loaded — approved_boq_path=%s  standard_model_path=%s",
        approved_boq_path, standard_model_path,
    )
    return library


# ─── Loaders ──────────────────────────────────────────────────────────────────

def _load_from_approved_boq(filepath: str, library: dict) -> None:
    """Read G303-BOQ sheet from the approved BOQ file."""
    try:
        import openpyxl
    except ImportError:
        log.warning("openpyxl not installed — item library cannot be loaded")
        return

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    except Exception as exc:
        log.warning("Could not open approved BOQ '%s': %s", filepath, exc)
        return

    sheet_name = "G303-BOQ"
    if sheet_name not in wb.sheetnames:
        # Try first sheet
        sheet_name = wb.sheetnames[0] if wb.sheetnames else None
        if not sheet_name:
            wb.close()
            return
        log.warning("Sheet 'G303-BOQ' not found — using '%s'", sheet_name)

    ws     = wb[sheet_name]
    DATA_START = 9
    section    = "MAIN"

    for row_idx, row in enumerate(
        ws.iter_rows(min_row=DATA_START, values_only=True), DATA_START
    ):
        if len(row) < 2:
            continue
        sc   = row[0]
        desc = row[1]
        qty  = row[2] if len(row) > 2 else None
        unit = row[3] if len(row) > 3 else None

        if sc is None and desc is None:
            continue

        desc_str = str(desc).strip() if desc is not None else ""
        sc_str   = str(sc).strip()   if sc   is not None else ""

        # Section header rows
        if sc is None and qty is None and desc_str:
            section = desc_str
            continue

        if not desc_str:
            continue

        norm = _normalise(desc_str)
        if not norm:
            continue

        library[norm] = {
            "stock_code":  sc_str or None,
            "description": desc_str,
            "unit":        str(unit).strip() if unit is not None else "",
            "section":     section,
            "row_idx":     row_idx,
        }

    wb.close()


def _load_from_standard_model(filepath: str, library: dict) -> None:
    """Read Standard BOQ sheet from the standard model file."""
    try:
        import openpyxl
    except ImportError:
        return

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    except Exception as exc:
        log.warning("Could not open standard model '%s': %s", filepath, exc)
        return

    sheet_name = "Standard BOQ"
    if sheet_name not in wb.sheetnames:
        wb.close()
        return

    ws = wb[sheet_name]
    # Find header row
    header_row = 1
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), 1):
        cells = [str(v).strip().lower() if v is not None else "" for v in row]
        if "description" in cells or "stock code" in cells:
            header_row = i
            break

    headers = None
    desc_idx = 1
    sc_idx   = 0
    unit_idx = 3

    for row_idx, row in enumerate(
        ws.iter_rows(min_row=header_row, values_only=True), header_row
    ):
        if headers is None:
            headers = [str(h or "").strip().lower() for h in row]
            desc_idx = _find_col(headers, ["description", "desc"])
            sc_idx   = _find_col(headers, ["stock code", "stock_code", "code"])
            unit_idx = _find_col(headers, ["unit", "uom"])
            continue

        if not row or all(v is None for v in row):
            continue

        desc = row[desc_idx] if desc_idx < len(row) else None
        sc   = row[sc_idx]   if sc_idx   < len(row) else None
        unit = row[unit_idx] if unit_idx < len(row) else None

        desc_str = str(desc).strip() if desc is not None else ""
        sc_str   = str(sc).strip()   if sc   is not None else ""

        if not desc_str:
            continue

        norm = _normalise(desc_str)
        if not norm:
            continue

        library[norm] = {
            "stock_code":  sc_str or None,
            "description": desc_str,
            "unit":        str(unit).strip() if unit is not None else "",
            "section":     "Standard BOQ",
            "row_idx":     row_idx,
        }

    wb.close()


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _normalise(text: str) -> str:
    """Normalise description for dictionary keying."""
    t = text.lower()
    t = re.sub(r"[^\w\s]", " ", t)
    t = re.sub(r"\s+", " ", t).strip()
    return t


def _find_col(headers: list[str], candidates: list[str]) -> int:
    for c in candidates:
        if c in headers:
            return headers.index(c)
    return 0


# ── Keyword-based category matcher (for custom project matching) ───────────────

# Maps category keywords found in a candidate description to keywords that
# should appear in a library entry description for it to be considered a match.
_KEYWORD_CATEGORY_MAP: list[tuple[list[str], list[str]]] = [
    # (candidate keywords, library keywords)
    (["floor area", "floor_area", "fc floor", "flooring"],
     ["floor", "flooring", "fc floor", "vinyl", "ceramic", "tile"]),
    (["wall frame", "wall_frame", "lgs wall", "c-stud", "c stud"],
     ["wall frame", "wall stud", "c-stud", "lgs"]),
    (["roof batten", "roofing batten", "rfg batten"],
     ["roof batten", "roofing batten", "batten"]),
    (["ceiling batten", "ceil batten", "clg batten"],
     ["ceiling batten", "ceil batten"]),  # no bare "batten" — avoids matching roof batten
    (["roof truss", "truss"],
     ["truss", "roof truss"]),
    (["floor panel", "floor joist"],
     ["floor panel", "floor joist", "floor frame"]),
    (["roof area", "roof sheet", "corrugated"],
     ["roof sheet", "corrugated", "roofing"]),
    # FC sheets must come BEFORE the generic "ceiling" entry to avoid
    # "FC Sheet — Ceiling" matching "Ceiling Batten" instead of an FC sheet item.
    (["fc sheet", "fc wall", "fc ceiling", "fibre cement", "hardiflex"],
     ["fc sheet", "fc wall", "fibre cement", "hardiflex"]),
    # Generic ceiling — triggered only by specific ceiling keywords, NOT bare "ceiling"
    # (bare "ceiling" is too broad and would match ceiling battens for FC sheet items)
    (["plasterboard ceiling", "ceiling lining", "ceiling board"],
     ["ceiling", "fc ceiling", "plasterboard"]),
    (["door", "timber door"],
     ["door", "timber door", "solid core", "hollow core"]),
    (["window", "timber window", "louvre window"],
     ["window", "timber window", "louvre"]),
    (["stair", "staircase"],
     ["stair", "staircase", "stringer"]),
    (["external wall", "ext wall"],
     ["weatherboard", "external wall", "cladding", "fc sheet"]),
    (["verandah", "deck", "balcony"],
     ["verandah", "deck", "decking", "wpc"]),
]


def find_by_keyword_category(
    candidate_desc: str,
    library: dict,
) -> tuple[dict | None, str | None]:
    """
    Find a library entry by keyword category matching.

    Used when exact/fuzzy word-overlap matching fails.
    Maps known construction item category keywords to library entry keywords.
    Returns (entry, norm_key) or (None, None).
    """
    cand_lower = candidate_desc.lower()

    for cand_keywords, lib_keywords in _KEYWORD_CATEGORY_MAP:
        # Check if any candidate keyword appears in the candidate description
        if not any(kw in cand_lower for kw in cand_keywords):
            continue
        # Find a library entry whose description contains any lib keyword
        for norm_key, entry in library.items():
            entry_lower = norm_key.lower()
            if any(lkw in entry_lower for lkw in lib_keywords):
                log.debug(
                    "Keyword category match: '%s' → '%s'",
                    candidate_desc[:50], entry.get("description", "")[:50],
                )
                return entry, norm_key
        break  # matched category but no library entry — stop searching

    return None, None
