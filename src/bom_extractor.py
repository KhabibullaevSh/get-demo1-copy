"""
bom_extractor.py — Read Framecad BOM (Excel/CSV) and IFC files.

BOM/IFC data is highest priority for structural framing items.
"""

from __future__ import annotations
import csv
import logging
import re
from pathlib import Path
from typing import Any

from src.utils import normalise_text, safe_float

log = logging.getLogger("boq.bom_extractor")

# Keywords to map BOM rows into structural categories
_CATEGORY_MAP = {
    "wall_frame": ["wall frame", "wall stud", "wall track", "wall top", "wall bottom",
                   "wall member", "c-section", "c section"],
    "roof_truss": ["roof truss", "truss", "rafter"],
    "floor_joist": ["floor joist", "joist", "bearer", "floor member", "floor frame"],
    "floor_panel": ["floor panel", "panel"],
    "ceiling_batten": ["ceiling batten", "ceil batten", "clg batten"],
    "roof_batten": ["roof batten", "rfg batten", "roofing batten"],
    "bracing": ["bracing", "brace", "diagonal"],
    "connection": ["screw", "bolt", "nail", "bracket", "connector", "clip",
                   "strap", "plate", "tek screw"],
    "insulation": ["insulation", "batts", "blanket"],
    "door": ["door"],
    "window": ["window"],
}


def extract_bom(bom_files: list[dict]) -> dict[str, Any]:
    """Extract and normalise all BOM/IFC files.

    Returns::

        {
          "raw_items": [...],
          "normalized": {
            "wall_frame_lm": ...,
            "roof_truss_qty": ...,
            "floor_panels": [...],
            "ceiling_batten_lm": ...,
            "roof_batten_lm": ...,
            ...
          },
          "warnings": [...]
        }
    """
    result: dict[str, Any] = {
        "raw_items": [],
        "normalized": {},
        "warnings": [],
    }

    if not bom_files:
        return result

    for entry in bom_files:
        path = Path(entry["path"])
        ext = path.suffix.lower()
        try:
            if ext in (".xlsx", ".xls"):
                items = _read_excel_bom(path)
            elif ext == ".csv":
                items = _read_csv_bom(path)
            elif ext == ".ifc":
                items = _read_ifc(path, result)
                result["raw_items"].extend(items)
                continue
            else:
                result["warnings"].append(f"Unsupported BOM format: {path.name}")
                continue

            result["raw_items"].extend(items)
            log.info("BOM %s: %d raw items", path.name, len(items))
        except Exception as exc:
            result["warnings"].append(f"BOM read error {path.name}: {exc}")
            log.warning("BOM read error %s: %s", path.name, exc)

    _normalise(result)
    return result


# ─── Excel reader ─────────────────────────────────────────────────────────────

def _read_excel_bom(path: Path) -> list[dict]:
    try:
        import openpyxl
    except ImportError:
        log.warning("openpyxl not installed")
        return []

    wb = openpyxl.load_workbook(str(path), data_only=True)
    items = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        # Find header row (contains "description" or "qty" or "length")
        header_idx = _find_header(rows)
        if header_idx is None:
            continue
        headers = [str(h).strip().lower() if h else "" for h in rows[header_idx]]
        for row in rows[header_idx + 1:]:
            if not row or all(v is None for v in row):
                continue
            item = {}
            for i, h in enumerate(headers):
                if i < len(row) and h:
                    item[h] = row[i]
            if item:
                item["_sheet"] = sheet_name
                item["_source"] = path.name
                items.append(_normalise_bom_row(item))
    wb.close()
    return items


def _find_header(rows: list) -> int | None:
    header_kw = {"description", "desc", "qty", "quantity", "length", "item", "mark", "part"}
    for i, row in enumerate(rows[:20]):
        cells = {str(v).strip().lower() for v in row if v is not None}
        if cells & header_kw:
            return i
    return None


def _read_csv_bom(path: Path) -> list[dict]:
    items = []
    try:
        with open(path, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                norm = _normalise_bom_row({k.lower().strip(): v for k, v in row.items()})
                norm["_source"] = path.name
                items.append(norm)
    except Exception as exc:
        log.warning("CSV BOM read error %s: %s", path.name, exc)
    return items


def _normalise_bom_row(row: dict) -> dict:
    """Map various column names to canonical keys."""
    desc = (
        row.get("description") or row.get("desc") or
        row.get("item") or row.get("member") or ""
    )
    qty_raw = (
        row.get("qty") or row.get("quantity") or row.get("count") or
        row.get("no") or row.get("number") or None
    )
    length_raw = (
        row.get("length") or row.get("len") or row.get("length_mm") or None
    )
    return {
        "description": str(desc).strip() if desc else "",
        "qty": safe_float(qty_raw),
        "length_mm": safe_float(length_raw),
        "unit": str(row.get("unit", "") or "").strip(),
        "mark": str(row.get("mark") or row.get("part") or "").strip(),
        "category": _categorise(str(desc)),
        "_sheet": row.get("_sheet", ""),
        "_source": row.get("_source", ""),
    }


def _categorise(description: str) -> str:
    desc_lower = normalise_text(description)
    for cat, keywords in _CATEGORY_MAP.items():
        if any(kw in desc_lower for kw in keywords):
            return cat
    return "other"


# ─── IFC reader ───────────────────────────────────────────────────────────────

def _read_ifc(path: Path, result: dict) -> list[dict]:
    """Use ifc_extractor for full property/quantity extraction."""
    try:
        from src.ifc_extractor import extract_ifc
    except ImportError:
        result["warnings"].append("ifc_extractor module not found — IFC extraction skipped")
        return []

    ifc_result = extract_ifc(path)
    result["warnings"].extend(ifc_result.get("warnings", []))

    # Store geometry for merger to pick up
    if ifc_result.get("geometry"):
        result["ifc_geometry"] = ifc_result["geometry"]

    # Store structural for _normalise to merge in
    if ifc_result.get("structural"):
        result["ifc_structural"] = ifc_result["structural"]

    # Return empty raw_items — quantities flow via ifc_structural/ifc_geometry
    return []


# ─── Normalise aggregates ─────────────────────────────────────────────────────

def _normalise(result: dict) -> None:
    """Aggregate raw items into normalised summary values."""
    items = result["raw_items"]
    n = result["normalized"]

    def total_length(cat: str) -> float:
        return sum(
            (i["qty"] or 1) * (i["length_mm"] or 0)
            for i in items if i.get("category") == cat
        )

    def total_qty(cat: str) -> float:
        return sum(i["qty"] or 1 for i in items if i.get("category") == cat)

    n["wall_frame_lm"] = round(total_length("wall_frame") / 1000, 2)
    n["ceiling_batten_lm"] = round(total_length("ceiling_batten") / 1000, 2)
    n["roof_batten_lm"] = round(total_length("roof_batten") / 1000, 2)
    n["floor_joist_lm"] = round(total_length("floor_joist") / 1000, 2)
    n["bracing_lm"] = round(total_length("bracing") / 1000, 2)
    n["roof_truss_qty"] = round(total_qty("roof_truss"))
    n["floor_panel_qty"] = round(total_qty("floor_panel"))
    n["connection_qty"] = round(total_qty("connection"))
    n["door_qty"] = round(total_qty("door"))
    n["window_qty"] = round(total_qty("window"))

    # Floor panels detail
    panels = [i for i in items if i.get("category") == "floor_panel"]
    n["floor_panels"] = panels

    # Fill in from IFC structural where BOM raw items gave nothing
    ifc_struct = result.get("ifc_structural", {})
    if ifc_struct:
        for key in ("wall_frame_lm", "ceiling_batten_lm", "roof_batten_lm",
                    "verandah_batten_lm", "floor_joist_lm", "bracing_lm"):
            if n.get(key, 0) == 0 and ifc_struct.get(key, 0) > 0:
                n[key] = ifc_struct[key]
        for key in ("roof_truss_qty", "floor_panel_qty", "post_qty"):
            if n.get(key, 0) == 0 and ifc_struct.get(key, 0) > 0:
                n[key] = ifc_struct[key]
        # Propagate source flags
        for src_key in ("wall_frame_source", "roof_truss_source", "floor_panel_source"):
            if src_key in ifc_struct:
                n[src_key] = ifc_struct[src_key]

    log.info(
        "BOM normalized: wall=%.1flm  ceil_batten=%.1flm  roof_batten=%.1flm  "
        "trusses=%d  floor_panels=%d",
        n["wall_frame_lm"], n["ceiling_batten_lm"], n["roof_batten_lm"],
        n["roof_truss_qty"], n["floor_panel_qty"],
    )
