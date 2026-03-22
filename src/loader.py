"""
loader.py — Loads standard_model_G303.xlsx into memory.

Reads three sheets:
  1. Standard BOQ — all items with stock codes, descriptions, units, standard qty, rate
  2. Standard Geometry — dimensions of standard G303
  3. Rules Library — scaling rules for dependent items

Returns structured dicts/lists for downstream pipeline steps.
"""

import os
from typing import Any
import openpyxl


def load_standard_model(filepath: str) -> dict[str, Any]:
    """Load the standard model Excel workbook and return parsed data."""
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"Standard model file not found: {filepath}")

    wb = openpyxl.load_workbook(filepath, data_only=True)

    required_sheets = ["Standard BOQ", "Standard Geometry", "Rules Library"]
    missing = [s for s in required_sheets if s not in wb.sheetnames]
    if missing:
        raise ValueError(f"Missing required sheets in standard model: {missing}")

    standard_boq = _load_standard_boq(wb["Standard BOQ"])
    standard_geometry = _load_standard_geometry(wb["Standard Geometry"])
    rules = _load_rules_library(wb["Rules Library"])

    wb.close()

    return {
        "standard_boq": standard_boq,
        "standard_geometry": standard_geometry,
        "rules": rules,
    }


def _find_header_row(ws, key_indicators: list[str]) -> int:
    """Scan worksheet to find the row index (1-based) of the actual header row.

    Uses exact cell-value matching (strip + lowercase) against the key_indicators.
    Returns 1 if not found (fall back to first row).
    """
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), 1):
        cells = [str(v).strip().lower() if v is not None else "" for v in row]
        if any(cell == indicator for indicator in key_indicators for cell in cells if cell):
            return i
    return 1


def _load_standard_boq(ws) -> list[dict]:
    """Parse the Standard BOQ sheet into a list of item dicts."""
    # Find actual header row (contains 'description' or 'stock code')
    header_row_idx = _find_header_row(ws, ["description", "stock code", "standard qty"])

    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not rows or header_row_idx > len(rows):
        return []

    headers = [str(h).strip().lower() if h else "" for h in rows[header_row_idx - 1]]
    items = []
    item_counter = 1

    for row in rows[header_row_idx:]:
        if not row or all(cell is None for cell in row):
            continue
        # Skip category separator rows (only first cell has content, rest are None)
        non_none = [v for v in row if v is not None]
        if len(non_none) <= 1:
            continue

        item = {}
        for i, header in enumerate(headers):
            if i < len(row):
                item[header] = row[i]

        item = _normalise_boq_keys(item)

        # Require description to be non-empty
        if not item.get("description"):
            continue

        # Auto-assign item number if missing
        if not item.get("item_no"):
            item["item_no"] = item_counter
        item_counter += 1

        items.append(item)

    return items


def _normalise_boq_keys(item: dict) -> dict:
    """Map various possible column names to canonical keys."""
    key_map = {
        "item no": "item_no",
        "item_no": "item_no",
        "item number": "item_no",
        "no": "item_no",
        "#": "item_no",
        "stock code": "stock_code",
        "stock_code": "stock_code",
        "code": "stock_code",
        "description": "description",
        "desc": "description",
        "unit": "unit",
        "uom": "unit",
        "qty": "qty",
        "quantity": "qty",
        "standard qty": "qty",
        "standard_qty": "qty",
        "rate": "rate",
        "rate (pgk)": "rate",
        "rate_pgk": "rate",
        "amount": "amount",
        "amount (pgk)": "amount",
        "category": "category",
        "section": "category",
        "notes": "notes",
        "confidence": "confidence",
        "source": "source",
    }
    normalised = {}
    for k, v in item.items():
        canonical = key_map.get(str(k).strip().lower(), str(k).strip().lower())
        normalised[canonical] = v
    return normalised


def _load_standard_geometry(ws) -> dict:
    """Parse the Standard Geometry sheet into a dict of element → value.

    Expected format: Element | Description | Value | Unit | Notes | Source
    """
    # Find header row containing 'element' and 'value'
    header_row_idx = _find_header_row(ws, ["element", "value"])

    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not rows or header_row_idx > len(rows):
        return {}

    headers = [str(h).strip().lower() if h else "" for h in rows[header_row_idx - 1]]

    # Find the index of the 'value' column
    try:
        value_col_idx = headers.index("value")
    except ValueError:
        value_col_idx = 2  # Default to third column

    raw_geometry = {}

    for row in rows[header_row_idx:]:
        if not row or row[0] is None:
            continue
        # Skip category separator rows (only element name, rest None)
        if all(cell is None for cell in row[1:]):
            continue

        key = str(row[0]).strip().lower().replace(" ", "_").replace("/", "_").replace(".", "")
        # Remove brackets/parens
        key = key.replace("(", "").replace(")", "").replace(",", "")
        # Collapse multiple underscores
        while "__" in key:
            key = key.replace("__", "_")
        key = key.strip("_")

        value = row[value_col_idx] if value_col_idx < len(row) else None
        try:
            value = float(value) if value is not None else 0.0
        except (ValueError, TypeError):
            value = 0.0

        raw_geometry[key] = value

    return _normalise_geometry_keys(raw_geometry)


def _normalise_geometry_keys(raw: dict) -> dict:
    """Map raw geometry keys to canonical names expected by change_detector.

    Also computes derived aggregate values.
    """
    # Direct renames
    rename_map = {
        "total_floor_area": "total_floor_area",
        "internal_floor_area": "internal_floor_area",
        "verandah_area": "verandah_area",
        "roof_surface_area": "roof_area",
        "roof_plan_area": "roof_plan_area",
        "gutter_length": "roof_perimeter",
        "total_external_wall_length": "external_wall_length",
        "total_internal_wall_length": "internal_wall_length",
        "external_wall_area": "external_wall_area",
        "internal_wall_area": "internal_wall_area",
        "no_of_piers_posts": "post_count",
        "no_of_piers__posts": "post_count",
        "no_of_stair_flights": "stair_count",
        "no_of_stair_steps": "stair_steps",
        "no_of_downpipes": "downpipe_count",
        "floor_panel_count": "floor_panel_count",
        "ridge_length": "ridge_length",
        "barge_length": "barge_length",
        "fascia_length": "fascia_length",
        "balustrade_length": "balustrade_length",
        "building_length": "building_length",
        "building_width": "building_width",
        "wall_height": "wall_height",
        "floor_height": "floor_height",
        "roof_pitch": "roof_pitch",
    }

    geometry = {}
    for raw_key, canon_key in rename_map.items():
        if raw_key in raw:
            geometry[canon_key] = raw[raw_key]

    # Copy remaining keys that weren't renamed
    for k, v in raw.items():
        if k not in rename_map and k not in geometry:
            geometry[k] = v

    # Compute total_wall_length
    ext_wall = geometry.get("external_wall_length", 0.0)
    int_wall = geometry.get("internal_wall_length", 0.0)
    if ext_wall or int_wall:
        geometry["total_wall_length"] = round(ext_wall + int_wall, 2)

    # Ceiling area ≈ internal floor area (standard approximation)
    if "internal_floor_area" in geometry and "ceiling_area" not in geometry:
        geometry["ceiling_area"] = geometry["internal_floor_area"]

    # Aggregate door counts
    door_keys = [k for k in raw if k.startswith("door_") and k.endswith("_count")]
    total_doors = sum(int(raw.get(k, 0) or 0) for k in door_keys)
    if total_doors > 0:
        geometry["door_count"] = total_doors
        # Store individual door types for type-level comparison
        geometry["door_types"] = {
            k.replace("_count", "").replace("door_", "Door ").title(): int(raw.get(k, 0) or 0)
            for k in door_keys if int(raw.get(k, 0) or 0) > 0
        }

    # Aggregate window counts
    window_keys = [k for k in raw if k.startswith("window_") and k.endswith("_count")]
    total_windows = sum(int(raw.get(k, 0) or 0) for k in window_keys)
    if total_windows > 0:
        geometry["window_count"] = total_windows
        geometry["window_types"] = {
            k.replace("_count", "").replace("window_", "Window ").title(): int(raw.get(k, 0) or 0)
            for k in window_keys if int(raw.get(k, 0) or 0) > 0
        }

    # Room count: count rooms with area > 0
    room_keys = [
        "bedroom_1", "bedroom_2", "bedroom_3", "bathroom", "toilet",
        "kitchen", "living__dining", "laundry", "storage", "corridor__hallway",
    ]
    room_count = sum(1 for k in room_keys if raw.get(k, 0))
    if room_count > 0:
        geometry["room_count"] = room_count

    return geometry


# Mapping from rules sheet "Driven By" values (normalised) to canonical geometry keys
_RULE_DEPENDS_ON_MAP = {
    "roof_surface_area_m²": "roof_area",
    "roof_surface_area_m": "roof_area",
    "ceiling_area_m²": "ceiling_area",
    "ceiling_area_m": "ceiling_area",
    "external_wall_area_m²": "external_wall_area",
    "external_wall_area_m": "external_wall_area",
    "internal_wall_area_m²": "internal_wall_area",
    "internal_wall_area_m": "internal_wall_area",
    "floor_area_m²": "total_floor_area",
    "floor_area_m": "total_floor_area",
    "vinyl_floor_area_m²": "total_floor_area",
    "tile_floor_area_m²": "total_floor_area",
    "tile_area_m²": "total_floor_area",
    "gutter_length_lm": "roof_perimeter",
    "ridge_length_lm": "ridge_length",
    "barge_length_lm": "barge_length",
    "fascia_length_lm": "fascia_length",
    "ridge_+_barge_length_lm": "roof_perimeter",
}


def _load_rules_library(ws) -> list[dict]:
    """Parse the Rules Library sheet into a list of rule dicts."""
    # Find header row (contains 'rule id' or 'item')
    header_row_idx = _find_header_row(ws, ["rule id", "item (dependent)", "formula"])

    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not rows or header_row_idx > len(rows):
        return []

    headers = [str(h).strip().lower() if h else "" for h in rows[header_row_idx - 1]]
    rules = []

    # Column name normalization for rules sheet
    rules_key_map = {
        "rule id": "rule_id",
        "rule_id": "rule_id",
        "item (dependent)": "target_item",
        "item_(dependent)": "target_item",
        "target_item": "target_item",
        "target item": "target_item",
        "driven by (bulk item)": "depends_on",
        "driven_by_(bulk_item)": "depends_on",
        "depends_on": "depends_on",
        "depends on": "depends_on",
        "formula / rule": "formula",
        "formula_/_rule": "formula",
        "formula": "formula",
        "unit": "unit",
        "example": "example",
        "notes / source": "notes",
        "notes_/_source": "notes",
        "notes": "notes",
        "multiplier": "multiplier",
        "description": "description",
    }

    for row in rows[header_row_idx:]:
        if not row or all(cell is None for cell in row):
            continue
        # Skip category separator rows
        non_none = [v for v in row if v is not None]
        if len(non_none) <= 1:
            continue

        rule = {}
        for i, header in enumerate(headers):
            if i < len(row):
                canon = rules_key_map.get(header, header.replace(" ", "_"))
                rule[canon] = row[i]

        if rule.get("rule_id") or rule.get("target_item"):
            # Normalise depends_on to geometry key format
            dep = str(rule.get("depends_on", "")).strip()
            dep_norm = dep.lower().replace(" ", "_").replace("(", "").replace(")", "").replace("/", "_")
            while "__" in dep_norm:
                dep_norm = dep_norm.replace("__", "_")
            # Map rule depends_on to canonical geometry key names
            dep_norm = _RULE_DEPENDS_ON_MAP.get(dep_norm, dep_norm)
            rule["depends_on"] = dep_norm
            rules.append(rule)

    return rules


def load_rate_library(filepath: str) -> list[dict]:
    """Load the rate library Excel file."""
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"Rate library file not found: {filepath}")

    wb = openpyxl.load_workbook(filepath, data_only=True)

    # Try to use "Rate Library" sheet, otherwise use active sheet
    if "Rate Library" in wb.sheetnames:
        ws = wb["Rate Library"]
    else:
        ws = wb.active

    # Find actual header row (contains 'description' and 'rate')
    header_row_idx = _find_header_row(ws, ["description", "rate (pgk)", "rate"])

    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not rows or header_row_idx > len(rows):
        wb.close()
        return []

    headers = [str(h).strip().lower() if h else "" for h in rows[header_row_idx - 1]]
    rates = []

    key_map = {
        "#": "seq_no",
        "stock code": "stock_code",
        "stock_code": "stock_code",
        "code": "stock_code",
        "description": "description",
        "desc": "description",
        "unit": "unit",
        "uom": "unit",
        "rate": "rate",
        "rate (pgk)": "rate",
        "rate_pgk": "rate",
        "category": "category",
        "notes": "notes",
        "source sheet": "source_sheet",
    }

    for row in rows[header_row_idx:]:
        if not row or all(cell is None for cell in row):
            continue
        # Skip category separator rows (description column is None or not a real item)
        # In rate library, description is at column index 2 typically
        desc_idx = next((i for i, h in enumerate(headers) if h == "description"), None)
        if desc_idx is not None and (desc_idx >= len(row) or row[desc_idx] is None):
            continue

        item = {}
        for i, header in enumerate(headers):
            if i < len(row):
                canonical = key_map.get(header, header.replace(" ", "_"))
                item[canonical] = row[i]

        if item.get("description") and item.get("rate"):
            # Validate rate is numeric
            try:
                item["rate"] = float(item["rate"])
                rates.append(item)
            except (ValueError, TypeError):
                pass

    wb.close()
    return rates


def load_structural_baseline(filepath: str) -> list[dict]:
    """Load the Structural Baseline sheet from the complete standard model.

    Returns a list of dicts with keys: package, component, qty, unit, source, notes.
    """
    if not os.path.exists(filepath):
        return []

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception:
        return []

    if "Structural Baseline" not in wb.sheetnames:
        wb.close()
        return []

    ws = wb["Structural Baseline"]
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not rows:
        wb.close()
        return []

    # First row is header
    headers = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
    key_map = {
        "package": "package",
        "component": "component",
        "standard qty": "qty",
        "qty": "qty",
        "unit": "unit",
        "source": "source",
        "notes": "notes",
    }

    baseline = []
    for row in rows[1:]:
        if not row or all(v is None for v in row):
            continue
        item = {}
        for i, hdr in enumerate(headers):
            if i < len(row):
                canon = key_map.get(hdr, hdr)
                item[canon] = row[i]
        if item.get("component") and item.get("qty") is not None:
            try:
                item["qty"] = float(item["qty"])
            except (TypeError, ValueError):
                continue
            baseline.append(item)

    wb.close()
    return baseline


def load_approved_boq(filepath: str, sheet_name: str = "G303-BOQ") -> list[dict]:
    """Load item list from the approved project BOQ (G303-BOQ sheet).

    Returns list of item dicts with keys:
        stock_code, description, qty, unit, _row_idx, _section, _window_context
    """
    import logging
    log = logging.getLogger("boq.loader")

    if not os.path.exists(filepath):
        raise FileNotFoundError(f"Approved BOQ file not found: {filepath}")

    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in {filepath}")

    ws = wb[sheet_name]
    items: list[dict] = []
    DATA_START_ROW = 9   # rows 1-8 are title / column headers

    section = "FIRST_FLOOR"
    window_context = None   # "A", "B", "D", or None

    for row_idx, row in enumerate(ws.iter_rows(min_row=DATA_START_ROW, values_only=True), DATA_START_ROW):
        if len(row) < 2:
            continue
        sc   = row[0]
        desc = row[1]
        qty  = row[2] if len(row) > 2 else None
        unit = row[3] if len(row) > 3 else None

        if sc is None and desc is None:
            continue

        desc_str = str(desc).strip() if desc is not None else ""
        sc_str   = str(sc).strip()   if sc   is not None else ""
        desc_lower = desc_str.lower()

        # Section-header rows: no stock code, no qty
        if sc is None and qty is None:
            if desc_str:
                if any(k in desc_lower for k in ["laundry", "ground level", "ground floor"]):
                    section = "LAUNDRY"
                    window_context = None
                # other section headers don't change section
            continue

        qty_val = None
        if qty is not None:
            try:
                qty_val = float(qty)
            except (TypeError, ValueError):
                qty_val = None

        # Update window type context from stock-code parent rows
        if sc_str:
            dl = desc_lower
            if "timber window - a" in dl or "timber window-a" in dl:
                window_context = "A"
            elif "timber window - b" in dl or "timber window-b" in dl:
                window_context = "B"
            elif "timber window - c" in dl or "timber window-c" in dl:
                window_context = "C"
            elif "timber window - d" in dl or "timber window-d" in dl:
                window_context = "D"
            elif not any(k in dl for k in ["timber window", "timber frame window"]):
                window_context = None

        items.append({
            "stock_code":       sc_str or None,
            "description":      desc_str,
            "qty":              qty_val,
            "unit":             str(unit).strip() if unit is not None else None,
            "_row_idx":         row_idx,
            "_section":         section,
            "_window_context":  window_context,
        })

    wb.close()
    log.info("Approved BOQ loaded: %d items from '%s'", len(items), sheet_name)
    return items
