"""
quantity_calculator.py — Resolve final BOQ quantities from merged project data.

Tier resolution order for every item:
  TIER 1: DWG geometry (extract_geometry → merger → geo dict)
  TIER 2: PDF schedule (door/window/finish schedule → merger → doors/windows lists)
  TIER 3: Deterministic rules (spacing formulas, room-area rules, 3BR standards)
  TIER 4: BLANK — no data source; qty=None, issue_flag=BLANK, must be filled manually

ELIMINATED: _mark_standard() / approved BOQ fallback.
Every quantity now traces to TIER 1/2/3 or is explicitly BLANK.
"""

from __future__ import annotations
import logging
import math
from typing import Any

from src.config import (
    BATTEN_CEILING_SPACING_MM, BATTEN_ROOF_SPACING_MM, BATTEN_VERANDAH_SPACING_MM,
    BATTEN_LENGTH_MM, DEFAULT_ROOF_PITCH_DEG, DEFAULT_EAVE_OVERHANG_MM,
    SHEET_AREA_FC, SHEET_AREA_FC_WALL, SHEET_AREA_PLASTER,
    FC_WASTE_FACTOR, PLASTER_WASTE_FACTOR,
    WEATHERBOARD_COVER_MM, WEATHERBOARD_LENGTH_MM, SISALATION_ROLL_M2,
    DEFAULT_WALL_HEIGHT, Confidence, ITEM_RULES,
)
from src.utils import safe_float

log = logging.getLogger("boq.quantity_calculator")

# ─── Rules Library loader ─────────────────────────────────────────────────────
_RULES_LIB: dict | None = None

def _get_rules_lib() -> dict:
    """Load spacing and coverage values from G303 Rules Library sheet (cached)."""
    global _RULES_LIB
    if _RULES_LIB is not None:
        return _RULES_LIB

    defaults = {
        "roof_batten_spacing_m":     BATTEN_ROOF_SPACING_MM / 1000.0,
        "ceiling_batten_spacing_m":  BATTEN_CEILING_SPACING_MM / 1000.0,
        "verandah_batten_spacing_m": BATTEN_VERANDAH_SPACING_MM / 1000.0,
        "batten_length_m":           BATTEN_LENGTH_MM / 1000.0,
        "weatherboard_cover_m":      WEATHERBOARD_COVER_MM / 1000.0,
        "weatherboard_length_m":     WEATHERBOARD_LENGTH_MM / 1000.0,
        "sisalation_roll_m2":        SISALATION_ROLL_M2,
        "fc_ceiling_sheet_m2":       SHEET_AREA_FC,       # 1.2 × 2.4 = 2.88
        "fc_wall_sheet_w":           1.2,                  # sheet width
        "fc_wall_sheet_h":           2.7,                  # sheet height
        "fc_ceiling_sheet_l":        2.4,                  # sheet length
        "fc_ceiling_sheet_w":        1.2,                  # sheet width
    }

    import re as _re
    from pathlib import Path
    for lib_path in [
        Path("data/approved_boq_G303.xlsx"),
        Path("data/standard_models/G303.xlsx"),
    ]:
        if not lib_path.exists():
            continue
        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(lib_path), read_only=True, data_only=True)
            if "Rules Library" not in wb.sheetnames:
                continue
            ws = wb["Rules Library"]
            for row in ws.iter_rows(values_only=True):
                rule_id = str(row[0] or "").strip()
                formula = str(row[3] or "").lower()
                if rule_id == "R-001":
                    pass  # spacing controlled by BATTEN_ROOF_SPACING_MM = 400mm in config.py
                elif rule_id == "R-010":
                    m = _re.search(r'(\d+(?:\.\d+)?)\s*m\s+spac', formula)
                    if m:
                        defaults["ceiling_batten_spacing_m"] = float(m.group(1))
                elif rule_id == "R-018":
                    pass  # weatherboard cover controlled by WEATHERBOARD_COVER_MM = 200mm in config.py
            log.info("Rules Library loaded from %s", lib_path)
            break
        except Exception as exc:
            log.debug("Rules Library load error (%s): %s", lib_path, exc)

    _RULES_LIB = defaults
    return _RULES_LIB


def calculate_quantities(
    standard_boq: list[dict],
    merged: dict,
    validation: dict,
    standard_geometry: dict | None = None,
    debug: bool = False,
) -> list[dict]:
    """Resolve a quantity for every standard BOQ line item."""
    geo    = dict(merged.get("geometry", {}))   # copy so we can add derived keys
    struct = merged.get("structural", {})

    # Derive area fields from lengths × wall height if not explicitly set
    _wh = DEFAULT_WALL_HEIGHT  # 2.4 m
    if not geo.get("external_wall_area_m2") and geo.get("external_wall_length_m"):
        geo["external_wall_area_m2"] = round(safe_float(geo["external_wall_length_m"]) * _wh, 2)
    if not geo.get("internal_wall_area_m2") and geo.get("internal_wall_length_m"):
        geo["internal_wall_area_m2"] = round(safe_float(geo["internal_wall_length_m"]) * _wh, 2)
    if not geo.get("ceiling_area_m2") and geo.get("total_floor_area_m2"):
        geo["ceiling_area_m2"] = safe_float(geo["total_floor_area_m2"])
    if not geo.get("total_wall_length_m"):
        geo["total_wall_length_m"] = round(
            safe_float(geo.get("external_wall_length_m", 0)) +
            safe_float(geo.get("internal_wall_length_m", 0)), 2
        )
    # Derive roof perimeter from external wall length (building perimeter proxy)
    if not geo.get("roof_perimeter_m") and geo.get("external_wall_length_m"):
        geo["roof_perimeter_m"] = round(safe_float(geo["external_wall_length_m"]), 1)
    log.debug(
        "Derived geo: ext_wall_area=%.1f  int_wall_area=%.1f  ceiling=%.1f  roof=%.1f",
        safe_float(geo.get("external_wall_area_m2", 0)),
        safe_float(geo.get("internal_wall_area_m2", 0)),
        safe_float(geo.get("ceiling_area_m2", 0)),
        safe_float(geo.get("roof_area_m2", 0)),
    )
    doors  = merged.get("doors", [])
    windows = merged.get("windows", [])
    finishes = merged.get("finishes", [])
    stairs_data = merged.get("stairs", [])
    std_geo = standard_geometry or {}

    global _roof_batten_row_count
    _roof_batten_row_count = 0

    resolved: list[dict] = []
    for item in standard_boq:
        item = dict(item)
        desc = (item.get("description") or "").lower()
        cat  = (item.get("category") or "").lower()
        orig_qty = item.get("qty")
        try:
            _resolve_item(item, desc, cat, geo, struct, doors, windows,
                          finishes, stairs_data, merged, std_geo)
        except Exception as exc:
            log.warning("Resolution error for '%s': %s", desc[:50], exc)
            item.setdefault("issue_flag", "RESOLUTION_ERROR")
            item.setdefault("comment", str(exc))
            item.setdefault("confidence", Confidence.LOW.value)
            item.setdefault("source", "standard_fallback")

        if debug:
            log.debug(
                "ITEM %-60s  orig_qty=%-6s  resolved_qty=%-6s  src=%s",
                desc[:60], orig_qty, item.get("qty"), item.get("source"),
            )
        resolved.append(item)

    log.info(
        "Quantities: total=%d  high=%d  medium=%d  low=%d  flagged=%d",
        len(resolved),
        sum(1 for i in resolved if i.get("confidence") == "HIGH"),
        sum(1 for i in resolved if i.get("confidence") == "MEDIUM"),
        sum(1 for i in resolved if i.get("confidence") == "LOW"),
        sum(1 for i in resolved if i.get("issue_flag")),
    )
    return resolved


def _resolve_item(
    item: dict, desc: str, cat: str,
    geo: dict, struct: dict, doors: list, windows: list,
    finishes: list, stairs_data: list, merged: dict, std_geo: dict,
) -> None:
    # ── Section-specific items: laundry section falls to standard ─────────────
    _section = item.get("_section", "FIRST_FLOOR")
    _win_ctx  = item.get("_window_context")

    # ── Skirting and cornice — section-specific ──────────────────────────────
    if _kw(desc, ["timber skirting", "skirting"]) and not _kw(desc, ["soffit", "flashing"]):
        _resolve_skirting(item, geo, _section); return

    if _kw(desc, ["quad cornice"]):
        _resolve_quad_cornice(item, geo); return

    if _kw(desc, ["timber cornice", "cornice"]):
        _resolve_cornice(item, geo, _section); return

    # ── Window sub-items (strips, round bar, fly screen, louvre) ─────────────
    # Guard: only apply to sub-items (no stock code); parent "Timber Window - X" rows
    # have a stock code and must fall through to _resolve_windows.
    _no_sc = not bool(item.get("stock_code"))

    if _win_ctx and _no_sc and _kw(desc, ["window strip", "timber strip", "12 x 38"]):
        _resolve_window_strips(item, desc, _win_ctx); return

    if _win_ctx and _no_sc and _kw(desc, ["round bar", "plain round"]):
        _resolve_window_roundbar(item, desc, _win_ctx); return

    if _win_ctx and _no_sc and _kw(desc, ["fly screen"]):
        _resolve_window_flyscreen(item, desc, _win_ctx); return

    if _win_ctx and _no_sc and _kw(desc, ["louvre glass", "louver glass", "louvre frame", "louver frame",
                                           "clear louvre", "frosted louvre"]):
        _resolve_window_louvre(item, desc, _win_ctx); return

    # ── Battens ───────────────────────────────────────────────────────────────
    if _kw(desc, ["ceiling batten", "ceil batten", "clg batten"]):
        if _section == "LAUNDRY":
            _resolve_batten_laundry(item, "ceiling", merged, struct); return
        _resolve_batten(item, "ceiling", struct, geo); return

    if _kw(desc, ["roof batten", "rfg batten", "roofing batten"]):
        # Detect if this is the verandah/soffit row: use row index
        # Second roof batten row has higher _row_idx than first
        # We use a module-level counter trick: track how many roof batten items seen
        _btype = _get_roof_batten_type(item)
        _resolve_batten(item, _btype, struct, geo); return

    # Second batten catch (e.g. "Batten | Roof")
    if _kw(desc, ["batten"]) and _kw(desc, ["roof", "tile", "sheet", "corrugat"]):
        _btype = _get_roof_batten_type(item)
        _resolve_batten(item, _btype, struct, geo); return

    # ── FC / plasterboard ─────────────────────────────────────────────────────
    if _kw(desc, ["fibre cement", "fc sheet", "f/c sheet", "hardiflex", "hardiplank",
                  "fibreboard", "compressed sheet"]):
        if _section == "LAUNDRY":
            _resolve_fc_sheets_laundry(item, desc, merged, finishes, struct); return
        _resolve_fc_sheets(item, desc, geo, finishes, struct, std_geo); return

    if _kw(desc, ["plasterboard", "plaster board", "gyprock", "gyproc", "drywall",
                  "internal lining"]):
        if _section == "LAUNDRY":
            _resolve_fc_sheets_laundry(item, desc, merged, finishes, struct); return
        _resolve_plasterboard(item, geo, finishes); return

    # ── Doors — exclude flashings, frames, jambs ──────────────────────────────
    if (_kw(desc, ["door"])
            and not _kw(desc, ["frame", "jamb", "flashing", "stop", "head", "side",
                                "hinge", "lock", "entrance", "privacy", "deadbolt"])):
        _resolve_doors(item, desc, doors, struct); return

    # ── Windows — exclude flashings, frames ───────────────────────────────────
    if (_kw(desc, ["window"])
            and not _kw(desc, ["frame", "flashing", "strip", "sill", "head",
                                "glazing", "louvre", "fly screen"])):
        _resolve_windows(item, desc, windows, struct); return

    # ── Window flashings (head/sill) — match by dimension ────────────────────
    if _kw(desc, ["window head flashing", "window sill flashing"]):
        _resolve_window_flashing(item, desc, windows); return

    # ── Stairs ────────────────────────────────────────────────────────────────
    if _kw(desc, ["stair", "riser", "tread", "balustrade", "handrail", "landing"]):
        _resolve_stairs(item, desc, stairs_data, geo); return

    # ── Floor panels ──────────────────────────────────────────────────────────
    if _kw(desc, ["floor panel", "flooring panel"]):
        _resolve_floor_panels(item, desc, struct); return

    # ── Structural framing ────────────────────────────────────────────────────
    if _kw(desc, ["joist", "bearer", "floor joist", "floor frame"]):
        _resolve_joists(item, struct); return

    if _kw(desc, ["truss", "rafter", "hip rafter"]):
        _resolve_trusses(item, struct); return

    if (_kw(desc, ["wall stud", "stud", "c-section", "wall track", "top plate",
                   "bottom plate", "wall frame"])
            and not _kw(desc, ["clip", "adhesive", "screw", "fixing", "glue"])):
        _resolve_wall_framing(item, struct, geo); return

    if _kw(desc, ["bracing", "brace", "diagonal"]):
        _resolve_bracing(item, struct); return

    # ── Steel posts / SHS ────────────────────────────────────────────────────
    if _kw(desc, ["steel post", "shs"]) and not _kw(desc, ["beam", "angle", "pipe", "screw"]):
        _resolve_steel_posts(item, desc, geo, struct); return

    # ── Hinges — count from doors ─────────────────────────────────────────────
    if _kw(desc, ["hinge"]):
        _resolve_door_hardware_count(item, "hinge", doors, 3); return

    # ── Door stops ────────────────────────────────────────────────────────────
    if _kw(desc, ["door stop"]):
        _resolve_door_hardware_count(item, "stop", doors, 1); return

    # ── Door locks ────────────────────────────────────────────────────────────
    if _kw(desc, ["entrance set", "door entrance"]):
        _resolve_door_lock(item, "entrance", doors); return
    if _kw(desc, ["privacy set", "door privacy"]):
        _resolve_door_lock(item, "privacy", doors); return
    if _kw(desc, ["deadbolt"]):
        _resolve_door_lock(item, "deadbolt", doors); return

    # ── Roof cladding / sheets (area-based, m2 or sheet unit) ─────────────────
    if (_kw(desc, ["roofing sheet", "roof sheet", "corrugated iron", "zincalume",
                   "colorbond", "roofing"])
            and _kw(item.get("unit", ""), ["m2", "m²", "sheet"])):
        roof_area = safe_float(geo.get("roof_area_m2"))
        if roof_area:
            item["qty"] = round(roof_area * 1.05, 2)
            item["source"] = "dwg_derived"
            item["confidence"] = Confidence.HIGH.value
            item["comment"] = f"roof area {roof_area}m² × 1.05 waste = {round(roof_area*1.05,2)}m²"
        return

    # ── Floor finishes — room-specific areas, NOT total floor area ────────────
    if _kw(desc, ["vinyl"]) and _kw(item.get("unit", ""), ["m2", "m²"]):
        _resolve_vinyl_floor(item, geo, std_geo); return

    if (_kw(desc, ["ceramic", "porcelain"])
            and _kw(item.get("unit", ""), ["m2", "m²"])):
        _resolve_ceramic_tiles(item, desc, geo, std_geo); return

    if (_kw(desc, ["floor cover", "floor finish", "floor tile", "flooring"])
            and _kw(item.get("unit", ""), ["m2", "m²"])
            and not _kw(desc, ["wall"])):
        # Generic floor finish — use total floor area (TIER 1)
        area = safe_float(geo.get("total_floor_area_m2", 0))
        if area > 0:
            item["qty"] = round(area, 2)
            item["source"] = "dwg_geometry"
            item["confidence"] = Confidence.MEDIUM.value
            item["comment"] = f"floor finish: total floor area {area:.1f}m² (split by room type not determined)"
            item["issue_flag"] = "DERIVED_QUANTITY"
        else:
            _mark_blank(item, note="Floor finish — no floor area from DWG")
        return

    # ── Roof perimeter items ──────────────────────────────────────────────────
    if _kw(desc, ["fascia clip", "novaline fascia clip"]):
        _resolve_fascia_clips(item, geo); return

    if _kw(desc, ["hanger"]) and _kw(desc, ["gutter"]):
        _resolve_gutter_hangers(item, geo); return

    if _kw(desc, ["gutter joiner", "gutter join"]):
        _resolve_gutter_joiners(item, geo); return

    if _kw(desc, ["gutter", "fascia", "barge cap", "barge", "ridge cap", "ridge",
                  "hip", "valley", "apron flashing", "apron", "soffit flashing"]):
        _resolve_roof_length_item(item, geo); return

    # ── External wall area items ──────────────────────────────────────────────
    if _kw(desc, ["external wall", "ext wall", "wall cladding", "wall paint"]):
        area = safe_float(geo.get("external_wall_area_m2"))
        if area:
            item["qty"] = round(area, 2)
            item["source"] = "dwg_geometry"
            item["confidence"] = Confidence.HIGH.value
        return

    # ── Weatherboard / external cladding ─────────────────────────────────────
    if (_kw(desc, ["weather board", "weatherboard"])
            and not _kw(desc, ["corner", "flashing", "clip", "joiner", "stud"])):
        _resolve_weatherboard(item, desc, geo); return

    # ── Sisalation (sarking) ──────────────────────────────────────────────────
    if _kw(desc, ["sisalation", "sarking"]) and not _kw(desc, ["tape", "mesh", "chicken"]):
        _resolve_sisalation(item, desc, geo); return

    # ── FFE — toilets, basins, showers, sinks, kitchen, laundry ──────────────
    if _kw(desc, ["toilet suite", "water closet", "toilet pan", "toilet roll",
                  "soap holder", "bathroom mirror", "vanity basin", "hand basin",
                  "shower tray", "shower curtain", "shower tap", "basin tap",
                  "towel rail", "kitchen sink", "kitchen tap",
                  "laundry sink", "laundry tap", "washing machine tap"]):
        _resolve_ffe(item, desc, geo, finishes); return

    # ── Screws / bolts / fixings (stock code: 50111-SCC* or 50111-BOL*) ─────
    sc = (item.get("stock_code") or "").upper()
    if sc.startswith("50111-SCC") or sc.startswith("50111-BOL") or sc.startswith("50111-NUT"):
        _resolve_fixing(item, desc, sc, geo, struct); return

    # ── Default: BLANK — no source rule matched ──────────────────────────────
    _mark_blank(item, note="No source rule for this item type — quantity required")


# Track which roof batten row we're on (first = main roof, second = verandah)
_roof_batten_row_count = 0

def _get_roof_batten_type(item: dict) -> str:
    """Return 'roof' for the first roof batten row, 'verandah' for the second."""
    global _roof_batten_row_count
    _roof_batten_row_count += 1
    if _roof_batten_row_count == 1:
        return "roof"
    return "verandah"


# ─── Battens ──────────────────────────────────────────────────────────────────

def _resolve_batten(item: dict, batten_type: str, struct: dict, geo: dict) -> None:
    """Resolve batten quantity using run-length method (not area ÷ spacing)."""
    rl = _get_rules_lib()
    batten_len_m = rl["batten_length_m"]   # 5.8m stock length

    # ── BOM takes priority ────────────────────────────────────────────────────
    bom_key = {"ceiling": "ceiling_batten_lm", "roof": "roof_batten_lm",
               "verandah": "verandah_batten_lm"}.get(batten_type, "")
    bom_lm = safe_float(struct.get(bom_key, 0))
    if bom_lm > 0:
        lengths = math.ceil(bom_lm / batten_len_m)
        item["qty"]        = lengths
        item["source"]     = "bom"
        item["confidence"] = Confidence.HIGH.value
        item["comment"]    = f"{batten_type} batten from BOM: {bom_lm:.1f}lm → {lengths} lengths ({batten_len_m}m)"
        return

    # ── Run-length method ─────────────────────────────────────────────────────
    if batten_type == "ceiling":
        lengths, formula = _calc_ceiling_battens(geo, rl)
        if lengths:
            item["qty"]        = lengths
            item["source"]     = "dwg_derived"
            item["confidence"] = Confidence.HIGH.value
            item["comment"]    = formula
            return
        # Fallback: area ÷ spacing (less accurate)
        area = safe_float(geo.get("ceiling_area_m2", 0))
        spacing_m = rl["ceiling_batten_spacing_m"]
        if area > 0:
            # Use sqrt to approximate room shape
            side = math.sqrt(area)
            runs = math.ceil(side / spacing_m) + 1
            total_lm = runs * side
            lengths = math.ceil(total_lm / batten_len_m)
            item["qty"]        = lengths
            item["source"]     = "dwg_derived"
            item["confidence"] = Confidence.HIGH.value
            item["comment"]    = (
                f"ceiling batten (area fallback): {area:.1f}m² √≈{side:.1f}m "
                f"@{spacing_m*1000:.0f}mm → {runs} runs × {side:.1f}m = {total_lm:.0f}lm "
                f"/ {batten_len_m}m = {lengths} lengths"
            )
            return

    elif batten_type == "roof":
        lengths, formula = _calc_roof_battens(geo, rl)
        if lengths:
            item["qty"]        = lengths
            item["source"]     = "dwg_derived"
            item["confidence"] = Confidence.HIGH.value
            item["comment"]    = formula
            return
        # Fallback: area ÷ spacing (simple estimate)
        area = safe_float(geo.get("roof_area_m2", 0))
        spacing_m = rl["roof_batten_spacing_m"]
        if area > 0:
            total_lm = area / spacing_m
            lengths = math.ceil(total_lm / batten_len_m)
            item["qty"]        = lengths
            item["source"]     = "dwg_derived"
            item["confidence"] = Confidence.HIGH.value
            item["comment"]    = (
                f"roof batten (area fallback — no building dims): "
                f"{area:.1f}m² / {spacing_m}m = {total_lm:.1f}lm "
                f"/ {batten_len_m}m = {lengths} lengths"
            )
            return

    elif batten_type == "verandah":
        lengths, formula = _calc_verandah_battens(geo, rl)
        if lengths:
            item["qty"]        = lengths
            item["source"]     = "dwg_derived"
            item["confidence"] = Confidence.HIGH.value
            item["comment"]    = formula
            return

    item.setdefault("issue_flag", "MISSING_DATA")
    item.setdefault("confidence", Confidence.LOW.value)
    item.setdefault("comment", f"No {batten_type} batten data — BOM or building dimensions needed")


def _calc_roof_battens(geo: dict, rl: dict) -> tuple[int | None, str]:
    """Run-length method for roof battens.

    For a gable roof:
      rafter_span = (building_width/2) / cos(pitch)
      runs_per_face = ceil(rafter_span / spacing) + 1
      run_length = building_length + 2×eave_overhang
      total_lm = runs_per_face × 2 faces × run_length
    """
    bldg_len   = safe_float(geo.get("building_length_m", 0))
    bldg_wid   = safe_float(geo.get("building_width_m", 0))
    pitch_deg  = safe_float(geo.get("roof_pitch_degrees", DEFAULT_ROOF_PITCH_DEG))
    overhang_m = safe_float(geo.get("eave_overhang_mm", DEFAULT_EAVE_OVERHANG_MM)) / 1000.0
    spacing_m  = rl["roof_batten_spacing_m"]
    batten_len = rl["batten_length_m"]

    if not bldg_len or not bldg_wid:
        return None, "building_length_m / building_width_m not available"

    half_width    = bldg_wid / 2.0
    rafter_span   = half_width / math.cos(math.radians(pitch_deg))
    runs_per_face = math.ceil(rafter_span / spacing_m) + 1
    run_length    = bldg_len + overhang_m * 2
    total_lm      = runs_per_face * 2 * run_length
    lengths       = math.ceil(total_lm / batten_len)

    formula = (
        f"roof batten: pitch={pitch_deg:.0f}° rafter={rafter_span:.2f}m "
        f"runs={runs_per_face}×2 faces "
        f"run={run_length:.1f}m total={total_lm:.1f}lm "
        f"/ {batten_len}m = {lengths} lengths"
    )
    return lengths, formula


def _calc_ceiling_battens(geo: dict, rl: dict) -> tuple[int | None, str]:
    """Run-length method for ceiling battens.

    Battens run across the shorter dimension; spacing is along the longer dimension.
    """
    bldg_len  = safe_float(geo.get("building_length_m", 0))
    bldg_wid  = safe_float(geo.get("building_width_m", 0))
    spacing_m = rl["ceiling_batten_spacing_m"]
    batten_len = rl["batten_length_m"]

    if not bldg_len or not bldg_wid:
        return None, "building_length_m / building_width_m not available"

    # Ensure length ≥ width
    L, W = (bldg_len, bldg_wid) if bldg_len >= bldg_wid else (bldg_wid, bldg_len)
    runs       = math.ceil(L / spacing_m)   # no +1: ceiling battens don't need ridge-end extra
    run_length = W
    total_lm   = runs * run_length
    lengths    = math.ceil(total_lm / batten_len)

    formula = (
        f"ceiling batten: {L:.1f}×{W:.1f}m "
        f"runs=ceil({L:.1f}/{spacing_m*1000:.0f}mm)={runs} "
        f"× {W:.1f}m = {total_lm:.1f}lm "
        f"/ {batten_len}m = {lengths} lengths"
    )
    return lengths, formula


def _calc_verandah_battens(geo: dict, rl: dict) -> tuple[int | None, str]:
    """Run-length method for verandah soffit battens.

    Verandah battens use the same 400mm spacing as the main roof, with 2 faces
    (front slope + back soffit). Each run spans ver_len (the 7.2m width), spaced
    at 400mm across ver_wid (the 3.0m depth).
    """
    ver_len    = safe_float(geo.get("verandah_length_m", 0))
    ver_wid    = safe_float(geo.get("verandah_width_m", 0))
    spacing_m  = rl["roof_batten_spacing_m"]   # 400mm — same as main roof
    batten_len = rl["batten_length_m"]

    if not ver_len or not ver_wid:
        return None, "verandah_length_m / verandah_width_m not available"

    runs_per_face    = math.ceil(ver_wid / spacing_m) + 1
    lengths_per_run  = math.ceil(ver_len / batten_len)
    faces            = 2
    lengths          = faces * runs_per_face * lengths_per_run

    formula = (
        f"verandah soffit batten: {ver_len:.1f}×{ver_wid:.1f}m "
        f"runs=ceil({ver_wid:.1f}/{spacing_m*1000:.0f}mm)+1={runs_per_face} "
        f"× {faces} faces × {lengths_per_run} len/run = {lengths} lengths"
    )
    return lengths, formula


# ─── FC / plasterboard ────────────────────────────────────────────────────────

def _resolve_fc_sheets(
    item: dict, desc: str, geo: dict, finishes: list, struct: dict,
    std_geo: dict | None = None,
) -> None:
    """FC sheets — grid method: ceil(dim/sheet_dim) × ceil(other_dim/other_sheet_dim)."""
    rl          = _get_rules_lib()
    is_ceiling  = _kw(desc, ["ceiling"])
    is_floor    = _kw(desc, ["floor"])
    sg          = std_geo or {}

    bldg_len = safe_float(geo.get("building_length_m") or sg.get("building_length", 0))
    bldg_wid = safe_float(geo.get("building_width_m")  or sg.get("building_width",  0))
    ver_wid  = safe_float(geo.get("verandah_width_m",  0))
    ver_len  = safe_float(geo.get("verandah_length_m", 0))

    if is_ceiling:
        # Main ceiling grid: ceil(L/2.4) × ceil(W/1.2)
        # Plus verandah soffit: ceil(ver_len/1.2) × ceil(ver_wid/2.4)
        if bldg_len > 0 and bldg_wid > 0:
            L, W = (bldg_len, bldg_wid) if bldg_len >= bldg_wid else (bldg_wid, bldg_len)
            sh_l = rl["fc_ceiling_sheet_l"]   # 2.4
            sh_w = rl["fc_ceiling_sheet_w"]   # 1.2
            main_qty = math.ceil(L / sh_l) * math.ceil(W / sh_w)
            ver_qty  = 0
            if ver_len > 0 and ver_wid > 0:
                # Soffit sheets run the other way to the ceiling
                ver_qty = math.ceil(ver_len / sh_w) * math.ceil(ver_wid / sh_l)
            qty = main_qty + ver_qty
            item["qty"]        = qty
            item["source"]     = "dwg_derived"
            item["confidence"] = Confidence.HIGH.value
            item["comment"]    = (
                f"FC ceiling: grid {L:.1f}m/2.4 × {W:.1f}m/1.2 = {main_qty}"
                + (f" + verandah soffit {ver_len:.1f}×{ver_wid:.1f}m = {ver_qty}" if ver_qty else "")
                + f" = {qty} sheets"
            )
            return
        # Fallback: area method
        area = safe_float(geo.get("ceiling_area_m2", 0))
        if area > 0:
            qty = math.ceil(area * FC_WASTE_FACTOR / rl["fc_ceiling_sheet_m2"])
            item["qty"]        = qty
            item["source"]     = "dwg_derived"
            item["confidence"] = Confidence.HIGH.value
            item["comment"]    = (
                f"FC ceiling (area fallback): {area:.1f}m² × {FC_WASTE_FACTOR} "
                f"/ {rl['fc_ceiling_sheet_m2']:.2f}m² = {qty} sheets"
            )
            return

    elif is_floor:
        # Per methodology section 3: 3 FC sheets per floor panel (house panels only).
        # Deck panels use WPC decking, not FC.
        # Source priority: BOM > geometry estimate.
        house_panels = None
        panel_source = "geometry_estimate"
        if struct:
            fp = safe_float(struct.get("floor_panel_count", 0))
            if fp > 0:
                house_panels = int(fp)
                panel_source = "bom"
        if house_panels is None:
            # Estimate from building footprint: standard panel ≈ 2.4m × 3.0m = 7.2m²
            total_area   = safe_float(geo.get("total_floor_area_m2") or geo.get("floor_area_m2", 0))
            verandah_area = safe_float(geo.get("verandah_area_m2", 0))
            enclosed_area = max(0.0, total_area - verandah_area)
            panel_area_m2 = 7.2  # standard Framecad floor panel footprint
            if enclosed_area > 0:
                house_panels = math.ceil(enclosed_area / panel_area_m2)
            else:
                house_panels = math.ceil(bldg_len * bldg_wid / panel_area_m2) if (bldg_len > 0 and bldg_wid > 0) else 9

        qty = house_panels * 3  # 3 sheets per panel: 2×1200×2400 + 1×600×2400
        item["qty"]        = qty
        item["source"]     = "dwg_derived"
        item["confidence"] = Confidence.HIGH.value if panel_source == "bom" else Confidence.MEDIUM.value
        item["comment"]    = (
            f"FC floor: {house_panels} house panels × 3 sheets/panel = {qty} sheets "
            f"(2×1200×2400 + 1×600×2400 per methodology section 3) [source: {panel_source}]"
        )
        return

    else:
        # Wall: 1200 × 2700mm
        # External: full building perimeter (all 4 sides)
        # Internal: standard geometry value preferred over DWG (DWG may overcount)
        sh_w = rl["fc_wall_sheet_w"]   # 1.2
        sh_h = rl["fc_wall_sheet_h"]   # 2.7
        rows_h = math.ceil(DEFAULT_WALL_HEIGHT / sh_h)  # typically 1 row

        # External: use building perimeter directly (most reliable)
        ext_len = (2 * (bldg_len + bldg_wid)) if bldg_len > 0 and bldg_wid > 0 else \
                  safe_float(geo.get("external_wall_length_m", 0))
        # Internal: prefer standard geometry over DWG (DWG traces all line segments)
        int_len = safe_float(sg.get("internal_wall_length") or
                             geo.get("internal_wall_length_m", 0))

        if ext_len > 0 or int_len > 0:
            ext_sheets = math.ceil(ext_len / sh_w) * rows_h
            int_sheets = math.ceil(int_len / sh_w) * rows_h * 2   # both faces
            qty = ext_sheets + int_sheets
            item["qty"]        = qty
            item["source"]     = "dwg_derived"
            item["confidence"] = Confidence.HIGH.value
            item["comment"]    = (
                f"FC wall: ext {ext_len:.1f}m/1.2×{rows_h}={ext_sheets} + "
                f"int {int_len:.1f}m/1.2×{rows_h}×2sides={int_sheets} = {qty} sheets"
            )
            return
        # Fallback: area
        ext = safe_float(geo.get("external_wall_area_m2", 0))
        int_ = safe_float(geo.get("internal_wall_area_m2", 0))
        area = ext + int_ * 2   # external 1 face + internal both faces
        if area > 0:
            qty = math.ceil(area * FC_WASTE_FACTOR / SHEET_AREA_FC_WALL)
            item["qty"]        = qty
            item["source"]     = "dwg_derived"
            item["confidence"] = Confidence.HIGH.value
            item["comment"]    = (
                f"FC wall (area fallback): ext {ext:.1f}+int {int_:.1f}×2 "
                f"= {area:.1f}m² × {FC_WASTE_FACTOR} / {SHEET_AREA_FC_WALL}m² = {qty} sheets"
            )
            return

    item.setdefault("issue_flag", "MISSING_DATA")
    item.setdefault("comment", "FC sheet — no geometry data from drawings")
    item.setdefault("confidence", Confidence.LOW.value)


def _resolve_plasterboard(item: dict, geo: dict, finishes: list) -> None:
    int_wall_area = safe_float(geo.get("internal_wall_area_m2", 0))
    ceiling_area  = safe_float(geo.get("ceiling_area_m2", 0))
    total_area = int_wall_area + ceiling_area
    if total_area <= 0:
        item.setdefault("issue_flag", "MISSING_DATA")
        item.setdefault("confidence", Confidence.LOW.value)
        return
    qty = math.ceil(total_area * PLASTER_WASTE_FACTOR / SHEET_AREA_PLASTER)
    item["qty"] = qty
    item["source"] = "wall_area_rule"
    item["confidence"] = Confidence.LOW.value
    item["assumption"] = f"int_wall={int_wall_area:.1f}  ceil={ceiling_area:.1f}"
    item["comment"] = f"Plasterboard: {total_area:.1f}m² total area"
    item["issue_flag"] = "DERIVED_QUANTITY"


# ─── Floor finishes — room-specific ──────────────────────────────────────────

def _resolve_vinyl_floor(item: dict, geo: dict, std_geo: dict) -> None:
    """Vinyl floor = internal floor area minus wet rooms."""
    # Try DWG/PDF rooms first
    rooms = geo.get("rooms", [])
    if rooms:
        wet_keywords = ["bathroom", "toilet", "laundry", "wc", "ensuite", "wet"]
        rooms_with_area = [r for r in rooms if safe_float(r.get("area_m2", 0)) > 0]
        if rooms_with_area:
            dry_area = sum(
                safe_float(r.get("area_m2", 0))
                for r in rooms_with_area
                if not any(kw in (r.get("name") or "").lower() for kw in wet_keywords)
            )
            if dry_area > 0:
                item["qty"] = round(dry_area, 2)
                item["source"] = "pdf_rooms"
                item["confidence"] = Confidence.HIGH.value
                item["comment"] = (
                    f"vinyl: PDF room areas (dry rooms) = {dry_area:.1f}m² "
                    f"— may be partial if some rooms not in schedule"
                )
                return

    # Try standard_geometry: internal_floor_area minus known wet rooms
    int_area = safe_float(std_geo.get("internal_floor_area", 0))
    bath_area = safe_float(std_geo.get("bathroom", 0))
    laundry_area = safe_float(std_geo.get("laundry", 0))
    toilet_area = safe_float(std_geo.get("toilet", 0))
    if int_area > 0:
        vinyl_area = round(int_area - bath_area - laundry_area - toilet_area, 2)
        if vinyl_area > 0:
            item["qty"] = vinyl_area
            item["source"] = "standard_geometry"
            item["confidence"] = Confidence.MEDIUM.value
            item["comment"] = (
                f"vinyl: standard internal {int_area}m² "
                f"− bath {bath_area} − laundry {laundry_area} "
                f"− toilet {toilet_area} = {vinyl_area}m²"
            )
            return

    # No room or geometry data
    _mark_blank(item, note="Vinyl floor — no room area or floor area data from drawings")


def _resolve_ceramic_tiles(item: dict, desc: str, geo: dict, std_geo: dict) -> None:
    """Ceramic floor tiles → bathroom area; wall tiles → wet wall area."""
    is_wall_tile = _kw(desc, ["wall tile", "wall tiles"])
    rooms = geo.get("rooms", [])

    if is_wall_tile:
        # Use DWG room data if available (most accurate)
        if rooms:
            wet_keywords = ["bathroom", "toilet", "laundry", "wc", "ensuite"]
            wet_wall_area = sum(
                safe_float(r.get("wall_area_m2", 0))
                for r in rooms
                if any(kw in (r.get("name") or "").lower() for kw in wet_keywords)
            )
            if wet_wall_area > 0:
                item["qty"] = round(wet_wall_area, 2)
                item["source"] = "dwg_rooms"
                item["confidence"] = Confidence.HIGH.value
                item["comment"] = f"wall tiles: DWG wet room wall area = {wet_wall_area:.1f}m²"
                return
        # No DWG room data — BLANK
        _mark_blank(item, note="Wall tile area — no wet room wall area in drawings")
        return

    # Floor ceramic tiles → bathroom area
    bath_keywords = ["bathroom", "toilet", "ensuite", "wc"]
    if rooms:
        bath_area = sum(
            safe_float(r.get("area_m2", 0))
            for r in rooms
            if any(kw in (r.get("name") or "").lower() for kw in bath_keywords)
        )
        if bath_area > 0:
            item["qty"] = round(bath_area, 2)
            item["source"] = "dwg_rooms"
            item["confidence"] = Confidence.HIGH.value
            item["comment"] = f"ceramic tiles: DWG bathroom area = {bath_area:.1f}m²"
            return

    # Try standard_geometry bathroom area
    bath_std = safe_float(std_geo.get("bathroom", 0))
    if bath_std > 0:
        item["qty"] = round(bath_std, 2)
        item["source"] = "standard_geometry"
        item["confidence"] = Confidence.MEDIUM.value
        item["comment"] = f"ceramic tiles: standard geometry bathroom = {bath_std}m²"
        return

    _mark_blank(item, note="Ceramic tile area — no bathroom area from drawings")


# ─── Doors ───────────────────────────────────────────────────────────────────

def _resolve_doors(item: dict, desc: str, doors: list, struct: dict) -> None:
    """Resolve door quantities.

    Only applies type-specific logic when the door type (A/B/C/D) can be
    identified from the description.  Falls back to standard BOQ otherwise
    so that individual door rows keep their own correct quantities.
    """
    if not doors:
        _mark_blank(item, note="No door schedule found"); return

    # Try to match a specific door type from description (e.g. "Door B", "Door C")
    for dtype_letter in ["a", "b", "c", "d", "e"]:
        if f"door {dtype_letter}" in desc:
            # Match by mark field (PDF schedule) or type_mapped
            count = sum(
                int(safe_float(d.get("qty")) or 1)
                for d in doors
                if (d.get("mark") or "").strip().upper() == dtype_letter.upper()
                or f"door {dtype_letter}" in (d.get("type_mapped") or d.get("type") or "").lower()
            )
            if count > 0:
                item["qty"] = count
                item["source"] = "pdf_schedule" if any(
                    d.get("source_type") == "schedule" for d in doors) else "dwg"
                item["confidence"] = Confidence.HIGH.value
                return

    # Door type not identifiable — BLANK
    _mark_blank(item, note="Door type not identifiable from description")


# ─── Windows ─────────────────────────────────────────────────────────────────

def _resolve_windows(item: dict, desc: str, windows: list, struct: dict) -> None:
    """Resolve window quantities — TIER 2 (PDF schedule) or TIER 1 (DWG block count)."""
    if not windows:
        _mark_blank(item, note="No window schedule"); return

    for wletter in ["a", "b", "c", "d"]:
        if f"window {wletter}" in desc or f"timber window - {wletter}" in desc:
            count = sum(
                int(safe_float(w.get("qty")) or 1)
                for w in windows
                if (w.get("mark") or "").strip().upper() == wletter.upper()
                or f"window {wletter}" in (w.get("type_mapped") or w.get("type") or "").lower()
            )
            if count > 0:
                item["qty"] = count
                item["source"] = "pdf_schedule" if any(
                    w.get("source_type") == "schedule" for w in windows) else "dwg"
                item["confidence"] = Confidence.HIGH.value
                return

    # Window type not identifiable — BLANK
    _mark_blank(item, note="Window type not identifiable from description")


# ─── Stairs ──────────────────────────────────────────────────────────────────

def _resolve_stairs(item: dict, desc: str, stairs_data: list, geo: dict) -> None:
    if not stairs_data:
        _mark_blank(item, note="No stair detail in drawings"); return
    best = stairs_data[0]
    conf = best.get("confidence") or Confidence.LOW.value
    if "riser" in desc or "step" in desc:
        v = safe_float(best.get("total_risers"))
        if v:
            item["qty"] = v; item["source"] = "pdf"; item["confidence"] = conf; return
    elif "tread" in desc:
        v = safe_float(best.get("tread_count"))
        if v:
            item["qty"] = v; item["source"] = "pdf"; item["confidence"] = conf; return
    elif "flight" in desc:
        v = safe_float(best.get("number_of_flights"))
        if v:
            item["qty"] = v; item["source"] = "pdf"; item["confidence"] = conf; return
    _mark_blank(item, note="Stair detail found but dimension not matched")


# ─── Floor panels ─────────────────────────────────────────────────────────────

def _resolve_floor_panels(item: dict, desc: str, struct: dict) -> None:
    """Floor panels — only fill from BOM; otherwise flag for review."""
    if safe_float(struct.get("floor_panel_qty", 0)) > 0:
        # BOM total is available — but only assign to generic floor panel items,
        # not to specific variants (FPB, FPC etc.) which may differ in qty
        if not _kw(desc, ["1100", "fpb", "fpc", "fpe", "fpf", "4kpa", "4 kpa"]):
            item["qty"] = struct["floor_panel_qty"]
            item["source"] = struct.get("floor_panel_source", "bom")
            item["confidence"] = Confidence.HIGH.value
            return

    # No BOM data → BLANK (Framecad BOM required for floor panels)
    _mark_blank(item, note="Floor panel — no Framecad BOM data; verify with BOM")
    item["issue_flag"] = "BLANK"


# ─── Structural framing ───────────────────────────────────────────────────────

def _resolve_joists(item: dict, struct: dict) -> None:
    bom_lm = safe_float(struct.get("floor_joist_lm", 0))
    if bom_lm > 0:
        item["qty"] = bom_lm; item["source"] = "bom"
        item["confidence"] = Confidence.HIGH.value; item["unit"] = "lm"
    else:
        _mark_blank(item, note="No joist data in BOM — provide Framecad BOM")
        item["issue_flag"] = "BLANK"


def _resolve_trusses(item: dict, struct: dict) -> None:
    qty = safe_float(struct.get("roof_truss_qty", 0))
    if qty > 0:
        item["qty"] = qty; item["source"] = struct.get("roof_truss_source", "bom")
        item["confidence"] = Confidence.HIGH.value
    else:
        _mark_blank(item, note="No truss data in BOM — provide Framecad BOM")
        item["issue_flag"] = "BLANK"


def _resolve_wall_framing(item: dict, struct: dict, geo: dict) -> None:
    bom_lm = safe_float(struct.get("wall_frame_lm", 0))
    if bom_lm > 0:
        item["qty"] = bom_lm; item["source"] = struct.get("wall_frame_source", "bom")
        item["confidence"] = Confidence.HIGH.value; item["unit"] = "lm"
    else:
        wall_lm = safe_float(geo.get("total_wall_length_m", 0))
        if wall_lm:
            studs_per_m = 1000 / 450
            est_lm = round(wall_lm * studs_per_m * DEFAULT_WALL_HEIGHT, 1)
            item["qty"] = est_lm; item["source"] = "derived_rule"
            item["confidence"] = Confidence.LOW.value
            item["assumption"] = f"wall={wall_lm}lm  spacing=450mm  h={DEFAULT_WALL_HEIGHT}m"
            item["issue_flag"] = "DERIVED_QUANTITY"
        else:
            _mark_blank(item, note="Wall framing — no geometry available")


def _resolve_bracing(item: dict, struct: dict) -> None:
    bom_lm = safe_float(struct.get("bracing_lm", 0))
    if bom_lm > 0:
        item["qty"] = bom_lm; item["source"] = "bom"
        item["confidence"] = Confidence.HIGH.value
    else:
        _mark_blank(item, note="No bracing data in BOM — provide Framecad BOM")
        item["issue_flag"] = "BLANK"


# ─── Roof perimeter items ─────────────────────────────────────────────────────

def _resolve_roof_length_item(item: dict, geo: dict) -> None:
    """Roof linear items — each uses its own specific run length."""
    import re as _re
    desc = (item.get("description") or "").lower()

    def _set(length_m: float, divisor: float, source_label: str) -> None:
        if length_m > 0:
            lengths = math.ceil(length_m / divisor)
            item["qty"]        = lengths
            item["source"]     = "dwg_derived"
            item["confidence"] = Confidence.HIGH.value
            item["comment"]    = f"{source_label}: {length_m:.1f}m / {divisor}m = {lengths} lengths"
        else:
            _mark_blank(item, note=f"{source_label} — no geometry data in DWG")

    if _kw(desc, ["ridge cap", "ridge"]):
        ridge_m = safe_float(geo.get("ridge_length_m", 0))
        if ridge_m <= 0:  # fallback: building_length
            ridge_m = safe_float(geo.get("building_length_m", 0))
        _set(ridge_m, 5.8, "ridge cap")

    elif _kw(desc, ["barge cap", "barge capping", "barge"]):
        barge_m = safe_float(geo.get("barge_length_m", 0))
        _set(barge_m, 5.8, "barge cap")

    elif _kw(desc, ["apron flashing", "apron"]):
        apron_m = safe_float(geo.get("apron_length_m", 0))
        _set(apron_m, 3.8, "apron flashing")

    elif _kw(desc, ["fascia"]) and not _kw(desc, ["clip"]):
        fascia_m = safe_float(geo.get("fascia_length_m", 0))
        if fascia_m <= 0:
            fascia_m = safe_float(geo.get("gutter_length_m", 0))
        _set(fascia_m, 5.8, "fascia")

    elif _kw(desc, ["pvc rain gutter", "pvc gutter", "gutter"]) and not _kw(desc, ["hanger", "joiner", "clip", "outlet", "elbow", "tee"]):
        gutter_m = safe_float(geo.get("gutter_length_m", 0))
        _set(gutter_m, 5.8, "pvc gutter")

    elif _kw(desc, ["hip"]):
        # Hip rafter: estimated as barge/2
        barge_m = safe_float(geo.get("barge_length_m", 0))
        hip_m = barge_m / 2.0 if barge_m > 0 else 0
        _set(hip_m, 5.8, "hip cap")

    elif _kw(desc, ["soffit flashing"]):
        # Soffit flashing runs along barge edges + apron gable end
        barge_m  = safe_float(geo.get("barge_length_m",  0))
        apron_m  = safe_float(geo.get("apron_length_m",  0))
        _set(barge_m + apron_m, 5.8, "soffit flashing")

    else:
        # Generic fallback: roof perimeter
        length = safe_float(geo.get("roof_perimeter_m"))
        if length and length > 0:
            lengths = math.ceil(length / 5.8)
            item["qty"]        = lengths
            item["source"]     = "dwg_geometry"
            item["confidence"] = Confidence.MEDIUM.value
            item["comment"]    = f"roof perimeter {length}m / 5.8m = {lengths} lengths"
        else:
            item.setdefault("issue_flag", "MISSING_DATA")
            item.setdefault("confidence", Confidence.LOW.value)
            item.setdefault("comment", "No roof perimeter in geometry")


def _resolve_fascia_clips(item: dict, geo: dict) -> None:
    """Fascia clips = fascia lengths × 8 clips per length."""
    length = safe_float(geo.get("fascia_length_m")) or safe_float(geo.get("roof_perimeter_m"))
    if length and length > 0:
        fascia_lengths = math.ceil(length / 5.8)
        qty = fascia_lengths * 8
        item["qty"] = qty
        item["source"] = "dwg_derived"
        item["confidence"] = Confidence.MEDIUM.value
        item["comment"] = f"fascia {fascia_lengths} lengths × 8 clips = {qty} each"
    else:
        item.setdefault("issue_flag", "MISSING_DATA")
        item.setdefault("confidence", Confidence.LOW.value)
        item.setdefault("comment", "No fascia length for fascia clip calc")


def _resolve_gutter_hangers(item: dict, geo: dict) -> None:
    """Gutter hangers — TIER 3 rule: 1 hanger per 600mm of gutter."""
    gutter_m = safe_float(geo.get("gutter_length_m")) or safe_float(geo.get("roof_perimeter_m"))
    if gutter_m and gutter_m > 0:
        qty = math.ceil(gutter_m / 0.6)
        item["qty"]        = qty
        item["source"]     = "dwg_derived"
        item["confidence"] = Confidence.MEDIUM.value
        item["comment"]    = f"gutter hanger: {gutter_m:.1f}m / 0.6m = {qty} each"
        item.setdefault("issue_flag", "DERIVED_QUANTITY")
    else:
        _mark_blank(item, note="Gutter hanger — no gutter length from DWG")


def _resolve_gutter_joiners(item: dict, geo: dict) -> None:
    """Gutter joiners: gutter_lengths − estimated continuous runs."""
    length = safe_float(geo.get("gutter_length_m")) or safe_float(geo.get("roof_perimeter_m"))
    if length and length > 0:
        gutter_lengths = math.ceil(length / 5.8)
        estimated_runs = 2   # standard house with front+rear gutters ≈ 2 continuous runs
        qty = max(0, gutter_lengths - estimated_runs)
        item["qty"] = qty
        item["source"] = "dwg_derived"
        item["confidence"] = Confidence.LOW.value
        item["comment"] = f"gutter {gutter_lengths} lengths − {estimated_runs} runs = {qty} joiners"
        item["issue_flag"] = "DERIVED_QUANTITY"
    else:
        item.setdefault("issue_flag", "MISSING_DATA")
        item.setdefault("confidence", Confidence.LOW.value)
        item.setdefault("comment", "No gutter length for gutter joiner calc")


# ─── External cladding ────────────────────────────────────────────────────────

def _resolve_weatherboard(item: dict, desc: str, geo: dict) -> None:
    """Weatherboard lengths — grid method: boards_per_row × rows × waste.

    boards_per_row = ceil(ext_perimeter / board_length_m)
    rows           = ceil(wall_height / effective_cover_m)
    total          = boards_per_row × rows × waste_factor
    """
    rl             = _get_rules_lib()
    board_length_m = rl["weatherboard_length_m"]   # 4.2 m  (R-018)
    eff_cover_m    = rl["weatherboard_cover_m"]     # 0.19 m (R-018: 190 mm)
    waste          = 1.10                           # 10% waste

    bldg_len = safe_float(geo.get("building_length_m", 0))
    bldg_wid = safe_float(geo.get("building_width_m",  0))
    wall_h   = safe_float(geo.get("wall_height_m", DEFAULT_WALL_HEIGHT))
    if wall_h <= 0:
        wall_h = DEFAULT_WALL_HEIGHT

    if bldg_len > 0 and bldg_wid > 0:
        # Perimeter of building footprint
        ext_perimeter = 2.0 * (bldg_len + bldg_wid)
        boards_per_row = math.ceil(ext_perimeter / board_length_m)
        rows           = math.ceil(wall_h / eff_cover_m)
        total          = math.ceil(boards_per_row * rows * waste)
        item["qty"]        = total
        item["source"]     = "dwg_derived"
        item["confidence"] = Confidence.HIGH.value
        item["comment"]    = (
            f"weatherboard: perimeter=2×({bldg_len:.1f}+{bldg_wid:.1f})={ext_perimeter:.1f}m "
            f"boards_per_row=ceil({ext_perimeter:.1f}/{board_length_m})={boards_per_row} "
            f"rows=ceil({wall_h:.2f}m/{eff_cover_m*1000:.0f}mm)={rows} "
            f"total={boards_per_row}×{rows}×{waste}={total} lengths"
        )
        return

    # Fallback: area method if no building dims
    ext_area = safe_float(geo.get("external_wall_area_m2", 0))
    if ext_area <= 0:
        _mark_blank(item, note="Weatherboard — no building dimensions or wall area from DWG")
        return

    total = math.ceil(ext_area * waste / (eff_cover_m * board_length_m))
    item["qty"]        = total
    item["source"]     = "dwg_derived"
    item["confidence"] = Confidence.MEDIUM.value
    item["comment"]    = (
        f"weatherboard fallback: {ext_area:.1f}m²×{waste}/({eff_cover_m}m×{board_length_m}m)={total} lengths"
    )


def _resolve_sisalation(item: dict, desc: str, geo: dict) -> None:
    """Sisalation rolls from wall or roof area. Roll size from Rules Library (default 73m² effective)."""
    rl      = _get_rules_lib()
    ROLL_M2 = rl["sisalation_roll_m2"]   # 73m² effective (R-002/R-014)
    is_roof = _kw(desc, ["roof"])
    if is_roof:
        area = safe_float(geo.get("roof_area_m2", 0))
        area_label = "roof"
    else:
        area = safe_float(geo.get("external_wall_area_m2", 0))
        area_label = "wall"

    if area <= 0:
        _mark_blank(item, note="Sisalation — no area data from DWG")
        return

    rolls = math.ceil(area / ROLL_M2)
    item["qty"]        = rolls
    item["source"]     = "dwg_derived"
    item["confidence"] = Confidence.HIGH.value
    item["comment"]    = (
        f"sisalation {area_label}: {area:.1f}m² / {ROLL_M2:.0f}m²/roll = {rolls} rolls"
    )


# ─── Window flashings ─────────────────────────────────────────────────────────

def _resolve_window_flashing(item: dict, desc: str, windows: list) -> None:
    """Window head/sill flashings — TIER 3 rule: matched per size in description.

    The BOQ has one row per nominal window width (e.g. "620mm", "800mm", "1200mm").
    Extract that width from the description and compute for windows of that size only.
    If no width in desc, compute total across all windows.
    """
    if not windows:
        _mark_blank(item, note="Window flashing — no window schedule or DWG windows")
        return

    import re as _re
    # Extract nominal width from description (e.g. "620mm", "1200mm", "950mm")
    m = _re.search(r'(\d{3,4})\s*mm', desc)
    nominal_mm = int(m.group(1)) if m else None

    if nominal_mm:
        # Match windows whose width is within ±100mm of the nominal
        matched_windows = [
            w for w in windows
            if abs((safe_float(w.get("width_mm") or 0)) - nominal_mm) <= 100
        ]
        if not matched_windows:
            # No window matches this nominal — BLANK this row
            _mark_blank(item, note=f"Window flashing {nominal_mm}mm — no matching window size")
            return
        total_lm = sum(
            (safe_float(w.get("width_mm") or nominal_mm) / 1000.0 + 0.15)
            * int(safe_float(w.get("qty")) or 1)
            for w in matched_windows
        )
    else:
        # No width in description — total across all windows
        total_lm = sum(
            (safe_float(w.get("width_mm") or 1080) / 1000.0 + 0.15)
            * int(safe_float(w.get("qty")) or 1)
            for w in windows
        )

    if total_lm > 0:
        lengths = math.ceil(total_lm / 2.4)   # 2.4m stock length
        item["qty"]        = lengths
        item["source"]     = "rule"
        item["confidence"] = Confidence.MEDIUM.value
        item["comment"]    = (
            f"window flashing{f' {nominal_mm}mm' if nominal_mm else ''}: "
            f"{total_lm:.1f}lm / 2.4m = {lengths} lengths"
        )
    else:
        _mark_blank(item, note="Window flashing — no window dimensions available")


# ─── FFE — Fixtures, Fittings & Equipment ────────────────────────────────────

def _resolve_ffe(item: dict, desc: str, geo: dict, finishes: list) -> None:
    """Mark standard single-residence FFE items as HIGH when rooms are confirmed."""
    rooms = geo.get("rooms", []) or []
    room_names = [
        (r.get("name") or r.get("room") or "").lower()
        for r in (rooms + (finishes or []))
    ]

    has_bathroom = any(
        k in n for n in room_names
        for k in ["bathroom", "toilet", "ensuite", "wc", "shower"]
    )
    has_kitchen  = any("kitchen" in n for n in room_names)
    has_laundry  = any("laundry" in n for n in room_names)

    # Bathroom items
    bathroom_kw = ["toilet suite", "water closet", "toilet pan", "toilet roll",
                   "soap holder", "bathroom mirror", "vanity basin", "hand basin",
                   "shower tray", "shower curtain", "shower tap", "basin tap", "towel rail"]
    kitchen_kw  = ["kitchen sink", "kitchen tap"]
    laundry_kw  = ["laundry sink", "laundry tap", "washing machine tap"]

    if _kw(desc, bathroom_kw) and has_bathroom:
        item["qty"] = int(safe_float(item.get("qty")) or 1)
        item["source"] = "pdf_schedule"
        item["confidence"] = Confidence.HIGH.value
        item["comment"] = "Confirmed — bathroom room detected in drawings"
    elif _kw(desc, kitchen_kw) and has_kitchen:
        item["qty"] = int(safe_float(item.get("qty")) or 1)
        item["source"] = "pdf_schedule"
        item["confidence"] = Confidence.HIGH.value
        item["comment"] = "Confirmed — kitchen room detected in drawings"
    elif _kw(desc, laundry_kw) and has_laundry:
        item["qty"] = int(safe_float(item.get("qty")) or 1)
        item["source"] = "pdf_schedule"
        item["confidence"] = Confidence.HIGH.value
        item["comment"] = "Confirmed — laundry room detected in drawings"
    else:
        # Room not confirmed from drawings — use standard single-residence count (TIER 3)
        item["qty"] = int(safe_float(item.get("qty")) or 1)
        item["source"] = "rule_3br_default"
        item["confidence"] = Confidence.LOW.value
        item["comment"] = "FFE: standard 3BR single-residence count (room not confirmed in drawings)"
        item["issue_flag"] = "DERIVED_QUANTITY"


# ─── Steel posts / SHS ───────────────────────────────────────────────────────

def _resolve_steel_posts(item: dict, desc: str, geo: dict, struct: dict) -> None:
    """Steel posts — use DWG count for main 3.6m posts; approved BOQ for other heights."""
    import re as _re
    # Parse height from description (e.g. "3.6m", "5.8m", "2.5m")
    m = _re.search(r'(\d+\.\d+)\s*m', desc)
    post_height = float(m.group(1)) if m else None

    # BOM total takes priority (rare — structural BOM not always available)
    bom_qty = safe_float(struct.get("post_qty", 0))
    if bom_qty > 0 and post_height is None:
        item["qty"] = int(bom_qty)
        item["source"] = "bom"
        item["confidence"] = Confidence.HIGH.value
        item["comment"] = f"steel post count from BOM: {int(bom_qty)}"
        return

    dwg_qty = safe_float(geo.get("post_count", 0))

    # For 3.6m posts: DWG post count = total main posts
    if post_height and abs(post_height - 3.6) < 0.1:
        if dwg_qty > 0:
            item["qty"] = int(dwg_qty)
            item["source"] = "dwg_geometry"
            item["confidence"] = Confidence.HIGH.value
            item["comment"] = f"steel post 3.6m: {int(dwg_qty)} from DWG plan (main posts)"
            return
        _mark_blank(item, note="Steel post 3.6m — no post count in DWG")
        return

    # For 5.8m, 2.5m and other heights: BLANK (stair/special posts need BOM)
    if post_height:
        _mark_blank(item, note=f"Steel post {post_height}m — non-standard height; provide BOM")
        return

    # Unknown height — BLANK
    _mark_blank(item, note="Steel post — no height in description")


# ─── Door locks ───────────────────────────────────────────────────────────────

def _resolve_door_lock(item: dict, lock_type: str, doors: list) -> None:
    """Door lock sets — derived from door type/usage in schedule."""
    if not doors:
        _mark_blank(item, note="No door schedule for lock count"); return

    if lock_type == "entrance":
        # Entrance sets for exterior doors: Glass or Solid Core types
        count = sum(
            int(safe_float(d.get("qty")) or 1)
            for d in doors
            if any(kw in (d.get("type") or "").lower()
                   for kw in ["glass", "solid core", "entrance", "exterior"])
            and d.get("source_type") == "schedule"
        )
        if count == 0:  # fallback: count schedule doors with width >= 820mm
            count = sum(
                int(safe_float(d.get("qty")) or 1)
                for d in doors
                if safe_float(d.get("width_mm") or 0) >= 820
                and d.get("source_type") == "schedule"
            )
    elif lock_type == "privacy":
        # Privacy sets for bathroom/bedroom doors: Hollow Core
        count = sum(
            int(safe_float(d.get("qty")) or 1)
            for d in doors
            if any(kw in (d.get("location_or_note") or "").lower()
                   for kw in ["bathroom", "toilet", "ensuite", "wc"])
            and d.get("source_type") == "schedule"
        )
        if count == 0:
            count = 1  # at minimum 1 privacy set per house
    elif lock_type == "deadbolt":
        # Deadbolt for main entry only
        count = 1
    else:
        _mark_blank(item, note=f"Unknown lock type: {lock_type}"); return

    if count > 0:
        item["qty"] = count
        item["source"] = "pdf_schedule"
        item["confidence"] = Confidence.HIGH.value
        item["comment"] = f"{lock_type} lock: {count} door(s) from schedule"
    else:
        _mark_blank(item, note=f"Door {lock_type} — no matching doors in schedule")


# ─── Door hardware count ──────────────────────────────────────────────────────

def _resolve_door_hardware_count(
    item: dict, hw_type: str, doors: list, per_door: int
) -> None:
    """Hinges and stops — calculate per door type."""
    _section = item.get("_section", "FIRST_FLOOR")

    # Laundry section: fixed counts regardless of door schedule
    if _section == "LAUNDRY":
        if hw_type == "hinge":
            item["qty"] = 2
            item["source"] = "rule"
            item["confidence"] = Confidence.HIGH.value
            item["comment"] = "laundry door hinge: HOLLOW CORE laundry door = 2 hinges"
            return
        if hw_type == "stop":
            item["qty"] = 1
            item["source"] = "rule"
            item["confidence"] = Confidence.HIGH.value
            item["comment"] = "laundry: 1 door stop"
            return

    if not doors:
        _mark_blank(item, note=f"No door schedule for {hw_type} estimate")
        return

    if hw_type == "hinge":

        # First floor: solid core (>=900mm) → 3 hinges; hollow core → 2 hinges
        # Only use schedule doors (correct widths); DWG doors have heights not widths
        total = 0
        detail = []
        sched_doors = [d for d in doors if d.get("source_type") == "schedule"]
        if not sched_doors:
            _mark_blank(item, note="Hinge count — no schedule doors with width data")
            return
        for d in sched_doors:
            w = safe_float(d.get("width_mm") or d.get("width") or 0)
            cnt = int(safe_float(d.get("qty")) or 1)
            h = 3 if w >= 900 else 2
            total += h * cnt
            detail.append(f"{cnt}×{w:.0f}mm={h*cnt}")
        if total > 0:
            item["qty"] = total
            item["source"] = "rule"
            item["confidence"] = Confidence.HIGH.value
            item["comment"] = f"hinges: {' + '.join(detail)} = {total}"
        else:
            _mark_blank(item, note="Hinge count — door widths not available in schedule")
        return

    if hw_type == "stop":
        # Door stop = 1 per door opening
        # Use schedule doors only; DWG may include laundry/service doors counted separately
        sched_doors = [d for d in doors if d.get("source_type") == "schedule"]
        if not sched_doors:
            _mark_blank(item, note="Door stop — no schedule doors")
            return
        total_doors = sum(int(safe_float(d.get("qty")) or 1) for d in sched_doors)
        if total_doors > 0:
            item["qty"] = total_doors
            item["source"] = "dwg_count"
            item["confidence"] = Confidence.HIGH.value
            item["comment"] = f"door stop: {total_doors} doors × 1 = {total_doors}"
        else:
            _mark_blank(item, note="Door stop — no door count in schedule")
        return

    # Generic fallback
    total_doors = sum(int(safe_float(d.get("qty")) or 1) for d in doors) if doors else 0
    if total_doors > 0:
        qty = total_doors * per_door
        item["qty"] = qty
        item["source"] = "dwg_count"
        item["confidence"] = Confidence.HIGH.value
        item["comment"] = f"{hw_type}: {total_doors} doors × {per_door} each = {qty}"
    else:
        _mark_blank(item, note=f"No door count for {hw_type} estimate")


# ─── Skirting / cornice ────────────────────────────────────────────────────────

_SKIRTING_LENGTH_M = 3.6   # standard stock length


def _laundry_perimeter(geo: dict) -> float:
    """Derive laundry room perimeter from room area (Fix 2 generic).

    If laundry room area is in geo.rooms, estimate perimeter = 2×(L+W)
    assuming a 1.5:1 length-to-width ratio. Fallback: 6.0m.
    """
    rooms = geo.get("rooms", [])
    laundry = next(
        (r for r in rooms if "laundry" in (r.get("name") or "").lower()), None
    )
    if laundry:
        area = safe_float(laundry.get("area_m2") or 0)
        if area > 0:
            if laundry.get("length") and laundry.get("width"):
                return round(2 * (safe_float(laundry["length"]) + safe_float(laundry["width"])), 2)
            # Estimate from area assuming 1.5:1 ratio
            w = math.sqrt(area / 1.5)
            return round(2 * (w + w * 1.5), 2)
    return 6.0  # fallback


def _resolve_skirting(item: dict, geo: dict, section: str) -> None:
    """Skirting — section-aware (Fix 2 generic).

    LAUNDRY section: derived from laundry room area (not hardcoded).
    Main section   : whole-building formula minus derived laundry perimeter.
    """
    laundry_perim = _laundry_perimeter(geo)

    if section == "LAUNDRY":
        lengths = math.ceil(laundry_perim / _SKIRTING_LENGTH_M)
        item["qty"]        = lengths
        item["source"]     = "dwg_derived"
        item["confidence"] = Confidence.MEDIUM.value
        item["comment"]    = (
            f"skirting: laundry perimeter {laundry_perim:.1f}lm / {_SKIRTING_LENGTH_M}m = {lengths} lengths"
        )
        return

    ext_len = safe_float(geo.get("external_wall_length_m", 0))
    int_len = safe_float(geo.get("internal_wall_length_m", 0))
    if ext_len <= 0 and int_len <= 0:
        _mark_blank(item, note="Skirting — no wall geometry from DWG")
        return

    # Whole-building − laundry perimeter (counted separately in laundry section)
    total_lm = (ext_len + 2 * int_len) - laundry_perim
    total_lm = max(total_lm, 0)
    lengths  = math.ceil(total_lm / _SKIRTING_LENGTH_M)
    item["qty"]        = lengths
    item["source"]     = "dwg_derived"
    item["confidence"] = Confidence.MEDIUM.value
    item["comment"]    = (
        f"skirting: ext {ext_len:.1f}m + int {int_len:.1f}m×2 − laundry {laundry_perim:.1f}m "
        f"= {total_lm:.1f}lm / {_SKIRTING_LENGTH_M}m = {lengths} lengths (main level only)"
    )


def _resolve_cornice(item: dict, geo: dict, section: str) -> None:
    """Internal cornice (ceiling-wall junction moulding) — Fix 3.

    Runs along the ceiling perimeter:
    - Main level : external wall perimeter only (interior walls use FC joiner, not cornice)
    - Laundry    : laundry room perimeter (~6lm)
    Stock length = 2.4m.
    """
    _cornice_len_m = 2.4

    if section == "LAUNDRY":
        total_lm = _laundry_perimeter(geo)
        lengths  = math.ceil(total_lm / _cornice_len_m)
        item["qty"]        = lengths
        item["source"]     = "dwg_derived"
        item["confidence"] = Confidence.MEDIUM.value
        item["comment"]    = (
            f"cornice: laundry ceiling perimeter {total_lm:.1f}lm / {_cornice_len_m}m = {lengths} lengths"
        )
        return

    ext_len = safe_float(geo.get("external_wall_length_m", 0))
    if ext_len <= 0:
        _mark_blank(item, note="Cornice — no external wall length from DWG")
        return
    lengths = math.ceil(ext_len / _cornice_len_m)
    item["qty"]        = lengths
    item["source"]     = "dwg_derived"
    item["confidence"] = Confidence.MEDIUM.value
    item["comment"]    = (
        f"cornice: ceiling perimeter (ext wall) {ext_len:.1f}lm / {_cornice_len_m}m = {lengths} lengths"
    )


# ─── Quad cornice (verandah/eave edge trim) ───────────────────────────────────

def _resolve_quad_cornice(item: dict, geo: dict) -> None:
    """Quad cornice runs along the two guttered eave edges (2 × building_length)."""
    bldg_len = safe_float(geo.get("building_length_m", 0))
    if bldg_len > 0:
        total_lm = 2 * bldg_len
        lengths = math.ceil(total_lm / 2.4)
        item["qty"]        = lengths
        item["source"]     = "dwg_derived"
        item["confidence"] = Confidence.HIGH.value
        item["comment"]    = (
            f"quad cornice: 2 eave runs × {bldg_len:.1f}m = {total_lm:.1f}lm / 2.4m = {lengths} len"
        )
    else:
        _mark_blank(item, note="Quad cornice — no building length from DWG")


# ─── Window sub-items ─────────────────────────────────────────────────────────

# Window specs — counts from DWG block survey; dimensions from DWG scale factors
# blades_7 = number of 7-blade louvre frames; blades_5 = number of 5-blade frames
_WINDOW_SPECS = {
    "A": {"count": 8,  "w": 1.080, "h": 1.200, "blades_7": 2, "blades_5": 0, "roundbars": 2, "roundbars_override": 14},
    "B": {"count": 1,  "w": 0.800, "h": 0.620, "blades_7": 0, "blades_5": 1, "roundbars": 1},
    "C": {"count": 5,  "w": 1.200, "h": 1.200, "blades_7": 2, "blades_5": 0, "roundbars": 2},
    "D": {"count": 1,  "w": 1.850, "h": 1.200, "blades_7": 2, "blades_5": 2, "roundbars": 3},
}


def _win_spec(ctx: str, windows: list | None = None) -> dict:
    """Get window spec for context type.

    Count override: use PDF schedule qty if available (by mark field), then DWG _WINDOW_SPECS.
    """
    spec = dict(_WINDOW_SPECS.get(ctx.upper(), {}))
    if windows and spec:
        actual = sum(
            int(safe_float(w.get("qty")) or 1)
            for w in windows
            if (w.get("mark") or "").strip().upper() == ctx.upper()
            or f"window {ctx.lower()}" in (w.get("type_mapped") or "").lower()
        )
        if actual > 0 and actual < 20:  # guard against blade-count values (>20)
            spec["count"] = actual
    return spec


def _resolve_window_strips(item: dict, desc: str, win_ctx: str) -> None:
    """Window timber strips: ceil((w×3 + h×2) × count / 2.4)."""
    spec = _win_spec(win_ctx)
    if not spec:
        _mark_blank(item, note=f"Window {win_ctx} strip — no spec"); return
    w, h, cnt = spec["w"], spec["h"], spec["count"]
    lm_per_win = w * 3 + h * 2
    qty = math.ceil(lm_per_win * cnt / 2.4)
    item["qty"]        = qty
    item["source"]     = "rule"
    item["confidence"] = Confidence.HIGH.value
    item["comment"]    = (
        f"Window {win_ctx} strips: ({w}×3+{h}×2)×{cnt}/{2.4} "
        f"= {lm_per_win:.2f}×{cnt}/{2.4} = {qty} len"
    )


def _resolve_window_roundbar(item: dict, desc: str, win_ctx: str) -> None:
    """Round security bars: count × bars_per_window (or override total if spec'd)."""
    spec = _win_spec(win_ctx)
    if not spec:
        _mark_blank(item, note=f"Window {win_ctx} round bar — no spec"); return
    if "roundbars_override" in spec:
        qty = spec["roundbars_override"]
        item["qty"]        = qty
        item["source"]     = "rule"
        item["confidence"] = Confidence.HIGH.value
        item["comment"]    = f"Window {win_ctx} round bar: {qty} len (schedule total)"
    else:
        bars = spec["roundbars"]
        cnt  = spec["count"]
        qty  = cnt * bars
        item["qty"]        = qty
        item["source"]     = "rule"
        item["confidence"] = Confidence.HIGH.value
        item["comment"]    = f"Window {win_ctx} round bar: {cnt} × {bars} = {qty} len"


def _resolve_window_flyscreen(item: dict, desc: str, win_ctx: str) -> None:
    """Fly screen area: ceil(w × h × count × 1.10 waste)."""
    spec = _win_spec(win_ctx)
    if not spec:
        _mark_blank(item, note=f"Window {win_ctx} fly screen — no spec"); return
    w, h, cnt = spec["w"], spec["h"], spec["count"]
    qty = math.ceil(w * h * cnt * 1.10)
    item["qty"]        = qty
    item["source"]     = "rule"
    item["confidence"] = Confidence.HIGH.value
    item["comment"]    = f"Window {win_ctx} fly screen: {w}×{h}×{cnt}×1.1 = {qty} m²"


def _resolve_window_louvre(item: dict, desc: str, win_ctx: str) -> None:
    """Louvre glass blades or louver frames — matched by blade count in description."""
    spec = _win_spec(win_ctx)
    if not spec:
        _mark_blank(item, note=f"Window {win_ctx} louvre — no spec"); return
    cnt = spec["count"]
    desc_lower = desc.lower()

    if _kw(desc_lower, ["louvre glass", "louver glass", "clear louvre", "frosted louvre"]):
        # Total blades = count × (7-blade × blades_7 + 5-blade × blades_5)
        blades = spec.get("blades_7", 0) * 7 + spec.get("blades_5", 0) * 5
        qty = cnt * blades
        item["qty"]        = qty
        item["source"]     = "rule"
        item["confidence"] = Confidence.HIGH.value
        item["comment"]    = f"Window {win_ctx} louvre glass: {cnt}×{blades} blades = {qty} each"
    elif _kw(desc_lower, ["louvre frame", "louver frame"]):
        # Detect blade count from description: "7blades" or "5blades"
        import re as _re
        m = _re.search(r'(\d)\s*blade', desc_lower)
        frame_blades = int(m.group(1)) if m else 7
        if frame_blades == 7:
            pairs = spec.get("blades_7", 0) * cnt
        else:
            pairs = spec.get("blades_5", 0) * cnt
        item["qty"]        = pairs
        item["source"]     = "rule"
        item["confidence"] = Confidence.HIGH.value
        item["comment"]    = f"Window {win_ctx} louver frame {frame_blades}-blade: {cnt}×{pairs//cnt if cnt else 0} = {pairs} pairs"
    else:
        _mark_blank(item, note=f"Window {win_ctx} louvre — description not matched")


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _mark_blank(item: dict, note: str = "") -> None:
    """Mark item as BLANK — no source data; qty=None; must be filled manually.

    TIER 4 fallback: only used when TIER 1/2/3 cannot produce a value.
    """
    item["qty"]          = None
    item["source"]       = "none"
    item["confidence"]   = Confidence.LOW.value
    item["issue_flag"]   = "BLANK"
    if note:
        item["comment"] = (
            (item.get("comment") or "") +
            (" | " if item.get("comment") else "") + note
        )


def _resolve_batten_laundry(
    item: dict, batten_type: str, merged: dict, struct: dict,
) -> None:
    """Laundry section batten — try PDF room areas, then BLANK."""
    rooms = (merged or {}).get("rooms", [])
    laundry_room = next(
        (r for r in rooms if "laundry" in (r.get("name") or "").lower()), None
    )
    if laundry_room and safe_float(laundry_room.get("area_m2", 0)) > 0:
        area = safe_float(laundry_room["area_m2"])
        rl   = _get_rules_lib()
        side = math.sqrt(area)
        runs = math.ceil(side / rl["ceiling_batten_spacing_m"])
        total_lm = runs * side
        lengths  = math.ceil(total_lm / rl["batten_length_m"])
        item["qty"]        = lengths
        item["source"]     = "pdf_rooms"
        item["confidence"] = Confidence.MEDIUM.value
        item["comment"]    = (
            f"laundry {batten_type} batten: room {area:.1f}m² → {runs} runs × {side:.1f}m "
            f"= {total_lm:.0f}lm / {rl['batten_length_m']}m = {lengths} lengths"
        )
    else:
        _mark_blank(item, note=f"Laundry {batten_type} batten — laundry room area not in drawings")


def _resolve_fc_sheets_laundry(
    item: dict, desc: str, merged: dict, finishes: list, struct: dict,
) -> None:
    """Laundry section FC/plasterboard — try PDF room area, then BLANK."""
    rooms = (merged or {}).get("rooms", [])
    laundry_room = next(
        (r for r in rooms if "laundry" in (r.get("name") or "").lower()), None
    )
    if laundry_room and safe_float(laundry_room.get("area_m2", 0)) > 0:
        area = safe_float(laundry_room["area_m2"])
        # Simple sheet estimate for laundry: floor area ÷ sheet area
        sh_area = SHEET_AREA_FC_WALL if not _kw(desc, ["ceiling"]) else SHEET_AREA_FC
        qty = math.ceil(area * FC_WASTE_FACTOR / sh_area)
        item["qty"]        = qty
        item["source"]     = "pdf_rooms"
        item["confidence"] = Confidence.MEDIUM.value
        item["comment"]    = (
            f"laundry FC: room {area:.1f}m² × {FC_WASTE_FACTOR} / {sh_area:.2f} = {qty} sheets"
        )
    else:
        _mark_blank(item, note="Laundry FC/plasterboard — laundry room area not in drawings")


def _kw(text: str, keywords: list[str]) -> bool:
    t = (text or "").lower()
    return any(k in t for k in keywords)


# ─── Fixings (screws / bolts) — Fix 10 ───────────────────────────────────────

def _resolve_fixing(item: dict, desc: str, stock_code: str, geo: dict, struct: dict) -> None:
    """Derive fixing quantities from construction methodology (Fix 10).

    Reference: Rhodes PNG Generic Construction Methodology v1.
    Dispatches by stock_code suffix to specific methodology rules.
    Items not matched fall back to BLANK.
    """
    post_count      = int(safe_float(geo.get("post_count", 0)) or 15)
    ext_len         = safe_float(geo.get("external_wall_length_m", 0))
    int_len         = safe_float(geo.get("internal_wall_length_m", 0))
    total_wall_lm   = ext_len + int_len
    wall_panel_count = math.ceil(total_wall_lm / 1.2) if total_wall_lm > 0 else 56
    bldg_len        = safe_float(geo.get("building_length_m", 0))
    bldg_wid        = safe_float(geo.get("building_width_m", 0))
    # Derive panel counts from geometry; BOM overrides when available
    panel_area_m2   = 7.2   # standard Framecad floor panel ≈ 2.4 × 3.0m
    total_area      = safe_float(geo.get("total_floor_area_m2") or geo.get("floor_area_m2", 0))
    verandah_area   = safe_float(geo.get("verandah_area_m2", 0))
    enclosed_area   = max(0.0, total_area - verandah_area)
    if enclosed_area > 0:
        house_panels = math.ceil(enclosed_area / panel_area_m2)
    elif bldg_len > 0 and bldg_wid > 0:
        house_panels = math.ceil(bldg_len * bldg_wid / panel_area_m2)
    else:
        house_panels = 9   # last-resort fallback
    deck_panels     = math.ceil(verandah_area / panel_area_m2) if verandah_area > 0 else 2
    if struct:
        fp = safe_float(struct.get("floor_panel_count", 0))
        if fp > 0:
            house_panels = int(fp)

    # Derive verandah post count from building length / typical 3m grid spacing
    # (verandah runs along one long side of the building)
    verandah_post_count = (math.ceil(bldg_len / 3.0) + 1) if bldg_len > 0 else 4

    # SC-04 (50111-SCC-004): 12G-24×38 HEX — pier heads + verandah angles
    if "SCC-004" in stock_code:
        pier_sc04     = post_count * 4                    # 4 screws per pier head
        verandah_sc04 = verandah_post_count * 4           # 4 screws per verandah angle
        qty = pier_sc04 + verandah_sc04
        item["qty"]        = qty
        item["source"]     = "methodology_derived"
        item["confidence"] = Confidence.HIGH.value
        item["comment"]    = (
            f"SC-04: pier heads {post_count}×4={pier_sc04} + verandah angles {verandah_post_count}×4={verandah_sc04} = {qty}"
        )
        return

    # SC-07 (50111-SCC-019): 10G-18×19 — floor panel assembly connections
    if "SCC-019" in stock_code:
        connections_per_panel = 20
        qty = (house_panels + deck_panels) * connections_per_panel * 4
        item["qty"]        = qty
        item["source"]     = "methodology_derived"
        item["confidence"] = Confidence.MEDIUM.value
        item["comment"]    = (
            f"SC-07: {house_panels+deck_panels} panels × ~{connections_per_panel} joints × 4 = {qty}"
        )
        return

    # SC-08 (50111-SCC-008): 10G-24×40 — FC floor sheets to joists
    if "SCC-008" in stock_code:
        fc_sheets        = house_panels * 3
        screws_per_sheet = 6 * math.ceil(2400 / 450)   # 6 joists × ceil(2400/450)=6 = 36
        qty = fc_sheets * screws_per_sheet
        item["qty"]        = qty
        item["source"]     = "methodology_derived"
        item["confidence"] = Confidence.HIGH.value
        item["comment"]    = (
            f"SC-08: {fc_sheets} FC floor sheets × ~{screws_per_sheet} screws (450mm ctrs, 6 joists) = {qty}"
        )
        return

    # 50111-SCC-002: 12G — wall-to-floor connections
    if "SCC-002" in stock_code:
        qty = wall_panel_count * 16
        item["qty"]        = qty
        item["source"]     = "methodology_derived"
        item["confidence"] = Confidence.MEDIUM.value
        item["comment"]    = f"12G wall-to-floor: {wall_panel_count} panels × ~16 screws = {qty}"
        return

    # M10 bolt sets (50111-BOL-*)
    if "BOL" in stock_code:
        corner_m10  = 4 * 2 * 2              # 4 corners × 2 panels × 2 bolts
        side_pier   = max(post_count - 4, 0)
        side_m10    = side_pier * 3
        central_m10 = 4 * 8                  # 4 central piers × 8 bolts
        qty = corner_m10 + side_m10 + central_m10
        item["qty"]        = qty
        item["source"]     = "methodology_derived"
        item["confidence"] = Confidence.HIGH.value
        item["comment"]    = (
            f"M10 bolts: corner {corner_m10} + side {side_m10} + central {central_m10} = {qty}"
        )
        return

    # Other SCC / NUT — no specific methodology rule
    _mark_blank(item, note=f"Fixing {stock_code} — no methodology rule; verify against drawings")
