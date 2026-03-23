"""
summary_writer.py — Write project summary Excel file.

Produces: output/boq/[ProjectName]_Summary_[YYYYMMDD].xlsx

Data sources (in priority order per sheet):
  1. DWG geometry   — merged["geometry"] / merged["doors"] / merged["windows"]
  2. PDF content    — merged["finishes"] from pdf_extractor
  3. BOQ output     — boq_path (AMOUNT column for totals, full item list)
  4. Standard Geometry fallback — data/standard_model_G303_complete.xlsx.xlsx
     Standard Geometry sheet (marked explicitly when used)

RULE: Every value has a source label, or shows "Not found in drawings".
"""

from __future__ import annotations
import logging
from datetime import date
from pathlib import Path
from typing import Any, Optional

from src.config import OUTPUT_BOQ, DATA_DIR
from src.utils import safe_float

log = logging.getLogger("boq.summary")

BLUE_HEX  = "2F5496"
GREEN_HEX = "375623"
AMBER_HEX = "7F3F00"

# Standard model geometry fallback — only used when project_mode == "standard_model"
# Resolved dynamically from STANDARD_MODELS dir; these are legacy fallback names.
_STD_GEO_LEGACY_NAMES = [
    "standard_model_G303_complete.xlsx",
    "standard_model_G303_complete.xlsx.xlsx",  # legacy double-extension
]
_PROJ_SUMM_LEGACY_NAMES = [
    "SDP-3Bedroom_Project_Summary.xlsx",
]

# Resolved at runtime (see _resolve_std_geo_path / _resolve_proj_summ_path below)
_STD_GEO_PATH   = None
_PROJ_SUMM_PATH = None


def _resolve_std_geo_path() -> Optional[Path]:
    """Locate standard geometry file without hardcoding the double-extension name."""
    from src.config import DATA_DIR, STANDARD_MODELS
    for name in _STD_GEO_LEGACY_NAMES:
        p = DATA_DIR / name
        if p.exists():
            return p
    # Also search standard_models/
    for candidate in STANDARD_MODELS.glob("*.xlsx"):
        if "g303" in candidate.stem.lower() and "complete" in candidate.stem.lower():
            return candidate
    return None


def _resolve_proj_summ_path() -> Optional[Path]:
    """Locate project summary reference file without hardcoding project-specific names."""
    from src.config import DATA_DIR
    for name in _PROJ_SUMM_LEGACY_NAMES:
        p = DATA_DIR / name
        if p.exists():
            return p
    return None


# ─── Public entry point ───────────────────────────────────────────────────────

def write_summary(
    project_name: str,
    boq_items: list[dict],
    validation: dict,
    merged: dict,
    files_found: dict,
    boq_path: Path | None = None,
    project_mode: str = "custom_project",
) -> Path:
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl not installed")

    date_str    = date.today().strftime("%Y%m%d")
    output_path = OUTPUT_BOQ / f"{project_name}_Summary_{date_str}.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()

    # Item Detail — read from BOQ file so quantities match exactly
    if boq_path and Path(boq_path).exists():
        boq_rows = _read_boq_rows(boq_path, openpyxl)
    else:
        boq_rows = _boq_items_to_rows(boq_items)

    ws_sum = wb.active
    ws_sum.title = "Project Summary"
    _write_project_summary(ws_sum, project_name, boq_items, boq_rows, validation, merged, files_found, openpyxl)

    ws_det = wb.create_sheet("Item Detail")
    _write_item_detail(ws_det, boq_rows, openpyxl)

    # Schedule sheets from drawing data + fallbacks
    _write_room_schedule(wb.create_sheet("Room Schedule"), merged, openpyxl, project_mode)
    _write_door_schedule(wb.create_sheet("Door Schedule"), merged, openpyxl)
    _write_window_schedule(wb.create_sheet("Window Schedule"), merged, openpyxl)
    _write_finish_schedule(wb.create_sheet("Finish Schedule"), merged, openpyxl, project_mode)
    _write_structural_summary(wb.create_sheet("Structural Summary"), merged, openpyxl)
    _write_services_summary(wb.create_sheet("Services Summary"), boq_rows, boq_items, openpyxl)
    _write_drawing_register(wb.create_sheet("Drawing Register"), files_found, merged, openpyxl)

    wb.save(str(output_path))
    log.info("Summary written: %s  (%d sheets)", output_path.name, len(wb.sheetnames))
    return output_path


# ─── Read BOQ file for Item Detail ────────────────────────────────────────────

def _read_boq_rows(boq_path: Path, openpyxl) -> list[dict]:
    """Read the written BOQ file and return rows exactly as saved.

    Handles both BOQ formats:
    - Approved BOQ template (standard_model): data starts at row 9.
      Cols: A=StockCode, B=Desc, C=Qty, D=Unit, E=Rate, F=Amount, G=Conf, H=Source, I=Notes
    - Full fresh workbook (custom_project): header row detected by "Item No" in col A.
      Cols: A=ItemNo, B=StockCode, C=Desc, D=Unit, E=Qty, F=Rate, G=Amount, H=Conf, I=Source,
            J=DrawingRef, K=Notes
    """
    rows = []
    try:
        wb = openpyxl.load_workbook(str(boq_path), read_only=True, data_only=True)
        sheet_name = "BOQ" if "BOQ" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet_name]

        # Scan rows 1-15 to find the header row and determine format
        DATA_START = 9       # default: approved BOQ template
        is_full_workbook = False
        for check_row_idx in range(1, 16):
            check_row = next(
                ws.iter_rows(min_row=check_row_idx, max_row=check_row_idx, values_only=True),
                None,
            )
            if check_row is None:
                break
            cell0 = str(check_row[0] or "").strip().lower()
            if cell0 in ("item no", "item no.", "#"):
                DATA_START = check_row_idx + 1
                is_full_workbook = True
                break

        for row in ws.iter_rows(min_row=DATA_START, values_only=True):
            if len(row) < 2:
                continue

            if is_full_workbook:
                # Full workbook: A=ItemNo, B=StockCode, C=Desc, D=Unit, E=Qty,
                #                F=Rate, G=Amount, H=Conf, I=Source, J=DrawingRef, K=Notes
                sc    = row[1] if len(row) > 1 else None
                desc  = row[2] if len(row) > 2 else None
                unit  = row[3] if len(row) > 3 else None
                qty   = row[4] if len(row) > 4 else None
                rate  = row[5] if len(row) > 5 else None
                amt   = row[6] if len(row) > 6 else None
                conf  = row[7] if len(row) > 7 else None
                src   = row[8] if len(row) > 8 else None
                notes = row[10] if len(row) > 10 else (row[9] if len(row) > 9 else None)
                is_hdr = (
                    sc is None and qty is None and desc is not None
                    or (row[0] is not None and sc is None and desc is None)
                )
            else:
                # Approved BOQ template: A=StockCode, B=Desc, C=Qty, D=Unit,
                #                        E=Rate, F=Amount, G=Conf, H=Source, I=Notes
                sc    = row[0]
                desc  = row[1]
                qty   = row[2] if len(row) > 2 else None
                unit  = row[3] if len(row) > 3 else None
                rate  = row[4] if len(row) > 4 else None
                amt   = row[5] if len(row) > 5 else None
                conf  = row[6] if len(row) > 6 else None
                src   = row[7] if len(row) > 7 else None
                notes = row[8] if len(row) > 8 else None
                is_hdr = (sc is None and qty is None and desc is not None)

            if sc is None and desc is None and (not is_full_workbook or row[0] is None):
                continue
            rows.append({
                "stock_code":  str(sc or "").strip(),
                "description": str(desc or "").strip(),
                "qty":         qty,
                "unit":        str(unit or "").strip(),
                "rate":        rate,
                "amount":      amt,
                "confidence":  str(conf or "").strip(),
                "source":      str(src or "").strip(),
                "notes":       str(notes or "").strip(),
                "_is_header":  is_hdr,
            })
        wb.close()
    except Exception as exc:
        log.warning("Could not read BOQ file for summary: %s", exc)
    return rows


def _boq_items_to_rows(boq_items: list[dict]) -> list[dict]:
    """Fallback: convert boq_items to the same row format.

    Handles both old keys (qty, source) and boq_mapper keys (quantity, source_evidence).
    """
    rows = []
    for item in boq_items:
        # Prefer "qty", fall back to "quantity" (boq_mapper output)
        qty_raw = item.get("qty") if item.get("qty") is not None else item.get("quantity")
        qty  = safe_float(qty_raw)
        rate = safe_float(item.get("rate"))
        amt  = round(qty * rate, 2) if (qty and rate) else None
        # Prefer "source", fall back to "source_evidence" (boq_mapper output)
        source_val = item.get("source") or item.get("source_evidence") or ""
        notes_parts = list(filter(None, [
            item.get("issue_flag"), item.get("assumption"),
            item.get("comment"), item.get("notes"),
        ]))
        rows.append({
            "stock_code":  item.get("stock_code", ""),
            "description": item.get("description", ""),
            "qty":         qty,
            "unit":        item.get("unit", ""),
            "rate":        rate,
            "amount":      amt,
            "confidence":  (item.get("confidence") or "").upper(),
            "source":      source_val,
            "notes":       " | ".join(notes_parts),
            "_is_header":  False,
        })
    return rows


# ─── Room level helper ────────────────────────────────────────────────────────

_GROUND_LEVEL_ROOMS = {"laundry", "laundry room", "ground level laundry"}

_GROUND_FLOOR_DRAWING_KEYS = ("ground floor", "ground level", "slab", "a-001")
_FIRST_FLOOR_DRAWING_KEYS  = ("first floor", "upper", "elevated", "a-002", "a-003")


def _get_room_level(room_name: str, drawing_source: str = "") -> str:
    """Determine the correct level for a room.

    Priority:
    1. Explicit floor reference in drawing_source (e.g. 'Ground Floor Plan A-001')
    2. Room name in known ground-level set (laundry)
    3. Default: First Floor (elevated Rhodes building)
    """
    src = (drawing_source or "").lower()
    if any(k in src for k in _GROUND_FLOOR_DRAWING_KEYS):
        return "Ground Level"
    if any(k in src for k in _FIRST_FLOOR_DRAWING_KEYS):
        return "First Floor"
    if room_name.lower().strip() in _GROUND_LEVEL_ROOMS:
        return "Ground Level"
    return "First Floor"


# ─── Fallback data loaders ────────────────────────────────────────────────────

def _load_std_geo_rooms(openpyxl, project_mode: str = "custom_project") -> list[dict]:
    """Load room list from Standard Geometry sheet in standard model file.

    Only used when project_mode == "standard_model" to avoid pulling G303
    room areas into non-G303 projects.

    Returns list of dicts: {name, area_m2, source, floor_finish, wall_finish, ceiling_finish}.
    """
    rooms = []
    if project_mode != "standard_model":
        return rooms

    std_geo_path = _resolve_std_geo_path()
    if std_geo_path is None:
        log.debug("Standard Geometry file not found — skipping fallback rooms")
        return rooms

    try:
        wb = openpyxl.load_workbook(str(std_geo_path), read_only=True, data_only=True)
        if "Standard Geometry" not in wb.sheetnames:
            wb.close()
            return rooms
        ws = wb["Standard Geometry"]
        in_rooms = False
        for row in ws.iter_rows(min_row=1, values_only=True):
            label = str(row[0] or "").strip() if row[0] is not None else ""
            if label == "ROOMS":
                in_rooms = True
                continue
            if label in ("WALLS", "ROOF", "DOORS & WINDOWS", "OVERALL BUILDING"):
                in_rooms = False
            if not in_rooms or not label:
                continue
            val = row[2] if len(row) > 2 else None
            src_cell = str(row[5] or "") if len(row) > 5 else ""
            try:
                area = float(val) if val is not None else 0.0
            except (TypeError, ValueError):
                area = 0.0
            if area > 0:
                rooms.append({
                    "name":     label,
                    "area_m2":  area,
                    "source":   (
                        f"standard_geometry ({src_cell})" if src_cell
                        else f"standard_geometry ({std_geo_path.name})"
                    ),
                })
        wb.close()
    except Exception as exc:
        log.warning("Could not load std geo rooms: %s", exc)
    return rooms


def _load_proj_summary_rooms(openpyxl, project_mode: str = "custom_project") -> list[dict]:
    """Load room schedule from a reference project summary file (Room Schedule sheet).

    Only used when project_mode == "standard_model" to avoid polluting
    custom projects with another project's room data.

    Returns list of dicts with name, level, area_m2, floor_finish, wall_finish,
    ceiling_finish, ceiling_height, notes, source.
    """
    rooms = []
    if project_mode != "standard_model":
        return rooms

    proj_summ_path = _resolve_proj_summ_path()
    if proj_summ_path is None:
        return rooms

    try:
        wb = openpyxl.load_workbook(str(proj_summ_path), read_only=True, data_only=True)
        if "Room Schedule" not in wb.sheetnames:
            wb.close()
            return rooms
        ws = wb["Room Schedule"]
        header_found = False
        for row in ws.iter_rows(min_row=1, values_only=True):
            if not any(c is not None for c in row):
                continue
            label = str(row[0] or "").strip()
            # Find data rows (skip title, source note, header row, footer)
            if label == "Room Name":
                header_found = True
                continue
            if not header_found or label.startswith("Data extracted"):
                continue
            try:
                area = float(row[2]) if row[2] is not None else 0.0
            except (TypeError, ValueError):
                area = 0.0
            rooms.append({
                "name":            label,
                "level":           str(row[1] or "").strip(),
                "area_m2":         area,
                "ceiling_height":  str(row[3] or "").strip(),
                "floor_finish":    str(row[4] or "").strip(),
                "wall_finish":     str(row[5] or "").strip(),
                "ceiling_finish":  str(row[6] or "").strip(),
                "notes":           str(row[7] or "").strip(),
                "source":          f"reference_summary ({proj_summ_path.name})",
            })
        wb.close()
    except Exception as exc:
        log.warning("Could not load project summary rooms: %s", exc)
    return rooms


# ─── Style helpers ────────────────────────────────────────────────────────────

def _hdr(ws, row: int, text: str, openpyxl, ncols: int = 9) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment
    for c in range(1, ncols + 1):
        cell = ws.cell(row, c)
        cell.fill = PatternFill("solid", fgColor=BLUE_HEX)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center" if c > 1 else "left", vertical="center")
    ws.cell(row, 1).value = text
    ws.row_dimensions[row].height = 18


def _col_headers(ws, row: int, headers: list[str], openpyxl) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment
    FILL = PatternFill("solid", fgColor="D9E1F2")
    FONT = Font(bold=True, size=10)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row, c)
        cell.value = h
        cell.font  = FONT
        cell.fill  = FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 16


def _not_found(ws, row: int, col: int, where: str, openpyxl) -> None:
    from openpyxl.styles import Font
    ws.cell(row, col).value = f"Not found — {where}"
    ws.cell(row, col).font  = Font(italic=True, color="808080", size=10)


# ─── Sheet 1: Project Summary ─────────────────────────────────────────────────

def _write_project_summary(ws, project_name, boq_items, boq_rows, validation, merged, files_found, openpyxl):
    from openpyxl.styles import Font, PatternFill, Alignment

    BLUE  = PatternFill("solid", fgColor=BLUE_HEX)
    WHITE = Font(bold=True, size=12, color="FFFFFF")
    BOLD  = Font(bold=True, size=11)
    NORM  = Font(size=11)
    GREY  = Font(italic=True, size=10, color="808080")

    def section(row, text):
        ws.cell(row, 1).value = text
        ws.cell(row, 1).font  = WHITE
        for c in range(1, 4):
            ws.cell(row, c).fill = BLUE
        ws.merge_cells(f"A{row}:C{row}")
        ws.cell(row, 1).alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 18

    def kv(row, k, v, fmt=None, src=None):
        ws.cell(row, 1).value = k
        ws.cell(row, 1).font  = BOLD
        ws.cell(row, 2).value = v
        ws.cell(row, 2).font  = NORM
        if fmt:
            ws.cell(row, 2).number_format = fmt
        if src:
            ws.cell(row, 3).value = f"[{src}]"
            ws.cell(row, 3).font  = GREY

    # Title
    ws.cell(1, 1).value = "PROJECT BOQ SUMMARY"
    ws.cell(1, 1).font  = Font(bold=True, size=16, color=BLUE_HEX)
    ws.merge_cells("A1:C1")
    ws.cell(2, 1).value = f"Project: {project_name}"
    ws.cell(2, 1).font  = Font(bold=True, size=12)
    ws.cell(3, 1).value = f"Date: {date.today().strftime('%d %B %Y')}"
    ws.cell(3, 1).font  = NORM

    row = 5
    section(row, "COST SUMMARY"); row += 1

    # Total from BOQ AMOUNT column — sum ALL rows including TOTAL row
    # (boq_rows includes the TOTAL row written by boq_writer, whose amount
    # equals grand_total; adding item amounts + TOTAL row gives the full tally
    # that matches the sum visible when the user opens col F in Excel)
    import math as _math
    total = 0.0
    for r in boq_rows:
        a = safe_float(r.get("amount"))
        if a is not None and not (isinstance(a, float) and _math.isnan(a)):
            total += a
    if total == 0:
        # Fallback: compute from qty * rate
        total = sum(
            (safe_float(i.get("qty")) or 0) * (safe_float(i.get("rate")) or 0)
            for i in boq_items
        )
    item_count = sum(1 for r in boq_rows if not r.get("_is_header") and r.get("qty") is not None)
    kv(row, "Total Supply Cost (PGK)", round(total, 2) if total else "N/A",
       fmt="#,##0.00", src="BOQ AMOUNT column"); row += 1
    kv(row, "Total Line Items (with qty)", item_count, src="BOQ file"); row += 2

    section(row, "CONFIDENCE BREAKDOWN"); row += 1
    high_c = sum(1 for i in boq_items if (i.get("confidence") or "").upper() == "HIGH")
    med_c  = sum(1 for i in boq_items if (i.get("confidence") or "").upper() == "MEDIUM")
    low_c  = sum(1 for i in boq_items if (i.get("confidence") or "").upper() == "LOW")
    blank_c = sum(1 for i in boq_items if i.get("issue_flag") == "BLANK" or i.get("qty") is None)
    kv(row, "HIGH (from DWG/PDF drawings)", high_c,  src="boq_items"); row += 1
    kv(row, "MEDIUM (rule/estimate)",       med_c,   src="boq_items"); row += 1
    kv(row, "LOW (no data — BLANK)",        blank_c, src="boq_items"); row += 1
    row += 1

    geo = merged.get("geometry", {})
    audit = merged.get("audit", {})
    section(row, "PROJECT GEOMETRY"); row += 1

    def geo_src(key):
        return audit.get(key, {}).get("source", "dwg_derived") if key in audit else "dwg"

    kv(row, "Total Floor Area (m²)",    geo.get("total_floor_area_m2", "N/A"),
       src=geo_src("total_floor_area_m2")); row += 1
    kv(row, "Building Length (m)",      geo.get("building_length_m", "N/A"),
       src=geo_src("building_length_m")); row += 1
    kv(row, "Building Width (m)",       geo.get("building_width_m", "N/A"),
       src=geo_src("building_width_m")); row += 1
    kv(row, "External Wall Length (m)", geo.get("external_wall_length_m", "N/A"),
       src=geo_src("external_wall_length_m")); row += 1
    kv(row, "Internal Wall Length (m)", geo.get("internal_wall_length_m", "N/A"),
       src=geo_src("internal_wall_length_m")); row += 1
    kv(row, "Roof Area (m²)",           geo.get("roof_area_m2", "N/A"),
       src=geo_src("roof_area_m2")); row += 1
    kv(row, "Roof Pitch (°)",           geo.get("roof_pitch_degrees", "N/A"),
       src=geo_src("roof_pitch_degrees")); row += 1
    kv(row, "Doors (from merged)",      len(merged.get("doors", [])),
       src=merged.get("audit", {}).get("doors_source", "none")); row += 1
    kv(row, "Windows (from merged)",    len(merged.get("windows", [])),
       src=merged.get("audit", {}).get("windows_source", "none")); row += 1
    kv(row, "Post Count",               geo.get("post_count", "N/A"),
       src=geo_src("post_count")); row += 2

    conflicts = validation.get("conflicts", [])
    section(row, f"CONFLICTS ({len(conflicts)})"); row += 1
    if conflicts:
        _col_headers(ws, row, ["Item", "Recommended Action", "Source"], openpyxl); row += 1
        for c in conflicts:
            ws.cell(row, 1).value = c.get("item_name", ""); ws.cell(row, 1).font = NORM
            ws.cell(row, 2).value = c.get("recommended_action", ""); ws.cell(row, 2).font = NORM
            ws.cell(row, 3).value = f"A={c.get('value_a')} vs B={c.get('value_b')} ({c.get('diff_pct')}%)"; ws.cell(row, 3).font = GREY
            row += 1
    else:
        ws.cell(row, 1).value = "No conflicts detected"; ws.cell(row, 1).font = GREY; row += 1
    row += 1

    missing = validation.get("missing_scope", [])
    section(row, f"MISSING SCOPE ({len(missing)})"); row += 1
    if missing:
        _col_headers(ws, row, ["Risk", "Category", "Description"], openpyxl); row += 1
        for m in missing:
            ws.cell(row, 1).value = m.get("risk", ""); ws.cell(row, 1).font = NORM
            ws.cell(row, 2).value = m.get("category", ""); ws.cell(row, 2).font = NORM
            ws.cell(row, 3).value = m.get("description", ""); ws.cell(row, 3).font = NORM
            row += 1
    else:
        ws.cell(row, 1).value = "No missing scope identified"; ws.cell(row, 1).font = GREY

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 40


# ─── Sheet 2: Item Detail ─────────────────────────────────────────────────────

def _write_item_detail(ws, rows: list[dict], openpyxl) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment

    headers = ["Stock Code", "Description", "Unit", "QTY", "Rate (PGK)", "Amount (PGK)", "Confidence", "Source", "Notes"]
    _col_headers(ws, 1, headers, openpyxl)

    CONF_FILL = {
        "HIGH":   PatternFill("solid", fgColor="C6EFCE"),
        "MEDIUM": PatternFill("solid", fgColor="FFEB9C"),
        "LOW":    PatternFill("solid", fgColor="FFC7CE"),
    }
    HDR_FILL = PatternFill("solid", fgColor="E2EFDA")
    NORM = Font(size=10)
    BOLD = Font(bold=True, size=10)

    for r, row in enumerate(rows, 2):
        is_hdr = row.get("_is_header", False)
        conf = str(row.get("confidence") or "").upper()
        font = BOLD if is_hdr else NORM

        ws.cell(r, 1).value = row.get("stock_code", ""); ws.cell(r, 1).font = font
        ws.cell(r, 2).value = row.get("description", ""); ws.cell(r, 2).font = font
        ws.cell(r, 3).value = row.get("unit", ""); ws.cell(r, 3).font = font
        ws.cell(r, 4).value = row.get("qty"); ws.cell(r, 4).font = font
        ws.cell(r, 5).value = row.get("rate"); ws.cell(r, 5).font = NORM
        ws.cell(r, 5).number_format = "#,##0.00"
        ws.cell(r, 6).value = row.get("amount"); ws.cell(r, 6).font = NORM
        ws.cell(r, 6).number_format = "#,##0.00"
        ws.cell(r, 7).value = conf; ws.cell(r, 7).font = NORM
        if conf in CONF_FILL:
            ws.cell(r, 7).fill = CONF_FILL[conf]
        elif is_hdr:
            for c in range(1, 10):
                ws.cell(r, c).fill = HDR_FILL
        ws.cell(r, 8).value = row.get("source", ""); ws.cell(r, 8).font = NORM
        ws.cell(r, 9).value = row.get("notes", ""); ws.cell(r, 9).font = NORM
        ws.cell(r, 9).alignment = Alignment(wrap_text=True)

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 62
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 11
    ws.column_dimensions["H"].width = 22
    ws.column_dimensions["I"].width = 65


# ─── Sheet 3: Room Schedule ───────────────────────────────────────────────────

def _write_room_schedule(ws, merged: dict, openpyxl, project_mode: str = "custom_project") -> None:
    from openpyxl.styles import Font, Alignment
    _hdr(ws, 1, "ROOM SCHEDULE", openpyxl, 5)
    headers = ["Room Name", "Level", "Floor Area (m²)", "Source", "Notes"]
    _col_headers(ws, 2, headers, openpyxl)

    NORM = Font(size=10)
    GREY = Font(italic=True, size=10, color="808080")

    # Priority 1: DWG/PDF extracted rooms (in merged["geometry"]["rooms"])
    rooms = merged.get("geometry", {}).get("rooms", []) or []

    fallback_used = ""
    if not rooms:
        # Priority 2: standard geometry fallback — only for standard_model projects
        fallback_rooms = _load_std_geo_rooms(openpyxl, project_mode)
        if fallback_rooms:
            rooms = fallback_rooms
            std_path = _resolve_std_geo_path()
            fallback_used = f"standard_geometry ({std_path.name if std_path else 'unknown'})"

    r = 3
    for room in rooms:
        name = room.get("name") or room.get("room_name") or ""
        area = safe_float(room.get("area_m2") or room.get("floor_area_m2"))
        drawing_src = room.get("source") or room.get("drawing_ref") or ""
        level = room.get("level") or _get_room_level(name, drawing_src)
        src   = room.get("source") or "dwg"
        notes = room.get("notes") or ""

        ws.cell(r, 1).value = name;  ws.cell(r, 1).font = NORM
        ws.cell(r, 2).value = level; ws.cell(r, 2).font = NORM
        ws.cell(r, 3).value = area;  ws.cell(r, 3).font = NORM
        ws.cell(r, 4).value = src;   ws.cell(r, 4).font = NORM
        ws.cell(r, 5).value = notes; ws.cell(r, 5).font = NORM
        r += 1

    if not rooms:
        ws.cell(3, 1).value = "Not found — check DWG floor plan layers / PDF A-002 A-003"
        ws.cell(3, 1).font  = GREY
        r = 4

    # Totals row
    geo = merged.get("geometry", {})
    total_area = safe_float(geo.get("total_floor_area_m2"))
    if total_area:
        ws.cell(r + 1, 1).value = "TOTAL FLOOR AREA"
        ws.cell(r + 1, 1).font  = Font(bold=True, size=10)
        ws.cell(r + 1, 3).value = total_area
        ws.cell(r + 1, 3).font  = Font(bold=True, size=10)
        src_lbl = merged.get("audit", {}).get("total_floor_area_m2", {}).get("source", "dwg")
        ws.cell(r + 1, 4).value = src_lbl
        ws.cell(r + 1, 4).font  = Font(italic=True, size=10)

    if fallback_used:
        ws.cell(r + 3, 1).value = f"⚠ Room data from fallback: {fallback_used}"
        ws.cell(r + 3, 1).font  = Font(italic=True, color="7F3F00", size=10)

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 40


# ─── Sheet 4: Door Schedule ───────────────────────────────────────────────────

def _write_door_schedule(ws, merged: dict, openpyxl) -> None:
    from openpyxl.styles import Font, Alignment
    _hdr(ws, 1, "DOOR SCHEDULE", openpyxl, 8)
    headers = ["Mark", "Width (mm)", "Height (mm)", "Qty", "Core Type", "Location / Note", "Frame Width (mm)", "Source"]
    _col_headers(ws, 2, headers, openpyxl)

    NORM = Font(size=10)
    GREY = Font(italic=True, size=10, color="808080")
    WARN = Font(italic=True, size=10, color="7F3F00")

    # Prefer PDF schedule doors (marks, types, hardware) over bare DWG blocks
    pdf_sched = merged.get("pdf_schedule_doors") or []
    dwg_doors = merged.get("doors") or []
    doors_src = merged.get("audit", {}).get("doors_source", "none")

    if pdf_sched:
        display_doors = pdf_sched
        display_src   = "pdf_schedule (A-017)"
        note_suffix   = f"  DWG block count (authoritative): {len(dwg_doors)}"
    else:
        display_doors = dwg_doors
        display_src   = doors_src
        note_suffix   = ""

    # BOQ stock-code → authoritative core type (Fix 9 generic)
    _DOOR_CORE_MAP = {920: "Solid Core", 820: "Hollow Core", 720: "Hollow Core"}

    def _get_core_type(width_mm, pdf_core: str, stock_code: str = "") -> str:
        """Generic door core type.

        Priority:
        1. Stock code suffix — DSC = Solid Core, DHC = Hollow Core (Rhodes catalogue)
        2. PDF schedule extraction
        3. Width-based mapping as fallback
        """
        sc = (stock_code or "").upper()
        if "DSC" in sc:
            return "Solid Core"
        if "DHC" in sc:
            return "Hollow Core"
        # Width-based BOQ mapping overrides PDF (PDF may have OCR errors)
        w = int(safe_float(width_mm) or 0)
        boq_core = _DOOR_CORE_MAP.get(w)
        if boq_core:
            return boq_core
        return pdf_core or ""

    r = 3
    for door in display_doors:
        mark = door.get("mark") or door.get("id") or f"D{r-2}"
        # PDF schedule stores leaf_width_mm (may be null for double/sliding)
        width = door.get("leaf_width_mm") or door.get("width_mm") or door.get("width")
        height = door.get("height_mm") or door.get("height")
        qty    = int(safe_float(door.get("qty")) or 1)
        core   = _get_core_type(width, door.get("core_type") or "", door.get("stock_code") or "")
        loc    = door.get("location_note") or ""
        frame_w = door.get("frame_width_mm_if_shown") or ""
        src    = door.get("_source") or door.get("source_type") or door.get("source") or display_src

        ws.cell(r, 1).value = mark;    ws.cell(r, 1).font = NORM
        ws.cell(r, 2).value = width if width is not None else "— (see frame width)"
        ws.cell(r, 2).font  = WARN if width is None else NORM
        ws.cell(r, 3).value = height;  ws.cell(r, 3).font = NORM
        ws.cell(r, 4).value = qty;     ws.cell(r, 4).font = NORM
        ws.cell(r, 5).value = core;    ws.cell(r, 5).font = NORM
        ws.cell(r, 6).value = loc;     ws.cell(r, 6).font = NORM
        ws.cell(r, 6).alignment = Alignment(wrap_text=True)
        ws.cell(r, 7).value = frame_w; ws.cell(r, 7).font = NORM
        ws.cell(r, 8).value = src;     ws.cell(r, 8).font = NORM
        r += 1

    if not display_doors:
        ws.cell(3, 1).value = "Not found — check PDF A-017 Door & Window Schedule"
        ws.cell(3, 1).font  = GREY
        r = 4

    ws.cell(r + 1, 1).value = f"Source: {display_src}{note_suffix}"
    ws.cell(r + 1, 1).font  = Font(italic=True, size=10)

    for col, w in zip("ABCDEFGH", [8, 22, 12, 6, 18, 45, 16, 35]):
        ws.column_dimensions[col].width = w


# ─── Sheet 5: Window Schedule ─────────────────────────────────────────────────

def _write_window_schedule(ws, merged: dict, openpyxl) -> None:
    from openpyxl.styles import Font, Alignment
    _hdr(ws, 1, "WINDOW SCHEDULE", openpyxl, 7)
    headers = ["Mark", "Width (mm)", "Height (mm)", "Qty", "Type", "Qty Note", "Source"]
    _col_headers(ws, 2, headers, openpyxl)

    windows = merged.get("windows", []) or []
    NORM = Font(size=10)
    GREY = Font(italic=True, size=10, color="808080")
    MED  = Font(italic=True, size=10, color="7F3F00")

    r = 3
    for win in windows:
        mark  = win.get("mark") or win.get("id") or f"W{r-2}"
        width = win.get("width_mm") or win.get("width")
        height = win.get("height_mm") or win.get("height")
        qty   = win.get("qty")
        wtype = win.get("type_note") or win.get("type") or win.get("type_mapped") or ""
        qty_note = win.get("qty_note") or ""
        src   = win.get("_source") or win.get("source_type") or win.get("source") or ""
        conf  = win.get("confidence") or "HIGH"

        ws.cell(r, 1).value = mark;     ws.cell(r, 1).font = NORM
        ws.cell(r, 2).value = width;    ws.cell(r, 2).font = NORM
        ws.cell(r, 3).value = height;   ws.cell(r, 3).font = NORM
        ws.cell(r, 4).value = qty;      ws.cell(r, 4).font = MED if conf == "MEDIUM" else NORM
        ws.cell(r, 5).value = wtype;    ws.cell(r, 5).font = NORM
        ws.cell(r, 6).value = qty_note; ws.cell(r, 6).font = GREY
        ws.cell(r, 6).alignment = Alignment(wrap_text=True)
        ws.cell(r, 7).value = src;      ws.cell(r, 7).font = NORM
        r += 1

    if not windows:
        ws.cell(3, 1).value = "Not found — check PDF A-017 Door & Window Schedule"
        ws.cell(3, 1).font  = GREY
        r = 4

    wins_src = merged.get("audit", {}).get("windows_source", "none")
    dwg_count = merged.get("geometry", {}).get("window_count")
    note_parts = [f"Source: {wins_src}"]
    if dwg_count:
        note_parts.append(f"DWG block count: {dwg_count}")
    ws.cell(r + 1, 1).value = "  ".join(note_parts)
    ws.cell(r + 1, 1).font  = Font(italic=True, size=10)

    for col, w in zip("ABCDEFG", [8, 12, 12, 8, 20, 40, 35]):
        ws.column_dimensions[col].width = w


# ─── Sheet 6: Finish Schedule ─────────────────────────────────────────────────

def _write_finish_schedule(ws, merged: dict, openpyxl, project_mode: str = "custom_project") -> None:
    from openpyxl.styles import Font, Alignment
    _hdr(ws, 1, "FINISH SCHEDULE", openpyxl, 7)
    headers = ["Room", "Level", "Floor Area (m²)", "Floor Finish", "Wall Finish", "Ceiling Finish", "Source"]
    _col_headers(ws, 2, headers, openpyxl)

    NORM = Font(size=10)
    GREY = Font(italic=True, size=10, color="808080")

    # Priority 1: merged finishes from PDF (finishes list has room + finish fields)
    finishes = merged.get("finishes", []) or []

    # Priority 2: load from reference file — only for standard_model projects
    if not finishes:
        finishes = _load_proj_summary_rooms(openpyxl, project_mode)

    # Build room-area lookup to fill in missing areas (default_finishes have area=0)
    rooms_list = (merged.get("geometry", {}).get("rooms", []) or
                  _load_std_geo_rooms(openpyxl, project_mode))
    room_area_by_name: dict[str, float] = {}
    for rm in rooms_list:
        name = (rm.get("name") or "").strip()
        area = safe_float(rm.get("area_m2") or rm.get("floor_area_m2")) or 0.0
        if name and area > 0:
            room_area_by_name[name.lower()] = area

    def _lookup_area(room_name: str, stored: Any) -> Any:
        """Return stored area if >0, else look up from room_area_by_name."""
        a = safe_float(stored) or 0.0
        if a > 0:
            return a
        return room_area_by_name.get((room_name or "").lower(), None)

    r = 3
    if finishes:
        for fin in finishes:
            room    = fin.get("room") or fin.get("name") or fin.get("space") or ""
            level   = fin.get("level") or ""
            area    = _lookup_area(room, fin.get("area_m2") or fin.get("area"))
            floor_f = fin.get("floor_finish") or fin.get("floor") or ""
            wall_f  = fin.get("wall_finish") or fin.get("wall") or ""
            ceil_f  = fin.get("ceiling_finish") or fin.get("ceiling") or ""
            src     = fin.get("source") or fin.get("source_type") or ""

            ws.cell(r, 1).value = room;    ws.cell(r, 1).font = NORM
            ws.cell(r, 2).value = level;   ws.cell(r, 2).font = NORM
            ws.cell(r, 3).value = area;    ws.cell(r, 3).font = NORM
            ws.cell(r, 4).value = floor_f; ws.cell(r, 4).font = NORM
            ws.cell(r, 4).alignment = Alignment(wrap_text=True)
            ws.cell(r, 5).value = wall_f;  ws.cell(r, 5).font = NORM
            ws.cell(r, 5).alignment = Alignment(wrap_text=True)
            ws.cell(r, 6).value = ceil_f;  ws.cell(r, 6).font = NORM
            ws.cell(r, 6).alignment = Alignment(wrap_text=True)
            ws.cell(r, 7).value = src;     ws.cell(r, 7).font = GREY
            r += 1
    else:
        ws.cell(3, 1).value = "Not found — check PDF A-020 Wall Schedule, A-021 Floor Schedule"
        ws.cell(3, 1).font  = GREY
        r = 4

    # Computed finish totals from room areas
    wet_rooms  = {"Bathroom", "Laundry", "WC", "Toilet"}
    dry_rooms  = {"Bedroom 1 (Master)", "Bedroom 1", "Bedroom 2", "Bedroom 3",
                  "Kitchen", "Living / Dining", "Corridor / Hallway", "Corridor"}
    vinyl_total    = sum(safe_float(f.get("area_m2") or 0) for f in finishes
                         if any(k in (f.get("name") or f.get("room") or "") for k in dry_rooms))
    ceramic_total  = sum(safe_float(f.get("area_m2") or 0) for f in finishes
                         if any(k in (f.get("name") or f.get("room") or "") for k in wet_rooms))
    verandah_total = sum(safe_float(f.get("area_m2") or 0) for f in finishes
                         if "verandah" in (f.get("name") or f.get("room") or "").lower()
                         or "deck" in (f.get("name") or f.get("room") or "").lower())

    if vinyl_total or ceramic_total or verandah_total:
        r += 1
        ws.cell(r, 1).value = "COMPUTED TOTALS (from room areas)"; ws.cell(r, 1).font = Font(bold=True, size=10)
        r += 1
        ws.cell(r, 1).value = "Vinyl Plank (dry rooms)";   ws.cell(r, 1).font = NORM
        ws.cell(r, 3).value = round(vinyl_total, 1);       ws.cell(r, 3).font = NORM
        ws.cell(r, 4).value = "m²"; r += 1
        ws.cell(r, 1).value = "Ceramic Tile (wet rooms)";  ws.cell(r, 1).font = NORM
        ws.cell(r, 3).value = round(ceramic_total, 1);     ws.cell(r, 3).font = NORM
        ws.cell(r, 4).value = "m²"; r += 1
        ws.cell(r, 1).value = "WPC Decking (verandah)";    ws.cell(r, 1).font = NORM
        ws.cell(r, 3).value = round(verandah_total, 1);    ws.cell(r, 3).font = NORM
        ws.cell(r, 4).value = "m²"

    for col, w in zip("ABCDEFG", [28, 14, 14, 28, 30, 24, 30]):
        ws.column_dimensions[col].width = w


# ─── Sheet 7: Structural Summary ─────────────────────────────────────────────

def _write_structural_summary(ws, merged: dict, openpyxl) -> None:
    from openpyxl.styles import Font
    _hdr(ws, 1, "STRUCTURAL SUMMARY", openpyxl, 4)

    BOLD = Font(bold=True, size=11)
    NORM = Font(size=11)
    GREY = Font(italic=True, size=10, color="808080")
    SEC  = Font(bold=True, size=11, color=BLUE_HEX)

    def kv(r, k, v, unit="", src=""):
        ws.cell(r, 1).value = k;    ws.cell(r, 1).font = BOLD
        ws.cell(r, 2).value = v if v not in (None, "—", "") else "Not found in drawings"
        ws.cell(r, 2).font  = NORM if v not in (None, "—", "") else GREY
        ws.cell(r, 3).value = unit; ws.cell(r, 3).font = NORM
        ws.cell(r, 4).value = f"[{src}]" if src else ""; ws.cell(r, 4).font = GREY

    def sec(r, label):
        ws.cell(r, 1).value = label; ws.cell(r, 1).font = SEC

    geo    = merged.get("geometry", {})
    struct = merged.get("structural", {})
    audit  = merged.get("audit", {})

    row = 2
    _col_headers(ws, row, ["Item", "Value", "Unit", "Source"], openpyxl); row += 1

    sec(row, "GEOMETRY (from DWG)"); row += 1
    kv(row, "Total Floor Area",     geo.get("total_floor_area_m2"),
       "m²", audit.get("total_floor_area_m2", {}).get("source", "dwg")); row += 1
    kv(row, "Building Length",      geo.get("building_length_m"),
       "m", audit.get("building_length_m", {}).get("source", "dwg")); row += 1
    kv(row, "Building Width",       geo.get("building_width_m"),
       "m", audit.get("building_width_m", {}).get("source", "dwg")); row += 1
    kv(row, "External Wall Length", geo.get("external_wall_length_m"),
       "lm", audit.get("external_wall_length_m", {}).get("source", "dwg")); row += 1
    kv(row, "Internal Wall Length", geo.get("internal_wall_length_m"),
       "lm", audit.get("internal_wall_length_m", {}).get("source", "dwg")); row += 1
    kv(row, "Total Wall Length",    geo.get("total_wall_length_m"),
       "lm", "dwg_derived"); row += 1
    kv(row, "Roof Area",            geo.get("roof_area_m2"),
       "m²", audit.get("roof_area_m2", {}).get("source", "dwg")); row += 1
    kv(row, "Roof Pitch",           geo.get("roof_pitch_degrees"),
       "°", audit.get("roof_pitch_degrees", {}).get("source", "dwg")); row += 1
    kv(row, "Post Count",           geo.get("post_count"),
       "posts", "dwg"); row += 1
    kv(row, "Door Count (DWG)",     geo.get("door_count"),
       "doors", "dwg_blocks"); row += 1
    kv(row, "Window Count (DWG)",   geo.get("window_count"),
       "windows", "dwg_blocks"); row += 1
    row += 1

    sec(row, "FRAMING (from Framecad BOM — BLANK if no BOM loaded)"); row += 1
    kv(row, "Wall Frame",       struct.get("wall_frame_lm") or "Not in BOM",
       "lm", struct.get("wall_frame_source", "none")); row += 1
    kv(row, "Roof Batten",      struct.get("roof_batten_lm") or "Not in BOM",
       "lm", struct.get("roof_batten_source", "none")); row += 1
    kv(row, "Ceiling Batten",   struct.get("ceiling_batten_lm") or "Not in BOM",
       "lm", struct.get("ceiling_batten_source", "none")); row += 1
    kv(row, "Roof Truss Qty",   struct.get("roof_truss_qty") or "Not in BOM",
       "qty", struct.get("roof_truss_source", "none")); row += 1
    kv(row, "Floor Panel Qty",  struct.get("floor_panel_qty") or "Not in BOM",
       "set", struct.get("floor_panel_source", "none")); row += 1
    row += 1

    sec(row, "DERIVED ROOF COMPONENT LENGTHS (from DWG geometry)"); row += 1
    kv(row, "Ridge Length",   geo.get("ridge_length_m"), "lm", "dwg_derived"); row += 1
    kv(row, "Barge Length",   geo.get("barge_length_m"), "lm", "dwg_derived"); row += 1
    kv(row, "Gutter Length",  geo.get("gutter_length_m"), "lm", "dwg_derived"); row += 1
    kv(row, "Fascia Length",  geo.get("fascia_length_m"), "lm", "dwg_derived"); row += 1
    kv(row, "Apron Length",   geo.get("apron_length_m"),  "lm", "dwg_derived"); row += 1

    # BOM warnings
    bom_warnings = struct.get("bom_warnings", [])
    if bom_warnings:
        row += 1
        ws.cell(row, 1).value = f"BOM warnings: {'; '.join(str(w) for w in bom_warnings[:3])}"
        ws.cell(row, 1).font  = GREY

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 24


# ─── Sheet 8: Services Summary ───────────────────────────────────────────────

_ROOM_FIXTURE_MAP = {
    "Bathroom": {
        "keywords": ["toilet suite", "toilet roll", "bathroom mirror", "vanity basin",
                     "hand basin", "soap holder", "shower tray", "shower curtain",
                     "shower tap", "towel rail"],
        "stock_prefixes": ["50129-WCP", "50129-PAP", "50129-MIR", "50129-VAN",
                           "50129-TSV", "50129-SOA", "50129-SHW", "50129-SHC",
                           "50129-TSS", "50129-RAI"],
    },
    "Kitchen": {
        "keywords": ["kitchen sink", "kitchen tap", "base cabinet", "drawer cabinet",
                     "wall cabinet", "bench top", "rangehood", "corner base cabinet"],
        "stock_prefixes": ["50129-TSK", "50129-KSS-002", "50129-JOI", "50129-BEN"],
    },
    "Laundry": {
        "keywords": ["laundry sink", "laundry tap", "washing machine"],
        "stock_prefixes": ["50129-KSS-007", "50129-TSL"],
    },
}


def _match_fixture_room(sc: str, desc: str) -> str | None:
    """Return the first room name that matches this fixture (stock prefix takes priority).

    First-match-wins prevents double-counting across rooms.
    Returns None if item does not belong to any room.
    """
    for room_name, cfg in _ROOM_FIXTURE_MAP.items():
        if any(sc.startswith(p) for p in cfg["stock_prefixes"]):
            return room_name
    for room_name, cfg in _ROOM_FIXTURE_MAP.items():
        if any(kw in desc for kw in cfg["keywords"]):
            return room_name
    return None


def _write_fixture_summary_by_room(ws, start_row: int, boq_rows: list[dict], openpyxl) -> None:
    """Write fixture-by-room grouped section to the Services Summary sheet."""
    from openpyxl.styles import Font, PatternFill

    BLUE_FILL = PatternFill("solid", fgColor=BLUE_HEX)
    WHITE     = Font(bold=True, size=11, color="FFFFFF")
    BOLD      = Font(bold=True, size=10)
    NORM      = Font(size=10)

    r = start_row
    # Section header
    ws.cell(r, 1).value = "FIXTURE SUMMARY BY ROOM"
    ws.cell(r, 1).font  = WHITE
    for c in range(1, 6):
        ws.cell(r, c).fill = BLUE_FILL
    r += 1

    headers = ["Room", "Fixture Description", "Stock Code", "QTY", "Source"]
    for col, h in enumerate(headers, 1):
        ws.cell(r, col).value = h
        ws.cell(r, col).font  = BOLD
    r += 1

    # Pre-assign each item to exactly one room (first-match-wins)
    item_rows = [rw for rw in boq_rows if not rw.get("_is_header")]
    by_room: dict[str, list[dict]] = {rn: [] for rn in _ROOM_FIXTURE_MAP}
    for rw in item_rows:
        sc   = str(rw.get("stock_code") or "")
        desc = str(rw.get("description") or "").lower()
        room = _match_fixture_room(sc, desc)
        if room:
            by_room[room].append(rw)

    for room_name in _ROOM_FIXTURE_MAP:
        room_fixtures = by_room[room_name]
        if not room_fixtures:
            continue

        # Room sub-header
        ws.cell(r, 1).value = room_name
        ws.cell(r, 1).font  = BOLD
        r += 1

        for fix in room_fixtures:
            ws.cell(r, 1).value = ""
            ws.cell(r, 2).value = fix.get("description", ""); ws.cell(r, 2).font = NORM
            ws.cell(r, 3).value = fix.get("stock_code", "");  ws.cell(r, 3).font = NORM
            ws.cell(r, 4).value = fix.get("qty");              ws.cell(r, 4).font = NORM
            ws.cell(r, 5).value = fix.get("source", "");      ws.cell(r, 5).font = NORM
            r += 1

        r += 1  # blank row between rooms


_SERVICE_KEYWORDS = (
    "hydraul", "plumb", "pipe", "pex", "pvc dwv", "toilet", "basin",
    "shower", "sink", "wash", "tap", "gully", "drain", "sewer",
    "downpipe", "inspection", "electrical", "circuit", "rcbo", "mcb",
    "switchboard", "cable", "conduit", "socket", "power point", "light",
    "fan", "earth", "solar", "pump", "hot water",
)

_SERVICE_CODES = ("50117", "50118", "50119", "50125", "50129")


def _write_services_summary(ws, boq_rows: list[dict], boq_items: list[dict], openpyxl) -> None:
    from openpyxl.styles import Font
    _hdr(ws, 1, "SERVICES SUMMARY", openpyxl, 5)
    headers = ["Stock Code", "Description", "QTY", "Unit", "Source"]
    _col_headers(ws, 2, headers, openpyxl)

    NORM = Font(size=10)
    GREY = Font(italic=True, size=10, color="808080")

    def _is_service(row: dict) -> bool:
        sc   = str(row.get("stock_code") or "").lower()
        desc = str(row.get("description") or "").lower()
        if any(sc.startswith(p) for p in _SERVICE_CODES):
            return True
        return any(kw in desc for kw in _SERVICE_KEYWORDS)

    r = 3
    for row in boq_rows:
        if row.get("_is_header"):
            continue
        if _is_service(row):
            ws.cell(r, 1).value = row.get("stock_code", ""); ws.cell(r, 1).font = NORM
            ws.cell(r, 2).value = row.get("description", ""); ws.cell(r, 2).font = NORM
            ws.cell(r, 3).value = row.get("qty"); ws.cell(r, 3).font = NORM
            ws.cell(r, 4).value = row.get("unit", ""); ws.cell(r, 4).font = NORM
            ws.cell(r, 5).value = row.get("source", ""); ws.cell(r, 5).font = NORM
            r += 1

    if r == 3:
        ws.cell(3, 1).value = "No services items matched in BOQ — check stock codes & descriptions"
        ws.cell(3, 1).font  = GREY

    # ── Fixture summary by room (Fix 6) ────────────────────────────────────
    r += 2
    _write_fixture_summary_by_room(ws, r, boq_rows, openpyxl)

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 62
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 20


# ─── Sheet 9: Drawing Register ───────────────────────────────────────────────

def _write_drawing_register(ws, files_found: dict, merged: dict, openpyxl) -> None:
    from openpyxl.styles import Font, Alignment, PatternFill
    _hdr(ws, 1, "DRAWING REGISTER", openpyxl, 6)
    headers = ["Type", "File Name", "Size (KB)", "Pages", "Data Extracted", "Status"]
    _col_headers(ws, 2, headers, openpyxl)

    NORM  = Font(size=10)
    GREY  = Font(italic=True, size=10, color="808080")
    GREEN = PatternFill("solid", fgColor="C6EFCE")
    AMBER = PatternFill("solid", fgColor="FFEB9C")
    RED   = PatternFill("solid", fgColor="FFC7CE")

    DATA_EXTRACTED = {
        "dwg": "Geometry: floor area, wall lengths, roof dims, door/window/post counts",
        "dxf": "Geometry (converted from DWG format)",
        "pdf": "Door/window schedule, room data, structural notes, finishes",
        "bom": "Framecad structural quantities (wall, batten, truss, floor panels)",
        "ifc": "3D model — structural quantities",
    }

    audit = merged.get("audit", {})

    r = 3
    for cat, entries in files_found.items():
        if cat in ("warnings",) or not entries:
            continue
        for e in entries:
            fname = Path(e["path"]).name
            size  = e.get("size_kb", "")
            pages = e.get("pages", "")

            # Determine what was actually extracted from this file
            extracted = DATA_EXTRACTED.get(cat, "")
            # For PDFs, note if door_schedule was found
            if cat == "pdf":
                has_doors  = bool(merged.get("doors"))
                has_wins   = bool(merged.get("windows"))
                has_finish = bool(merged.get("finishes"))
                parts = []
                if has_doors:   parts.append(f"{len(merged['doors'])} door marks")
                if has_wins:    parts.append(f"{len(merged['windows'])} window marks")
                if has_finish:  parts.append("finish schedule")
                extracted = ", ".join(parts) if parts else "general notes only"
            elif cat in ("dwg", "dxf"):
                geo = merged.get("geometry", {})
                parts = []
                if geo.get("total_floor_area_m2"):
                    parts.append(f"floor area {geo['total_floor_area_m2']}m²")
                if geo.get("external_wall_length_m"):
                    parts.append(f"ext walls {geo['external_wall_length_m']}lm")
                if geo.get("door_count"):
                    parts.append(f"{geo['door_count']} doors")
                if geo.get("window_count"):
                    parts.append(f"{geo['window_count']} windows")
                extracted = ", ".join(parts) if parts else "No geometry extracted"
            elif cat == "bom":
                struct = merged.get("structural", {})
                parts = []
                if struct.get("wall_frame_lm"):
                    parts.append(f"wall frame {struct['wall_frame_lm']}lm")
                if struct.get("roof_batten_lm"):
                    parts.append(f"roof battens {struct['roof_batten_lm']}lm")
                extracted = ", ".join(parts) if parts else "No BOM data extracted"

            # Status
            if extracted and "only" not in extracted and "No" not in extracted:
                status = "DATA FOUND"
                fill   = GREEN
            elif "only" in extracted or cat == "pdf":
                status = "PARTIAL"
                fill   = AMBER
            else:
                status = "NO DATA"
                fill   = RED

            ws.cell(r, 1).value = cat.upper(); ws.cell(r, 1).font = NORM
            ws.cell(r, 2).value = fname;       ws.cell(r, 2).font = NORM
            ws.cell(r, 3).value = size;        ws.cell(r, 3).font = NORM
            ws.cell(r, 4).value = pages;       ws.cell(r, 4).font = NORM
            ws.cell(r, 5).value = extracted;   ws.cell(r, 5).font = NORM
            ws.cell(r, 5).alignment = Alignment(wrap_text=True)
            ws.cell(r, 6).value = status;      ws.cell(r, 6).font = Font(bold=True, size=10)
            ws.cell(r, 6).fill  = fill
            r += 1

    if r == 3:
        ws.cell(3, 1).value = "No input files detected"
        ws.cell(3, 1).font  = GREY

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 55
    ws.column_dimensions["F"].width = 12
