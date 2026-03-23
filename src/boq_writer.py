"""
boq_writer.py — Write final BOQ Excel workbook.

Primary mode: copy data/approved_boq_G303.xlsx (G303-BOQ sheet) as template,
rename sheet to "BOQ", fill/override col C from drawings, add cols E-I for
rate / amount / confidence / source / notes.

Fallback (no approved BOQ): fresh 11-column workbook.

Output: output/boq/[ProjectName]_BOQ_[YYYYMMDD].xlsx
"""

from __future__ import annotations
import logging
import re
from copy import copy
from datetime import date
from pathlib import Path

from src.config import OUTPUT_BOQ, DATA_DIR
from src.utils import safe_float

log = logging.getLogger("boq.writer")

# ── Extra column widths (E-I added to approved BOQ's A-D) ────────────────────
EXTRA_COL_WIDTHS = {
    "E": 14.0, "F": 16.0, "G": 12.0, "H": 22.0, "I": 44.0,
}
EXTRA_COL_HEADERS = {
    5: "RATE (PGK)", 6: "AMOUNT (PGK)",
    7: "CONFIDENCE", 8: "SOURCE", 9: "NOTES / DRAWING REF",
}
CONF_TEXT_COLORS = {
    "HIGH":   "375623",   # dark green
    "MEDIUM": "7F3F00",   # amber
    "LOW":    "C00000",   # red
}

# Header row in G303-BOQ sheet (rows 1-8 are title/headers; data starts row 9)
APPROVED_BOQ_HEADER_ROW = 8
APPROVED_BOQ_DATA_START  = 9

# Fallback full-workbook columns (no-template mode)
BOQ_COLS = [
    "Item No", "Stock Code", "Description", "Unit", "Qty",
    "Rate (PGK)", "Amount (PGK)", "Confidence", "Qty Basis",
    "Source", "Rule / Method", "Notes",
]
CONF_BG_COLOURS = {
    "HIGH": "C6EFCE", "MEDIUM": "FFEB9C", "LOW": "FFC7CE",
}
QTY_BASIS_COLOURS = {
    "measured":      "C6EFCE",   # green
    "derived":       "FFEB9C",   # amber
    "provisional":   "FFC7CE",   # light red
    "manual_review": "D9D9D9",   # grey
}


# ─── Public entry point ───────────────────────────────────────────────────────

def write_boq(
    project_name: str,
    boq_items: list[dict],
    validation: dict,
    merged: dict,
    project_mode: str = "custom_project",
    approved_boq_path: str | None = None,
    reference_path: str | None = None,
    structural_baseline: list[dict] | None = None,
) -> Path:
    """Write final BOQ Excel workbook.

    Args:
        project_name:      Job name string.
        boq_items:         Calculated BOQ item list.
        validation:        Validator output dict.
        merged:            Merged project data dict.
        project_mode:      "standard_model" | "custom_project".
                           custom_project → always writes full fresh workbook.
                           standard_model → may use approved BOQ template if file exists.
        approved_boq_path: Path to approved BOQ xlsx (used only for standard_model mode).
        reference_path:    (unused, kept for backward compat)
        structural_baseline: (unused, kept for backward compat)
    """
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl not installed")

    date_str    = date.today().strftime("%Y%m%d")
    output_path = OUTPUT_BOQ / f"{project_name}_BOQ_{date_str}.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # project_mode is now the authoritative routing key.
    # approved_boq_path is only used as the template source when mode = standard_model.
    effective_mode = project_mode

    if effective_mode == "standard_model":
        # Only use G303 template if the approved BOQ file actually exists
        if not approved_boq_path:
            for candidate in [
                DATA_DIR / "approved_boq_G303.xlsx",
                DATA_DIR / "approved_boq.xlsx",
            ]:
                if candidate.exists():
                    approved_boq_path = str(candidate)
                    break

        if approved_boq_path and Path(approved_boq_path).exists():
            log.info("Using approved BOQ as template: %s", Path(approved_boq_path).name)
            total = _write_from_approved_boq(
                approved_boq_path, output_path, project_name,
                boq_items, validation, merged, openpyxl,
            )
        else:
            log.warning(
                "standard_model mode but no approved BOQ found — writing full workbook"
            )
            total = _write_full_workbook(
                output_path, project_name, boq_items, validation, merged, openpyxl,
            )
    else:
        # custom_project: always use full fresh workbook — do NOT copy G303 template
        log.info("custom_project mode — writing full workbook (no G303 template)")
        total = _write_full_workbook(
            output_path, project_name, boq_items, validation, merged, openpyxl,
        )

    log.info("BOQ written: %s  items=%d  total=PGK %.2f",
             output_path.name, len(boq_items), total)
    return output_path


# ─── Approved-BOQ–based writer ───────────────────────────────────────────────

def _write_from_approved_boq(
    approved_boq_path: str,
    output_path: Path,
    project_name: str,
    boq_items: list[dict],
    validation: dict,
    merged: dict,
    openpyxl,
) -> float:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # Load approved BOQ as the base workbook
    wb = openpyxl.load_workbook(str(approved_boq_path))

    # Rename G303-BOQ → BOQ; remove Sheet1
    ws = wb["G303-BOQ"]
    ws.title = "BOQ"
    if "Sheet1" in wb.sheetnames:
        del wb["Sheet1"]

    # ── Column widths for new cols E-I ─────────────────────────────────────
    for col, w in EXTRA_COL_WIDTHS.items():
        ws.column_dimensions[col].width = w

    thin    = Side(style="thin")
    all_bdr = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Add E-I column headers in header row (row 8) ───────────────────────
    ref_hdr  = ws.cell(APPROVED_BOQ_HEADER_ROW, 1)
    hdr_font = Font(
        name=ref_hdr.font.name if ref_hdr.font else "Calibri",
        size=ref_hdr.font.size if ref_hdr.font else 11,
        bold=True,
        color="FFFFFF",
    )
    hdr_fill = copy(ref_hdr.fill) if (ref_hdr.fill and ref_hdr.fill.fill_type) else \
               PatternFill("solid", fgColor="2F5496")

    for col_idx, label in EXTRA_COL_HEADERS.items():
        wc           = ws.cell(APPROVED_BOQ_HEADER_ROW, col_idx)
        wc.value     = label
        wc.font      = copy(hdr_font)
        wc.fill      = copy(hdr_fill)
        wc.border    = copy(all_bdr)
        wc.alignment = Alignment(
            horizontal="left" if col_idx in (9,) else "center",
            vertical="center",
        )

    # ── Update title cells ─────────────────────────────────────────────────
    for trow in range(1, APPROVED_BOQ_HEADER_ROW):
        for tcol in range(1, ws.max_column + 1):
            cell = ws.cell(trow, tcol)
            if cell.value and isinstance(cell.value, str):
                if "Alice Kivali" in cell.value or "24138" in cell.value:
                    cell.value = project_name
                if "25_07_2025" in cell.value or "25/07/2025" in cell.value:
                    cell.value = date.today().strftime("%d/%m/%Y")

    # ── Build boq_item lookups ─────────────────────────────────────────────
    # Primary: positional match by Excel row index (fixes duplicate-description issues)
    boq_by_row: dict[int, dict] = {}
    item_by_desc: dict[str, dict] = {}
    items_by_code: dict[str, list[dict]] = {}
    for item in boq_items:
        ri = item.get("_row_idx")
        if ri:
            boq_by_row[int(ri)] = item
        sc = (item.get("stock_code") or "").strip().upper()
        if sc:
            items_by_code.setdefault(sc, []).append(item)
        dn = _norm(item.get("description") or "")
        if dn:
            item_by_desc[dn] = item

    # ── Body font (read from first data row) ───────────────────────────────
    ref_body = ws.cell(APPROVED_BOQ_DATA_START, 2)
    body_font = Font(
        name=ref_body.font.name if ref_body.font else "Calibri",
        size=ref_body.font.size if ref_body.font else 11,
    )

    grand_total = 0.0

    # Track current section and seen items for deduplication
    current_section = "MAIN"
    _seen_items: set[tuple[str, str, str]] = set()   # (section, stock_code, desc_norm_short)

    # ── Iterate all data rows ──────────────────────────────────────────────
    for row_idx in range(APPROVED_BOQ_DATA_START, ws.max_row + 1):
        sc_val      = ws.cell(row_idx, 1).value
        desc_val    = ws.cell(row_idx, 2).value
        approved_qty = ws.cell(row_idx, 3).value
        # unit_val  = ws.cell(row_idx, 4).value  (kept unchanged)

        # Blank rows → skip
        if sc_val is None and desc_val is None:
            continue

        # Section / sub-section headers: no qty (col C is None)
        is_section = (sc_val is None and approved_qty is None and desc_val is not None)
        if is_section:
            # Track current section for deduplication
            _sec_lower = str(desc_val or "").lower()
            if any(k in _sec_lower for k in ["laundry", "ground level", "ground floor"]):
                current_section = "LAUNDRY"
            else:
                current_section = "MAIN"
            # Add empty bordered E-I cells with matching section fill
            sec_fill = ws.cell(row_idx, 2).fill
            sec_font = ws.cell(row_idx, 2).font
            for col_idx in range(5, 10):
                wc = ws.cell(row_idx, col_idx)
                wc.value = None
                if sec_fill and sec_fill.fill_type:
                    wc.fill = copy(sec_fill)
                if sec_font:
                    wc.font = copy(sec_font)
                wc.border = copy(all_bdr)
            continue

        # ── Item / sub-item row ─────────────────────────────────────────────
        sc_key   = str(sc_val or "").strip().upper()
        desc_key = _norm(str(desc_val or ""))

        # Deduplication: same description within same section → clear qty (Fix 2)
        _dedup_key = (current_section, sc_key, desc_key[:60])
        _is_dup = _dedup_key in _seen_items
        _seen_items.add(_dedup_key)

        # Positional match first (exact row → item mapping), then description fallback
        boq_item = boq_by_row.get(row_idx) or _find_boq_item(sc_key, desc_key, item_by_desc, items_by_code)

        qty    = None
        rate   = None
        amount = None
        conf   = None
        source = None
        notes  = None

        # If this is a duplicate row, zero it out (keep for template structure but
        # don't contribute qty/amount — first occurrence already has the right value)
        if _is_dup:
            ws.cell(row_idx, 3).value = None   # clear qty in col C
            for _c in range(5, 10):
                wc = ws.cell(row_idx, _c)
                wc.value  = None
                wc.font   = copy(body_font)
                wc.border = copy(all_bdr)
            ws.cell(row_idx, 7).value = "DUP"
            ws.cell(row_idx, 9).value = "Duplicate row — see first occurrence above"
            ws.cell(row_idx, 9).font  = Font(name=body_font.name, size=body_font.size,
                                             italic=True, color="808080")
            continue

        if boq_item:
            calc_qty = boq_item.get("qty")
            if calc_qty is not None:
                qty  = safe_float(calc_qty)
                conf = (boq_item.get("confidence") or "").upper()
                source = boq_item.get("source") or ""
                notes  = _build_notes(boq_item)
            else:
                # qty=None means calculator hit TIER 4 (BLANK) — no project data available.
                # Do NOT fall back to approved BOQ qty — that is a different project.
                qty    = None
                conf   = "LOW"
                source = "none"
                notes  = "No source data — manual entry required"

            rate = safe_float(boq_item.get("rate"))
            if qty and rate:
                amount = round(qty * rate, 2)
                grand_total += amount

            # Override col C with calculated qty
            ws.cell(row_idx, 3).value = qty

        else:
            # No matching calculated item found for this template row.
            # Do NOT fall back to approved BOQ qty — that is a different project.
            # Leave qty blank so the estimator knows this needs manual entry.
            qty    = None
            conf   = "LOW"
            source = "none"
            notes  = "No source data — manual entry required"

        # ── Write cols E-I ──────────────────────────────────────────────────
        # E: Rate
        ec = ws.cell(row_idx, 5)
        ec.value        = rate
        ec.font         = copy(body_font)
        ec.border       = copy(all_bdr)
        ec.alignment    = Alignment(horizontal="right", vertical="center")
        ec.number_format = "#,##0.00"

        # F: Amount
        fc = ws.cell(row_idx, 6)
        fc.value        = amount
        fc.font         = copy(body_font)
        fc.border       = copy(all_bdr)
        fc.alignment    = Alignment(horizontal="right", vertical="center")
        fc.number_format = "#,##0.00"

        # G: Confidence (color-coded)
        gc = ws.cell(row_idx, 7)
        gc.value     = conf or ""
        gc.font      = Font(
            name=body_font.name, size=body_font.size,
            color=CONF_TEXT_COLORS.get(conf or "", "000000"),
        )
        gc.border    = copy(all_bdr)
        gc.alignment = Alignment(horizontal="center", vertical="center")

        # H: Source (grey italic)
        hc = ws.cell(row_idx, 8)
        hc.value     = source or ""
        hc.font      = Font(name=body_font.name, size=body_font.size,
                            italic=True, color="808080")
        hc.border    = copy(all_bdr)
        hc.alignment = Alignment(horizontal="left", vertical="center")

        # I: Notes (wrap text)
        ic = ws.cell(row_idx, 9)
        ic.value     = notes or ""
        ic.font      = copy(body_font)
        ic.border    = copy(all_bdr)
        ic.alignment = Alignment(horizontal="left", vertical="center",
                                 wrap_text=True)

    # ── TOTAL row ──────────────────────────────────────────────────────────
    last_row = ws.max_row
    while last_row > APPROVED_BOQ_DATA_START and not any(
        ws.cell(last_row, c).value for c in range(1, 5)
    ):
        last_row -= 1

    total_row = last_row + 2
    blue_fill = PatternFill("solid", fgColor="2F5496")
    wht_bold  = Font(name=body_font.name, size=body_font.size,
                     bold=True, color="FFFFFF")

    for col_idx in range(1, 10):
        tc = ws.cell(total_row, col_idx)
        tc.fill   = copy(blue_fill)
        tc.border = copy(all_bdr)
        tc.font   = copy(wht_bold)

    ws.cell(total_row, 2).value     = "TOTAL SUPPLY COST (PGK)"
    ws.cell(total_row, 2).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(total_row, 6).value     = round(grand_total, 2)
    ws.cell(total_row, 6).alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(total_row, 6).number_format = "#,##0.00"

    log.info("BOQ sheet: grand total = PGK %.2f", grand_total)

    # ── Supporting sheets ──────────────────────────────────────────────────
    _write_conflicts_sheet(wb, validation.get("conflicts", []), openpyxl)
    _write_missing_scope_sheet(wb, validation.get("missing_scope", []), openpyxl)
    _write_qa_summary_sheet(wb, validation, merged, openpyxl)

    wb.save(str(output_path))
    return round(grand_total, 2)


# ─── Item matching helpers ────────────────────────────────────────────────────

def _norm(text: str) -> str:
    t = (text or "").lower()
    t = re.sub(r"[^\w\s]", " ", t)
    t = re.sub(r"\s+", " ", t).strip()
    return t


def _find_boq_item(
    sc_key: str,
    desc_key: str,
    item_by_desc: dict,
    items_by_code: dict,
) -> dict | None:
    """Find best matching boq_item for a template row.

    Priority: exact description → fuzzy description → unique code → best desc match among dupes.
    """
    if desc_key and desc_key in item_by_desc:
        return item_by_desc[desc_key]

    fuzzy = _fuzzy_find_item(desc_key, item_by_desc)
    if fuzzy:
        return fuzzy

    if sc_key and sc_key in items_by_code:
        code_items = items_by_code[sc_key]
        if len(code_items) == 1:
            return code_items[0]
        return _best_item_desc_match(desc_key, code_items)

    return None


def _best_item_desc_match(desc_key: str, items: list[dict]) -> dict | None:
    query_words = set(desc_key.split())
    best_score, best_item = 0.0, items[0]
    for item in items:
        item_desc  = _norm(item.get("description") or "")
        item_words = set(item_desc.split())
        common     = query_words & item_words
        if not common:
            continue
        score = len(common) / max(len(query_words), len(item_words), 1)
        if score > best_score:
            best_score = score
            best_item  = item
    return best_item


def _fuzzy_find_item(norm_query: str, item_by_desc: dict) -> dict | None:
    query_words = set(norm_query.split())
    if len(query_words) < 2:
        return None
    best_score, best_item = 0.0, None
    for norm_key, item in item_by_desc.items():
        key_words = set(norm_key.split())
        common    = query_words & key_words
        if len(common) < 2:
            continue
        score = len(common) / max(len(query_words), len(key_words), 1)
        if score > best_score:
            best_score, best_item = score, item
    return best_item if best_score >= 0.35 else None


# ─── Shared note builder ──────────────────────────────────────────────────────

def _build_notes(item: dict) -> str:
    parts = []
    if item.get("issue_flag"):
        parts.append(item["issue_flag"])
    if item.get("assumption"):
        parts.append(item["assumption"])
    if item.get("comment"):
        parts.append(item["comment"])
    return " | ".join(parts)


# ─── Fallback: full fresh workbook (no approved BOQ template) ─────────────────

def _write_full_workbook(
    output_path: Path, project_name: str,
    boq_items: list[dict], validation: dict, merged: dict, openpyxl,
) -> float:
    wb = openpyxl.Workbook()
    ws_boq = wb.active
    ws_boq.title = "BOQ"
    total = _write_boq_sheet(ws_boq, boq_items, project_name, openpyxl)
    _write_conflicts_sheet(wb, validation.get("conflicts", []), openpyxl)
    _write_missing_scope_sheet(wb, validation.get("missing_scope", []), openpyxl)
    _write_qa_summary_sheet(wb, validation, merged, openpyxl)
    _write_assumptions_sheet(wb, boq_items, openpyxl)
    wb.save(str(output_path))
    return total


def _write_boq_sheet(ws, items: list[dict], project_name: str, openpyxl) -> float:
    from openpyxl.styles import Font, PatternFill, Alignment

    ws["A1"] = "BILL OF QUANTITIES"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Project: {project_name}"
    ws["A3"] = f"Date: {date.today().strftime('%d %B %Y')}"

    _write_header_row(ws, 5, BOQ_COLS, openpyxl)
    col_idx      = {h: i + 1 for i, h in enumerate(BOQ_COLS)}
    data_row     = 6
    total_amount = 0.0
    current_cat  = None

    for item in items:
        # Support both old keys (qty, source, category) and boq_mapper keys
        # (quantity, source_evidence, boq_section) — prefer old keys if present.
        cat = (item.get("category") or item.get("boq_section") or "").strip()
        if cat and cat != current_cat:
            current_cat = cat
            ws.cell(data_row, col_idx["Description"]).value = cat.upper()
            ws.cell(data_row, col_idx["Description"]).font = Font(bold=True, italic=True)
            ws.cell(data_row, col_idx["Description"]).fill = PatternFill("solid", fgColor="D9E1F2")
            data_row += 1

        # qty: prefer "qty" key, fall back to "quantity" (boq_mapper output)
        qty    = safe_float(item.get("qty") if item.get("qty") is not None else item.get("quantity"))
        rate   = safe_float(item.get("rate"))
        amount = round(qty * rate, 2) if (qty is not None and rate) else None
        if amount:
            total_amount += amount
        conf = (item.get("confidence") or "").upper()

        # source: prefer "source" key, fall back to "source_evidence" (boq_mapper output)
        source_val = item.get("source") or item.get("source_evidence") or ""
        # notes: build from item flags OR use "notes" field directly (boq_mapper)
        notes_val = _build_notes(item) or item.get("notes") or ""

        ws.cell(data_row, col_idx["Item No"]).value        = item.get("item_no") or ""
        ws.cell(data_row, col_idx["Stock Code"]).value     = item.get("stock_code") or ""
        ws.cell(data_row, col_idx["Description"]).value    = item.get("description") or ""
        ws.cell(data_row, col_idx["Unit"]).value           = item.get("unit") or ""
        ws.cell(data_row, col_idx["Qty"]).value            = qty
        ws.cell(data_row, col_idx["Rate (PGK)"]).value     = rate
        ws.cell(data_row, col_idx["Amount (PGK)"]).value   = amount
        ws.cell(data_row, col_idx["Confidence"]).value     = conf
        ws.cell(data_row, col_idx["Qty Basis"]).value      = item.get("quantity_basis") or ""
        ws.cell(data_row, col_idx["Source"]).value         = source_val
        ws.cell(data_row, col_idx["Rule / Method"]).value  = item.get("quantity_rule_used") or item.get("notes") or ""
        ws.cell(data_row, col_idx["Notes"]).value          = notes_val

        colour = CONF_BG_COLOURS.get(conf)
        if colour:
            ws.cell(data_row, col_idx["Confidence"]).fill = PatternFill("solid", fgColor=colour)

        basis_colour = QTY_BASIS_COLOURS.get(item.get("quantity_basis", ""))
        if basis_colour:
            ws.cell(data_row, col_idx["Qty Basis"]).fill = PatternFill("solid", fgColor=basis_colour)

        if item.get("issue_flag") in ("REVIEW_REQUIRED", "MISSING_DATA"):
            ws.cell(data_row, col_idx["Notes"]).font = Font(color="C00000")
        data_row += 1

    data_row += 1
    ws.cell(data_row, col_idx["Description"]).value  = "TOTAL"
    ws.cell(data_row, col_idx["Amount (PGK)"]).value = round(total_amount, 2)
    ws.cell(data_row, col_idx["Description"]).font   = Font(bold=True)
    ws.cell(data_row, col_idx["Amount (PGK)"]).font  = Font(bold=True)

    widths = {
        "Item No": 8, "Stock Code": 14, "Description": 52, "Unit": 8,
        "Qty": 10, "Rate (PGK)": 12, "Amount (PGK)": 14,
        "Confidence": 12, "Qty Basis": 14, "Source": 20, "Rule / Method": 46, "Notes": 30,
    }
    for col_name, width in widths.items():
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(col_idx[col_name])].width = width
    for col_name in ("Rate (PGK)", "Amount (PGK)"):
        for r in range(6, data_row + 1):
            ws.cell(r, col_idx[col_name]).number_format = "#,##0.00"

    log.info("BOQ sheet: %d items  total=PGK %.2f", len(items), total_amount)
    return total_amount


# ─── Supporting sheets ────────────────────────────────────────────────────────

def _write_conflicts_sheet(wb, conflicts: list, openpyxl) -> None:
    ws = wb.create_sheet("Conflicts")
    headers = ["Item", "Source A", "Value A", "Source B", "Value B",
               "Diff %", "Severity", "Recommended Action"]
    _write_header_row(ws, 1, headers, openpyxl)
    for i, c in enumerate(conflicts, 2):
        ws.cell(i, 1).value = c.get("item_name", "")
        ws.cell(i, 2).value = c.get("source_a", "")
        ws.cell(i, 3).value = c.get("value_a", "")
        ws.cell(i, 4).value = c.get("source_b", "")
        ws.cell(i, 5).value = c.get("value_b", "")
        ws.cell(i, 6).value = c.get("diff_pct", "")
        ws.cell(i, 7).value = c.get("severity", "")
        ws.cell(i, 8).value = c.get("recommended_action", "")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["H"].width = 80


def _write_missing_scope_sheet(wb, missing: list, openpyxl) -> None:
    ws = wb.create_sheet("Missing Scope")
    headers = ["Category", "Description", "Risk"]
    _write_header_row(ws, 1, headers, openpyxl)
    for i, m in enumerate(missing, 2):
        ws.cell(i, 1).value = m.get("category", "")
        ws.cell(i, 2).value = m.get("description", "")
        ws.cell(i, 3).value = m.get("risk", "")
    ws.column_dimensions["B"].width = 60


def _write_qa_summary_sheet(wb, validation: dict, merged: dict, openpyxl) -> None:
    from openpyxl.styles import Font
    ws  = wb.create_sheet("QA Summary")
    row = 1
    ws.cell(row, 1).value = "QA SUMMARY"
    ws.cell(row, 1).font  = Font(bold=True, size=14)
    row += 2

    ws.cell(row, 1).value = "RELATIONSHIP CHECKS"
    ws.cell(row, 1).font  = Font(bold=True)
    row += 1
    for chk in validation.get("relationship_checks", []):
        ws.cell(row, 1).value = chk.get("check_name", "")
        ws.cell(row, 2).value = chk.get("status", "")
        ws.cell(row, 3).value = chk.get("details", "")
        row += 1

    row += 1
    ws.cell(row, 1).value = "MISSING SCOPE"
    ws.cell(row, 1).font  = Font(bold=True)
    row += 1
    for m in validation.get("missing_scope", []):
        ws.cell(row, 1).value = f"[{m.get('risk')}]"
        ws.cell(row, 2).value = m.get("category", "")
        ws.cell(row, 3).value = m.get("description", "")
        row += 1

    row += 1
    ws.cell(row, 1).value = "OVERALL NOTES"
    ws.cell(row, 1).font  = Font(bold=True)
    row += 1
    for note in validation.get("overall_notes", []):
        ws.cell(row, 1).value = str(note)
        row += 1

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 70


def _write_assumptions_sheet(wb, items: list, openpyxl) -> None:
    ws = wb.create_sheet("Assumptions")
    headers = ["Item No", "Description", "Source", "Assumption / Comment"]
    _write_header_row(ws, 1, headers, openpyxl)
    row = 2
    for item in items:
        if item.get("assumption") or item.get("issue_flag"):
            ws.cell(row, 1).value = item.get("item_no", "")
            ws.cell(row, 2).value = item.get("description", "")
            ws.cell(row, 3).value = item.get("source", "")
            ws.cell(row, 4).value = _build_notes(item)
            row += 1
    ws.column_dimensions["B"].width = 52
    ws.column_dimensions["D"].width = 80


def _write_header_row(ws, row: int, headers: list, openpyxl) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment
    for col, h in enumerate(headers, 1):
        cell           = ws.cell(row, col)
        cell.value     = h
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = PatternFill("solid", fgColor="2F5496")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
