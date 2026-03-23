"""
boq_writer_v2.py — Write final BOQ Excel (V2 format).

Columns:
  Item No | Stock Code | Description | Unit | Qty
  Rate (PGK) | Amount (PGK)
  Qty Basis | Source | Rule/Method | Confidence | Notes

Colour coding:
  measured      → C6EFCE  (green)
  derived       → FFEB9C  (amber)
  provisional   → FFC7CE  (red / pink)
  manual_review → D9D9D9  (grey)
"""
from __future__ import annotations

import logging
from pathlib import Path

log = logging.getLogger("boq.v2.boq_writer_v2")

# Colour fills
_COLOURS = {
    "measured":    "C6EFCE",
    "derived":     "FFEB9C",
    "provisional": "FFC7CE",
    "manual_review": "D9D9D9",
}

_HEADERS = [
    "Item No", "Stock Code", "Description", "Unit", "Qty",
    "Rate (PGK)", "Amount (PGK)",
    "Qty Basis", "Source", "Rule/Method", "Confidence", "Notes",
]

_COL_WIDTHS = [8, 14, 45, 8, 10, 12, 14, 14, 22, 35, 12, 40]


def write_boq_excel(
    boq_items:  list[dict],
    output_dir: Path,
    project_name: str = "project",
) -> Path | None:
    """
    Write BOQ items to an xlsx file with V2 column layout and colour coding.

    Returns path written, or None if openpyxl not installed (stub mode).
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        log.warning("openpyxl not installed — BOQ Excel not written (stub mode)")
        return None

    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_dir / f"{project_name}_BOQ_V2.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BOQ"

    # Header style
    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill("solid", fgColor="2F5496")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_side    = Side(style="thin")
    thin_border  = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # Write headers
    for col_idx, (header, width) in enumerate(zip(_HEADERS, _COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font   = header_font
        cell.fill   = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22

    # Group items by section for section headers
    current_section = None
    row_idx = 2

    for item in boq_items:
        section = item.get("boq_section", "")

        # Write section header row
        if section != current_section:
            current_section = section
            ws.merge_cells(
                start_row=row_idx, start_column=1,
                end_row=row_idx, end_column=len(_HEADERS),
            )
            sec_cell = ws.cell(row=row_idx, column=1, value=section)
            sec_cell.font      = Font(bold=True, color="FFFFFF")
            sec_cell.fill      = PatternFill("solid", fgColor="4472C4")
            sec_cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[row_idx].height = 18
            row_idx += 1

        basis = item.get("quantity_basis", "provisional")
        is_manual = item.get("manual_review", False)
        colour_key = "manual_review" if is_manual else basis
        fill_colour = _COLOURS.get(colour_key, "FFFFFF")
        row_fill = PatternFill("solid", fgColor=fill_colour)

        values = [
            item.get("item_no"),
            item.get("stock_code", ""),
            item.get("description", ""),
            item.get("unit", ""),
            item.get("quantity"),
            item.get("rate_pgk"),
            item.get("amount_pgk"),
            basis,
            item.get("v2_extractor_source", ""),
            item.get("quantity_rule_used", ""),
            item.get("confidence", ""),
            item.get("notes", ""),
        ]

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill      = row_fill
            cell.border    = thin_border
            cell.alignment = Alignment(
                horizontal="left" if col_idx > 2 else "center",
                vertical="center",
                wrap_text=(col_idx in (3, 10, 12)),
            )

        ws.row_dimensions[row_idx].height = 15
        row_idx += 1

    # Freeze top row
    ws.freeze_panes = "A2"

    wb.save(str(out_path))
    log.info("BOQ Excel written → %s  (%d items)", out_path, len(boq_items))
    return out_path
