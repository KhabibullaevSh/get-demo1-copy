"""
item_library.py — Load the approved BOQ as a REFERENCE for stock codes / descriptions only.

CRITICAL: This module is REFERENCE ONLY.
  - Stock codes, descriptions, and units may be looked up here.
  - Quantities from the library are NEVER used.
  - Any data that enters the quantity model must come from DXF / IFC / PDF / BOM.
"""
from __future__ import annotations

import logging
from pathlib import Path

log = logging.getLogger("boq.v2.item_library")


def load_item_library(data_dir: Path) -> dict:
    """
    Load approved_boq.xlsx from *data_dir* as a stock-code / description reference.

    Returns:
        {
          "items": [
              {
                "stock_code":  str,
                "description": str,
                "unit":        str,
                "section":     str,
                "_source":     "reference_only — no quantities",
              }, ...
          ],
          "loaded": bool,
          "source_file": str | None,
          "warning": str,
        }

    NEVER returns any quantity values from the library.
    """
    candidates = [
        data_dir / "approved_boq.xlsx",
        data_dir / "SDP-3Bedroom_BOQ_Final_Format.xlsx",
    ]

    for boq_path in candidates:
        if boq_path.exists():
            return _load_xlsx(boq_path)

    log.warning("No approved BOQ found in %s — item library empty", data_dir)
    return {
        "items":       [],
        "loaded":      False,
        "source_file": None,
        "warning":     f"No approved BOQ xlsx found in {data_dir}",
    }


def _load_xlsx(boq_path: Path) -> dict:
    try:
        import openpyxl
    except ImportError:
        return {
            "items": [], "loaded": False,
            "source_file": str(boq_path),
            "warning": "openpyxl not installed — item library not loaded",
        }

    items: list[dict] = []
    try:
        wb = openpyxl.load_workbook(str(boq_path), data_only=True, read_only=True)
        ws = wb.active

        header_row_idx: int | None = None
        headers: dict[str, int]    = {}

        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            row_lower = [str(c).lower().strip() if c is not None else "" for c in row]
            if any(h in row_lower for h in ("stock code", "description", "item", "code")):
                header_row_idx = row_idx
                headers = {v: i for i, v in enumerate(row_lower) if v}
                break

        if header_row_idx is None:
            wb.close()
            return {
                "items": [], "loaded": False,
                "source_file": str(boq_path),
                "warning": "No header row found in approved BOQ",
            }

        code_col  = next((headers[k] for k in headers if "code" in k), None)
        desc_col  = next((headers[k] for k in headers if "desc" in k), None)
        unit_col  = next((headers[k] for k in headers if k == "unit"), None)
        sect_col  = next((headers[k] for k in headers if "section" in k or "group" in k), None)

        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            if all(c is None for c in row):
                continue
            code = str(row[code_col]).strip() if code_col is not None and row[code_col] is not None else ""
            desc = str(row[desc_col]).strip() if desc_col is not None and row[desc_col] is not None else ""
            unit = str(row[unit_col]).strip() if unit_col is not None and row[unit_col] is not None else ""
            sect = str(row[sect_col]).strip() if sect_col is not None and row[sect_col] is not None else ""
            if not code and not desc:
                continue
            items.append({
                "stock_code":  code,
                "description": desc,
                "unit":        unit,
                "section":     sect,
                "_source":     "reference_only — no quantities",
            })

        wb.close()
        log.info("Item library loaded: %d reference items from %s", len(items), boq_path.name)
        return {
            "items":       items,
            "loaded":      True,
            "source_file": str(boq_path),
            "warning":     "REFERENCE ONLY — stock codes and descriptions, no quantities",
        }

    except Exception as exc:
        log.error("Failed to load item library from %s: %s", boq_path, exc)
        return {
            "items": [], "loaded": False,
            "source_file": str(boq_path),
            "warning": f"Error loading item library: {exc}",
        }


def find_stock_code(item_library: dict, description_keywords: list[str]) -> str | None:
    """
    Look up a stock code from the reference library by description keywords.
    Returns the first matching stock_code, or None if not found.
    Never returns a quantity.
    """
    for item in item_library.get("items", []):
        desc_lower = item.get("description", "").lower()
        if all(kw.lower() in desc_lower for kw in description_keywords):
            return item.get("stock_code") or None
    return None
